from typing import List, Optional
from datetime import datetime
import uuid

import openpyxl
from openpyxl.utils import get_column_letter

from backend.Sub.models import SubTree, SubNode, TreeMeta

import re

def parse_qty(value) -> float:
    if value is None:
        return 1

    s = str(value).upper().replace(" ", "")
    # 숫자만 추출 (정수 / 소수)
    m = re.search(r"(\d+(\.\d+)?)", s)
    if not m:
        return 1

    try:
        return float(m.group(1))
    except Exception:
        return 1


def get_cell_value(ws, r, c):
    coord = f"{get_column_letter(c)}{r}"

    for merged in ws.merged_cells.ranges:
        if coord in merged:
            return ws.cell(
                merged.min_row,
                merged.min_col
            ).value

    return ws.cell(row=r, column=c).value

def read_right_value(ws, r, c):
    """
    라벨 셀(c) 기준, 바로 오른쪽부터
    최대 4칸까지만 값을 탐색한다.
    """
    start_col = c + 1
    max_scan_col = c + 4

    visited_merged = set()
    col = start_col

    while col <= max_scan_col:
        coord = f"{get_column_letter(col)}{r}"
        value = ws.cell(row=r, column=col).value

        for merged in ws.merged_cells.ranges:
            if coord in merged:
                key = (merged.min_row, merged.min_col)
                if key in visited_merged:
                    value = None
                else:
                    visited_merged.add(key)
                    value = ws.cell(
                        merged.min_row,
                        merged.min_col
                    ).value
                break

        if value not in [None, "", "부품명", "품번", "수량", "재질"]:
            return str(value).strip()

        col += 1

    return None

def find_row_with_label(ws, start_r, start_c, label):
    for offset in range(1, 10):
        row = start_r + offset
        col = start_c
        cell = ws.cell(row=row, column=col)

        if cell.value:
            s = str(cell.value).replace(" ", "").replace("\n", "").strip()
            if label in s:
                return offset

        coord = f"{get_column_letter(col)}{row}"
        for merged in ws.merged_cells.ranges:
            if coord in merged:
                top_left = ws.cell(merged.min_row, merged.min_col)
                if top_left.value:
                    s2 = str(top_left.value).replace(" ", "").replace("\n", "").strip()
                    if label in s2:
                        return offset

    raise ValueError(f"Label '{label}' not found below row {start_r}")

def make_stable_id(spec_name: str, row: int, col: int) -> str:
    return f"{spec_name}:{row}:{col}"

def parse_block(ws, spec_name: str, label_row: int, label_col: int):
    r = label_row
    c = label_col

    name = read_right_value(ws, r, c)

    try:
        part_row = r + find_row_with_label(ws, r, c, "품번")
        part_no = read_right_value(ws, part_row, c)
    except Exception:
        part_no = None

    try:
        qty_row = r + find_row_with_label(ws, r, c, "수량")
        raw_qty = read_right_value(ws, qty_row, c)
    except Exception:
        raw_qty = None

    try:
        mat_row = r + find_row_with_label(ws, r, c, "재질")
        material = read_right_value(ws, mat_row, c)
    except Exception:
        material = None

    return {
        "id": name,
        "name": make_stable_id(spec_name, r, c),
        "part_no": part_no,
        "qty": raw_qty,
        "material": material,
        "row": r,
        "col": c,
    }

def build_tree_from_sheet(
    ws,
    bom_id: str,
    bom_filename: str,
    spec_name: str
) -> SubTree:

    boxes = []

    # 1. "부품명" 기준 block 수집
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        max_col=ws.max_column
    ):
        for cell in row:
            if cell.value and str(cell.value).strip() == "부품명":
                box = parse_block(ws, spec_name, cell.row, cell.column)
                boxes.append(box)

    # 2. 위치 기준 정렬 (위→아래, 좌→우)
    boxes_sorted = sorted(boxes, key=lambda b: (b["row"], b["col"]))

    nodes: List[SubNode] = []

    # ⭐ col 기준 부모 추적 스택 ( parent_ui_name , col )
    stack: List[tuple[str, int]] = []

    # 같은 부모 내 order 관리
    order_counter: dict[Optional[str], int] = {}

    for box in boxes_sorted:

        # 이 노드의 UI name ( spec:row:col )
        ui_name = f"{spec_name}:{box['row']}:{box['col']}"

        # 현재 box보다 col이 크거나 같은 애들은 부모가 될 수 없음
        while stack and box["col"] <= stack[-1][1]:
            stack.pop()

        # ⭐ 부모 ui_name 얻기
        parent_ui_name = stack[-1][0] if stack else None

        order = order_counter.get(parent_ui_name, 0)
        order_counter[parent_ui_name] = order + 1

        node = SubNode(
            id=box["id"],                  # 내부 DB PK (유지)
            name=ui_name,                  # UI name
            parent_name=parent_ui_name,    # ⭐ 부모 UI name
            order=order,
            type="PART",
            part_no=box["part_no"],
            material=box["material"],
            qty=parse_qty(box.get("qty")),
            inhouse=False,
        )

        nodes.append(node)

        # ⭐ 현재 노드를 스택에 push
        stack.append((ui_name, box["col"]))

    meta = TreeMeta(
        bom_id=bom_id,
        bom_filename=bom_filename,
        spec_name=spec_name,
    )

    return SubTree(
        meta=meta,
        nodes=nodes
    )

