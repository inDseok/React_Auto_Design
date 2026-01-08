import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import os
import win32com.client


# ================================================================
# 0. 설정
# ================================================================
# 원본파일 = r"C:\Users\USER\Desktop\test\1215\CN8_BOM_TEST.xlsx"
원본파일 = "..."
결과파일 = "..."
선택사양 = "..."




# ================================================================
# 0. 도식화 핵심 변수 
# ================================================================
전체박스_가로 = 12
전체박스_세로 = 8

레벨_이동칸 = 16        # 하위 레벨일수록 이동하는 간격 (가로)
박스_높이 = 10          # 다음 박스 시작행 (세로)

도식화_시작행 = 8      # 트리 도식화 첫 행
도식화_시작열 = 3      # 트리 시작 열
셀너비 = 4.2           # 시트 전체 열 너비


# -------------------------------------------------------
# 0. 공통 스타일 정의
# -------------------------------------------------------

# 두꺼운 선 + 얇은 선
두꺼운선 = Side(border_style="thick", color="000000")
얇은선 = Side(border_style="thin", color="000000")

# 테두리(얇은 테두리)
테두리 = Border(left=얇은선, right=얇은선, top=얇은선, bottom=얇은선)

# ★ 두꺼운 테두리(박스 외곽용)
두꺼운테두리 = Border(left=두꺼운선, right=두꺼운선,
                    top=두꺼운선, bottom=두꺼운선)

# 가운데 정렬
가운데정렬 = Alignment(horizontal="center", vertical="center")

#채우기
회색채움 = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


# ================================================================
# 0-1. xls / xlsb / xlsm → xlsx 변환
# ================================================================
def 변환_xls_to_xlsx_안전(원본파일):

    if not 원본파일.lower().endswith((".xlsb", ".xls", ".xlsm")):

        return 원본파일


    변환파일 = 원본파일.rsplit(".", 1)[0] + "_converted.xlsx"

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.EnableEvents = False

    try:
        wb = excel.Workbooks.Open(
            원본파일,
            ReadOnly=True,
            UpdateLinks=0
        )

        try:
            excel.Calculation = -4135   # xlCalculationManual
        except Exception as e:
            print("   [경고] Calculation 설정 실패 (무시):", e)

        wb.SaveAs(변환파일, FileFormat=51)
        wb.Close(False)


    finally:
        excel.Quit()

    return 변환파일

# ================================================================
# 1. openpyxl 로드
# ================================================================
def 엑셀_로드_기본정보(xlsx파일):

    wb = openpyxl.load_workbook(xlsx파일, data_only=True)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column


    return wb, ws, max_row, max_col


# ================================================================
# 2. 셀 접근 헬퍼 (시트 의존 명확화)
# ================================================================
def 셀값(ws, r, c):
    return ws.cell(r, c).value


# ================================================================
# 2. '수량' 병합셀 탐색 (함수화)
# ================================================================
def 수량_병합셀_탐색(ws):

    수량행 = None
    수량_왼쪽열 = None
    수량_오른열 = None

    # --------------------------------------------------
    # 1️⃣ '수량' 병합셀 찾기 (기존 기능)
    # --------------------------------------------------
    for 병합 in ws.merged_cells.ranges:
        값 = ws.cell(병합.min_row, 병합.min_col).value
        if 값 and str(값).strip() == "수량":
            수량행 = 병합.min_row
            수량_왼쪽열 = 병합.min_col
            수량_오른열 = 병합.max_col
            break

    if 수량행 is None:
        raise Exception("❌ '수량' 병합셀을 찾지 못함")

    # --------------------------------------------------
    # 2️⃣ 수량 병합셀 기준 아래로 내려가며
    #    병합셀이 없는 첫 행 찾기
    # --------------------------------------------------
    병합없는_첫행 = None

    for r in range(수량행 + 1, ws.max_row + 1):
        병합존재 = False

        for 병합 in ws.merged_cells.ranges:
            # 행이 병합 범위 안에 있고,
            # 열 범위가 수량 열 범위와 겹치면 병합 존재
            if (
                병합.min_row <= r <= 병합.max_row and
                not (병합.max_col < 수량_왼쪽열 or 병합.min_col > 수량_오른열)
            ):
                병합존재 = True
                break

        if not 병합존재:
            병합없는_첫행 = r
            break

    if 병합없는_첫행 is None:
        raise Exception("❌ 병합이 끝나는 데이터 시작 행을 찾지 못함")

    # --------------------------------------------------
    # 3️⃣ 헤더 깊이 계산
    # --------------------------------------------------
    헤더깊이 = 병합없는_첫행 - 수량행 - 1 

    return 수량행, 수량_왼쪽열, 수량_오른열, 헤더깊이

# ================================================================
# 3. 수량 병합 해제 + 값 채우기 + 4. 사양 컬럼 비활성화
# ================================================================
def 수량병합해제_및_사양컬럼_비활성화(ws, 수량행, 수량_왼쪽열, 수량_오른열):

    # ------------------------------------------------------------
    # [3] 수량 병합 해제 + 값 채우기
    # ------------------------------------------------------------

    for 병합 in ws.merged_cells.ranges.copy():
        if 병합.min_col >= 수량_왼쪽열 and 병합.max_col <= 수량_오른열:
            값 = ws.cell(병합.min_row, 병합.min_col).value
            ws.unmerge_cells(str(병합))
            for r in range(병합.min_row, 병합.max_row + 1):
                for c in range(병합.min_col, 병합.max_col + 1):
                    ws.cell(r, c).value = 값

    # ------------------------------------------------------------
    # [4] 짝수 사양 컬럼 비활성화 마스크 생성
    # ------------------------------------------------------------

    사양컬럼_활성화 = {}

    for col in range(수량_왼쪽열, 수량_오른열 + 1):
        offset = col - 수량_왼쪽열
        사양컬럼_활성화[col] = (offset % 2 == 0)

    for col, active in 사양컬럼_활성화.items():
        상태 = "사용" if active else "비활성"

    return 수량_왼쪽열, 수량_오른열, 사양컬럼_활성화


# ================================================================
# 5. 사양명 생성 (비활성 컬럼 제외)
# ================================================================
def 사양명_생성(
    ws,
    시트명,
    수량행,
    수량_왼쪽열,
    수량_오른열,
    헤더깊이,
    사양컬럼_활성화
):

    사양목록 = []

    for col in range(수량_왼쪽열, 수량_오른열 + 1):

        # 🔴 비활성 컬럼 스킵
        if not 사양컬럼_활성화.get(col, True):
            continue

        토큰 = []

        for d in range(1, 헤더깊이 + 1):
            값 = ws.cell(수량행 + d, col).value
            if 값:
                clean = (
                    str(값)
                    .replace("/", "")
                    .replace(" ", "_")
                    .replace("'", "")
                )
                토큰.append(clean)

        사양명 = f"{시트명}_{'_'.join(토큰)}"
        사양목록.append((col, 사양명))

    return 사양목록





# ================================================================
# 6. '조립단위' 병합셀 탐색 (함수화)
# ================================================================
def 조립단위_병합셀_탐색(ws):


    조립단위행 = None
    조립단위_왼쪽열 = None
    조립단위_오른열 = None

    for 병합 in ws.merged_cells.ranges:
        값 = ws.cell(병합.min_row, 병합.min_col).value
        if 값 and str(값).strip() == "조립단위":
            조립단위행 = 병합.min_row
            조립단위_왼쪽열 = 병합.min_col
            조립단위_오른열 = 병합.max_col
            break

    if 조립단위행 is None:
        raise Exception("❌ '조립단위' 병합셀 못찾음")

    return 조립단위행, 조립단위_왼쪽열, 조립단위_오른열


# ================================================================
# 7. '*' 기반 데이터 시작/끝 행 탐색 (공백 허용 버전)
# ================================================================
def 데이터구간_탐색_스타기반(
    ws,
    조립단위행,
    조립단위_왼쪽열,
    조립단위_오른열,
    최대행,
    허용_공백_연속수
):

    데이터시작행 = None
    데이터끝행 = None

    연속_공백수 = 0
    데이터_진행중 = False

    for r in range(조립단위행 + 1, 최대행 + 1):

        # -------------------------------------------
        # 현재 행에 '*' 존재 여부
        # -------------------------------------------
        줄에_스타존재 = False
        for c in range(조립단위_왼쪽열, 조립단위_오른열 + 1):
            값 = ws.cell(r, c).value
            if isinstance(값, str) and 값.strip().startswith("*"):
                줄에_스타존재 = True
                break

        # -------------------------------------------
        # 1) 아직 시작 안 했고 '*' 발견 → 시작
        # -------------------------------------------
        if 데이터시작행 is None and 줄에_스타존재:
            데이터시작행 = r
            데이터_진행중 = True
            연속_공백수 = 0
            continue

        # -------------------------------------------
        # 2) 데이터 진행 중
        # -------------------------------------------
        if 데이터_진행중:

            if 줄에_스타존재:
                # '*' 다시 등장 → 공백 카운트 리셋
                연속_공백수 = 0
            else:
                연속_공백수 += 1

                # ❌ 허용 초과 → 데이터 종료
                if 연속_공백수 >= 허용_공백_연속수:
                    데이터끝행 = r - 연속_공백수
                    break

    # -------------------------------------------
    # 파일 끝까지 갔는데 종료 못한 경우
    # -------------------------------------------
    if 데이터시작행 and 데이터끝행 is None:
        데이터끝행 = 최대행

    if 데이터시작행 is None:
        raise Exception("❌ '*' 기반 데이터 시작행 없음")

    return 데이터시작행, 데이터끝행




# ================================================================
# 8. '품명' 열 탐색 (함수화)
# ================================================================
def 품명열_탐색(ws, 최대행, 최대열):

    품명열 = None
    품명행 = None

    for r in range(1, 최대행 + 1):
        for c in range(1, 최대열 + 1):
            값 = ws.cell(r, c).value
            if 값 is None:
                continue
            if str(값).strip() == "품명":
                품명행 = r
                품명열 = c
                break
        if 품명열:
            break

    if 품명열 is None:
        raise Exception("❌ '품명' 셀을 시트 전체에서 찾을 수 없음")

    return 품명열



# ================================================================
# 8-2. '품번' 열 탐색 (함수화)
# ================================================================
def 품번열_탐색(ws, 최대행, 최대열):

    품번열 = None
    품번행 = None

    for r in range(1, 최대행 + 1):
        for c in range(1, 최대열 + 1):
            값 = ws.cell(r, c).value
            if 값 is None:
                continue
            if str(값).strip() == "품번":
                품번행 = r
                품번열 = c
                break
        if 품번열:
            break

    if 품번열 is None:
        raise Exception("❌ '품번' 셀을 시트 전체에서 찾을 수 없음")

    return 품번열


# ================================================================
# 8-3. '재질' 열 탐색 (단일셀 → 병합셀 검증 구조)
# ================================================================
def 재질열_탐색(ws, 최대행, 최대열):

    # ------------------------------------------------------------
    # 1️⃣ 단일셀 기준 '재질' 열 후보 수집
    # ------------------------------------------------------------
    단일셀_재질열 = set()

    for r in range(1, 최대행 + 1):
        for c in range(1, 최대열 + 1):
            값 = ws.cell(r, c).value
            if 값 and str(값).strip() == "재질":
                단일셀_재질열.add(c)

    if len(단일셀_재질열) == 0:
        raise Exception("❌ 단일셀 기준 '재질' 열을 찾지 못함")

    # ------------------------------------------------------------
    # 2️⃣ 병합셀 기준 검증 (단일셀 후보 열만 검사)
    # ------------------------------------------------------------
    병합셀_유효열 = set()

    for m in ws.merged_cells.ranges:

        # ❌ 다중 컬럼 병합 제외
        if m.min_col != m.max_col:
            continue

        # 👉 단일셀 후보 열만 검사
        if m.min_col not in 단일셀_재질열:
            continue

        # 병합 영역 내 '재질' 존재 여부
        for r in range(m.min_row, m.max_row + 1):
            값 = ws.cell(r, m.min_col).value
            if 값 and str(값).strip() == "재질":
                병합셀_유효열.add(m.min_col)
                break
    # ------------------------------------------------------------
    # 3️⃣ 최종 재질열 판정
    # ------------------------------------------------------------
    if len(병합셀_유효열) == 1:
        재질열 = next(iter(병합셀_유효열))
    elif len(병합셀_유효열) == 0:
        if len(단일셀_재질열) == 1:
            재질열 = next(iter(단일셀_재질열))
        else:
            raise Exception(
                f"❌ 단일셀 기준 '재질' 열이 2개 이상이며 "
                f"병합셀로도 단일 열로 좁혀지지 않음: {sorted(단일셀_재질열)}"
            )
    else:
        raise Exception(
            f"❌ 병합셀 기준 '재질' 열이 2개 이상 감지됨: {sorted(병합셀_유효열)}"
        )

    return 재질열


# ================================================================
# 9. 전체 조립 노드 파싱 (함수화)
# ================================================================
def 조립노드_파싱(
    ws,
    데이터시작행,
    최대행,
    조립단위_왼쪽열,
    조립단위_오른열,
    품명열,
    품번열,
    재질열
):

    노드목록 = []

    # --------------------------------------------------
    # 노드 파싱
    # --------------------------------------------------
    for r in range(데이터시작행, 최대행 + 1):

        # '*' 탐색
        star_col = None
        for c in range(조립단위_왼쪽열, 조립단위_오른열 + 1):
            값 = ws.cell(r, c).value
            if isinstance(값, str) and 값.strip().startswith("*"):
                star_col = c
                break

        if star_col is None:
            continue

        # 레벨
        레벨 = star_col - 조립단위_왼쪽열

        # 품명
        이름 = ws.cell(r, 품명열).value
        if not 이름:
            continue

        # 품번
        품번값 = ws.cell(r, 품번열).value
        품번 = str(품번값).strip() if 품번값 else ""

        #재질 
        재질값 = ws.cell(r, 재질열).value
        재질 = str(재질값).strip() if 재질값 else ""

        # 노드 추가
        노드목록.append({
            "행": r,
            "레벨": 레벨,
            "이름": str(이름).strip(),
            "품번": 품번,
            "재질": 재질
        })

    return 노드목록


# ================================================================
# 10. 사양별 수량 매핑 (재질 포함)
# ================================================================
def 사양별_수량_매핑(ws, 사양목록, 노드목록):

    사양별_수량 = {사양명: {} for (_, 사양명) in 사양목록}

    for 열, 사양명 in 사양목록:
        for 노드 in 노드목록:
            행 = 노드["행"]
            이름 = 노드["이름"]
            레벨 = 노드["레벨"]
            품번 = 노드["품번"]
            재질 = 노드["재질"]    # ✅ 추가

            raw = ws.cell(행, 열).value

            # ------------------------------
            # 빈 값 제외
            # ------------------------------
            if raw in (None, "", " ", 0):
                continue

            # ------------------------------
            # 문자열 숫자 처리
            # ------------------------------
            if isinstance(raw, str):
                if raw.isdigit():
                    수량 = int(raw)
                else:
                    continue
            elif isinstance(raw, (int, float)):
                수량 = int(raw)
            else:
                continue

            if 수량 <= 0:
                continue

            # ------------------------------
            # 매핑 저장
            # ------------------------------
            사양별_수량[사양명][행] = {
                "품명": 이름,
                "품번": 품번,
                "레벨": 레벨,
                "수량": 수량,
                "재질": 재질      # ✅ 추가
            }


    return 사양별_수량


# ================================================================
# 10-1. 사양별 수량 매핑 보기용 구조 변환 (재질 포함)
# ================================================================
def 사양별_수량_보기용_변환(사양별_수량):

    사양별_수량_보기용 = {}

    for 사양명, 항목들 in 사양별_수량.items():
        보기리스트 = []

        for 행번호, 데이터 in 항목들.items():
            보기리스트.append({
                "행번호": 행번호,
                "품명": 데이터["품명"],
                "품번": 데이터["품번"],
                "레벨": 데이터["레벨"],
                "수량": 데이터["수량"],
                "재질": 데이터["재질"]     # ✅ 추가
            })

        # ⭐ BOM 트리 구조 유지를 위해 행번호 오름차순 정렬
        보기리스트.sort(key=lambda x: x["행번호"])

        사양별_수량_보기용[사양명] = 보기리스트

    return 사양별_수량_보기용


# ================================================================
# 11-1. 단일 사양 BOM 트리 생성 (재질 포함)
# ================================================================
def 생성_BOM_트리(보기리스트):

    트리 = []
    스택 = []

    for 항목 in 보기리스트:
        노드 = {
            "행번호": 항목["행번호"],
            "품명": 항목["품명"],
            "품번": 항목["품번"],
            "레벨": 항목["레벨"],
            "수량": 항목["수량"],
            "재질": 항목["재질"],   # ✅ 추가
            "자식": []
        }

        level = 항목["레벨"]

        # ✅ 방어 로직 1: 스택이 비어 있으면 무조건 루트 처리
        if level == 0 or not 스택:
            노드["레벨"] = 0
            트리.append(노드)
            스택 = [노드]
            continue

        # ✅ 방어 로직 2: 레벨 점프 방지
        if level > len(스택):
            level = len(스택)
            노드["레벨"] = level

        # 부모 레벨로 스택 정리
        while len(스택) > level:
            스택.pop()

        부모 = 스택[-1]
        부모["자식"].append(노드)

        # 스택 갱신
        if len(스택) == level:
            스택.append(노드)
        else:
            스택[level] = 노드

    return 트리



# ================================================================
# 11-2. 사양별 BOM 트리 생성 (함수화)
# ================================================================
def 사양별_BOM_트리_생성(사양별_수량_보기용):

    사양별_트리 = {}

    for 사양명, 보기리스트 in 사양별_수량_보기용.items():
        사양별_트리[사양명] = 생성_BOM_트리(보기리스트)

    return 사양별_트리


# ================================================================
# 12. 결과 파일 생성 (도식화 용) - 함수화
# ================================================================
def 결과파일_초기화():

    wb = openpyxl.Workbook()

    # 기본 sheet 제거
    기본시트 = wb.active
    wb.remove(기본시트)

    return wb




# -------------------------------------------------------
# 📌 양식 박스를 그리는 함수
# -------------------------------------------------------
def 양식박스_그리기(ws, 시작행, 시작열, 품명, 품번, 수량, 재질):

    r = 시작행
    c = 시작열

    # ------------------------------------------
    # 0) 왼쪽 이미지 큰 박스 (8행, 6칸 병합) 
    # ------------------------------------------

    m_start_row = r
    m_start_column = c
    m_end_row = r + 7
    m_end_column = c + 5   

    #병합
    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
                end_row=m_end_row, end_column=m_end_column)


    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리

    # ------------------------------------------
    # 1-1) 오른쪽 상단 - '부품명'
    # ------------------------------------------
    m_start_row = r
    m_start_column = c + 6
    m_end_row = r + 1
    m_end_column = c + 7

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
                end_row=m_end_row, end_column=m_end_column)

    cell = ws.cell(m_start_row, m_start_column)
    cell.value = "부품명"
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리
            ws.cell(rr, cc).fill = 회색채움  
    
    # ------------------------------------------
    # 1-2) 오른쪽 상단 - 부품명
    # ------------------------------------------
    m_start_row = r
    m_start_column = c + 8
    m_end_row = r + 1
    m_end_column = c + 11

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
                end_row=m_end_row, end_column=m_end_column)

    cell = ws.cell(m_start_row, m_start_column)
    cell.value = 품명
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리  

    # ------------------------------------------
    # 2-1) 품번
    # ------------------------------------------
    m_start_row = r+2
    m_start_column = c + 6
    m_end_row = r + 3
    m_end_column = c + 7

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
                end_row=m_end_row, end_column=m_end_column)

    cell = ws.cell(m_start_row, m_start_column)
    cell.value = "품번"
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리
            ws.cell(rr, cc).fill = 회색채움  

    # ------------------------------------------
    # 2-2) 품번값
    # ------------------------------------------
    m_start_row = r+2
    m_start_column = c + 8
    m_end_row = r + 3
    m_end_column = c + 11

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
            end_row=m_end_row, end_column=m_end_column)
    
    cell = ws.cell(m_start_row, m_start_column)
    cell.value = f"{품번}"
    cell.number_format = "@"       # ← ★ 텍스트로 인식되도록 지정
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리  



    # ------------------------------------------
    # 3-1) 재질
    # ------------------------------------------
    m_start_row = r+4
    m_start_column = c + 6
    m_end_row = r + 5
    m_end_column = c + 7

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
                end_row=m_end_row, end_column=m_end_column)

    cell = ws.cell(m_start_row, m_start_column)
    cell.value = "재질"
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리
            ws.cell(rr, cc).fill = 회색채움  

    # ------------------------------------------
    # 3-2) 재질값
    # ------------------------------------------
    m_start_row = r+4
    m_start_column = c + 8
    m_end_row = r + 5
    m_end_column = c + 11

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
            end_row=m_end_row, end_column=m_end_column)
    
    cell = ws.cell(m_start_row, m_start_column)
    cell.value = f"{재질}"
    cell.number_format = "@"       # ← ★ 텍스트로 인식되도록 지정
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리  





    # ------------------------------------------
    # 4-1 ) '수량'입력 (text)
    # ------------------------------------------
    m_start_row = r+6
    m_start_column = c + 6
    m_end_row = r + 7
    m_end_column = c + 7

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
            end_row=m_end_row, end_column=m_end_column)
    
    cell = ws.cell(m_start_row, m_start_column)
    cell.value = '수량'
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리
            ws.cell(rr, cc).fill = 회색채움  

    # ------------------------------------------
    # 4-2 ) 수량 입력 (value)
    # ------------------------------------------
    m_start_row = r + 6
    m_start_column = c + 8
    m_end_row = r + 7
    m_end_column = c + 10

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
                end_row=m_end_row, end_column=m_end_column)

    cell = ws.cell(m_start_row, m_start_column)
    cell.value = f"{수량}"
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리  



    # ------------------------------------------
    # 4-3 ) EA
    # ------------------------------------------
    # EA
    m_start_row = r + 6
    m_start_column = c + 11
    m_end_row = r + 7
    m_end_column = c + 11

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
                end_row=m_end_row, end_column=m_end_column)

    cell = ws.cell(m_start_row, m_start_column)
    cell.value = "EA"
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리  



    # ------------------------------------------
    # 5) 전체 박스 외곽 테두리 (4 LINE 최적화)                  ############박스_세로, 박스_간격도 함께 변경#######################
    # ------------------------------------------
    m_start_row = r 
    m_start_column = c 
    m_end_row = r + 전체박스_세로 - 1      # 총 8행 높이
    m_end_column = c + 전체박스_가로 - 1   # 총 12열 폭

    # 5-1) ★ 위쪽 라인
    for cc in range(m_start_column, m_end_column + 1):
        cell = ws.cell(m_start_row, cc)
        cell.border = Border(
            top=두꺼운선,
            left=cell.border.left,
            right=cell.border.right,
            bottom=cell.border.bottom
        )

    # 5-2) ★ 아래쪽 라인
    for cc in range(m_start_column, m_end_column + 1):
        cell = ws.cell(m_end_row, cc)
        cell.border = Border(
            bottom=두꺼운선,
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top
        )

    # 5-3) ★ 왼쪽 라인
    for rr in range(m_start_row, m_end_row + 1):
        cell = ws.cell(rr, m_start_column)
        cell.border = Border(
            left=두꺼운선,
            top=cell.border.top,
            bottom=cell.border.bottom,
            right=cell.border.right
        )

    # 5-4) ★ 오른쪽 라인
    for rr in range(m_start_row, m_end_row + 1):
        cell = ws.cell(rr, m_end_column)
        cell.border = Border(
            right=두꺼운선,
            top=cell.border.top,
            bottom=cell.border.bottom,
            left=cell.border.left
        )



    # 박스 하나의 높이는 총 8행 공간 사용 
    return r # 다음 박스가 시작될 행 반환



# -------------------------------------------------------
# 📌 트리 구조를 도식화하는 함수
# -------------------------------------------------------
def 트리_도식화(ws, 트리, 시작행, 시작열):     

    def 출력_재귀(노드, 행, 이전형제레벨):
        # -------------------------------------------------
        # 1. 현재 노드의 기본 정보 추출
        # -------------------------------------------------
        현재레벨 = 노드["레벨"]     # 트리 깊이 (0: 루트, 1: 하위, ...)
        품명 = 노드["품명"]         # 부품명
        수량 = 노드["수량"]         # 수량
        재질 = 노드["재질"]         # 재질


        if 이전형제레벨 is None:
            # 첫 번째 형제이거나, 부모 → 첫 자식인 경우
            출력행 = 행

        elif 현재레벨 == 이전형제레벨:
            # 동일 레벨의 형제 노드 → 겹치지 않도록 아래로 이동
            출력행 = 행 + 박스_높이

        else:
            # 레벨이 다른 형제 (이론상 거의 없음)
            출력행 = 행

        # -------------------------------------------------
        # 3. 출력 열(column) 위치 계산
        # -------------------------------------------------
        # 레벨에 따라 오른쪽으로 이동
        박스열 = 시작열 + (현재레벨 * 레벨_이동칸)

        # -------------------------------------------------
        # 4. 실제 엑셀 박스 출력
        # -------------------------------------------------

        다음행 = 양식박스_그리기(
            ws,
            시작행=출력행,
            시작열=박스열,
            품명=품명,
            품번=노드["품번"],
            수량=수량,
            재질=재질
        )

        # -------------------------------------------------
        # 5. 현재 노드의 박스 좌표 저장
        # -------------------------------------------------
        # 이후 연결선(부모-자식) 그리기 등에 사용
        노드["좌표"] = {
            "start_row": 출력행,
            "start_col": 박스열,
            "end_row": 출력행 + 7,
            "end_col": 박스열 + 11
        }

        # -------------------------------------------------
        # 6. 자식 노드 재귀 출력           
        # -------------------------------------------------

        이전형제 = None

        for 자식 in 노드["자식"]:
            다음행 = 출력_재귀(자식, 다음행, 이전형제)
            # 현재 자식을 다음 자식의 '이전 형제'로 설정
            이전형제 = 자식["레벨"]

        # 이 서브트리에서 마지막으로 사용된 행 반환
        return 다음행

    # -------------------------------------------------
    # 7. 루트 노드 출력 제어            
    # -------------------------------------------------

    이전루트레벨 = None
    현재행 = 시작행

    for 루트노드 in 트리:               
        현재행 = 출력_재귀(루트노드, 현재행, 이전루트레벨)
        # 다음 루트 출력을 위해 레벨 저장
        이전루트레벨 = 루트노드["레벨"]




# ================================================================
# 13. 사양별 시트 생성 + 도식화 작성 (함수화)
# ================================================================
def 사양별_시트_생성_및_도식화(
    결과통합문서,
    사양별_트리,
    셀너비,
    두꺼운테두리,
    도식화_시작행,
    도식화_시작열
):

    for 사양명, 트리 in 사양별_트리.items():

        # --------------------------------------------------
        # 1) 시트 생성
        # --------------------------------------------------
        ws = 결과통합문서.create_sheet(title=str(사양명)[:31])

        # --------------------------------------------------
        # 2) 열 너비 설정
        # --------------------------------------------------
        for col in range(1, 200):
            ws.column_dimensions[get_column_letter(col)].width = 셀너비

        # --------------------------------------------------
        # 3) 시트 기본 옵션
        # --------------------------------------------------
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 60

        # --------------------------------------------------
        # 4) 상단 제목 박스
        # --------------------------------------------------
        ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=24)

        title_cell = ws.cell(row=2, column=2)
        title_cell.value = f"{사양명} 조립단위"
        title_cell.font = Font(size=32, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # 병합 영역 전체 테두리
        for r in range(2, 6):
            for c in range(2, 25):
                ws.cell(row=r, column=c).border = 두꺼운테두리

        # --------------------------------------------------
        # 5) 트리 도식화
        # --------------------------------------------------
        트리_도식화(
            ws,
            트리,
            시작행=도식화_시작행,
            시작열=도식화_시작열
        )





# ================================================================
# 14. 박스 간 연결선 그리기
# ================================================================

# -------------------------------------------------------
# 가로선 / 세로선 그리기 기본 함수 (두꺼운선 사용)
# -------------------------------------------------------
def 가로선(ws, r, c1, c2):
    for c in range(c1, c2 + 1):
        cell = ws.cell(r, c)
        cell.border = Border(
            top=두꺼운선,
            left=cell.border.left,
            right=cell.border.right,
            bottom=cell.border.bottom
        )

def 세로선(ws, c, r1, r2):
    for r in range(r1, r2):
        cell = ws.cell(r, c)
        cell.border = Border(
            left=두꺼운선,
            top=cell.border.top,
            bottom=cell.border.bottom,
            right=cell.border.right
        )


# ================================================================
# 14-1. 부모–자식 연결선 그리기 (함수화)
# ================================================================
def 부모_자식_연결선_그리기(ws, 부모노드, 자식노드목록):

    if not 자식노드목록:
        return

    # 박스 크기 정보
    박스_세로 = 전체박스_세로
    박스_간격 = 레벨_이동칸 - 전체박스_가로
    박스_높이 = 박스_세로 + 박스_간격

    # 부모 박스 중앙 좌표
    p_top = 부모노드["좌표"]["start_row"]
    p_right = 부모노드["좌표"]["end_col"]

    parent_center_r = p_top + (박스_세로 // 2)
    parent_center_c = p_right

    # 허브 위치
    hub_r = parent_center_r
    hub_c = parent_center_c + 박스_간격 - 1

    # STEP 1) 부모 → 허브
    가로선(ws, parent_center_r, parent_center_c + 1, hub_c)

    # STEP 2) 첫 번째 자식
    첫째 = 자식노드목록[0]
    c1_top = 첫째["좌표"]["start_row"]
    c1_left = 첫째["좌표"]["start_col"]

    child1_center_r = c1_top + (박스_세로 // 2)
    child1_center_c = c1_left

    가로선(ws, child1_center_r, hub_c, child1_center_c - 1)

    # STEP 3) 두 번째 이후 자식
    기준_r = child1_center_r

    for child in 자식노드목록[1:]:

        기준_r += 박스_높이

        ch_top = child["좌표"]["start_row"]
        ch_left = child["좌표"]["start_col"]

        child_center_r = ch_top + (박스_세로 // 2)
        child_center_c = ch_left

        세로선(ws, hub_c, hub_r, child_center_r)
        가로선(ws, child_center_r, hub_c, child_center_c - 1)


# ================================================================
# 14-2. 트리 전체 순회하며 연결선 그리기 (함수화)
# ================================================================
def 트리_연결선_전체_그리기(결과통합문서, 사양별_트리):


    for 사양명, 트리 in 사양별_트리.items():

        ws = 결과통합문서[str(사양명)[:31]]

        def 순회(node):
            부모_자식_연결선_그리기(ws, node, node["자식"])
            for child in node["자식"]:
                순회(child)

        for root in 트리:
            순회(root)



# ================================================================
# 15. 결과 파일 저장 (함수화)
# ================================================================
def 결과파일_저장(결과통합문서, 결과파일):


    # 디렉토리 없으면 생성
    저장폴더 = os.path.dirname(결과파일)
    if 저장폴더 and not os.path.exists(저장폴더):
        os.makedirs(저장폴더)

    결과통합문서.save(결과파일)




def 단일시트_사양트리_생성(ws):


    시트명 = ws.title.strip()
    최대행 = ws.max_row
    최대열 = ws.max_column

    # [2]
    수량행, 수량_왼쪽열, 수량_오른열, 헤더깊이 = 수량_병합셀_탐색(ws)

    # [3][4]
    수량_왼쪽열, 수량_오른열, 사양컬럼_활성화 = 수량병합해제_및_사양컬럼_비활성화(
        ws, 수량행, 수량_왼쪽열, 수량_오른열
    )

    #[5] 사양명 생성
    사양목록 = 사양명_생성(
        ws,
        시트명,
        수량행,
        수량_왼쪽열,
        수량_오른열,
        헤더깊이,
        사양컬럼_활성화
    )

    # [6]
    조립단위행, 조립단위_왼쪽열, 조립단위_오른열 = 조립단위_병합셀_탐색(ws)

    # [7]
    데이터시작행, 데이터끝행 = 데이터구간_탐색_스타기반(
        ws,
        조립단위행,
        조립단위_왼쪽열,
        조립단위_오른열,
        최대행,
        허용_공백_연속수=30
    )

    # [8][8-2][8-3]
    품명열 = 품명열_탐색(ws, 최대행, 최대열)
    품번열 = 품번열_탐색(ws, 최대행, 최대열)
    재질열 = 재질열_탐색(ws, 최대행, 최대열)

    # [9]
    노드목록 = 조립노드_파싱(
        ws,
        데이터시작행,
        최대행,
        조립단위_왼쪽열,
        조립단위_오른열,
        품명열,
        품번열,
        재질열
    )

    # [10][10-1][11]
    사양별_수량 = 사양별_수량_매핑(ws, 사양목록, 노드목록)
    보기용 = 사양별_수량_보기용_변환(사양별_수량)
    사양별_트리 = 사양별_BOM_트리_생성(보기용)

    return 사양별_트리





# ================================================================
# 🚀 main 실행부 (최종)
# ================================================================
if __name__ == "__main__":

    import sys

    if len(sys.argv) >= 3:
        원본파일 = sys.argv[1]
        결과파일 = sys.argv[2]

    if len(sys.argv) >= 4:
        선택사양 = sys.argv[3]
    else:
        선택사양 = None

    # [0-1] xls → xlsx 변환
    원본파일_xlsx = 변환_xls_to_xlsx_안전(원본파일)
    
    print("[DEBUG] sys.argv =", sys.argv)
    print("[DEBUG] 원본파일 =", 원본파일)
    print("[DEBUG] 원본파일_xlsx =", 원본파일_xlsx)

    # [1] openpyxl 로드
    wb = openpyxl.load_workbook(원본파일_xlsx, data_only=True)

    결과통합문서 = 결과파일_초기화()

    전체_사양별_트리 = {}

    # ✅ 여기서 시트 선택 가능
    선택시트 = None          

    for ws in wb.worksheets:
        if 선택시트 and ws.title not in 선택시트:
            continue

        사양별_트리 = 단일시트_사양트리_생성(ws)
        전체_사양별_트리.update(사양별_트리)

    # [13]
    사양별_시트_생성_및_도식화(
        결과통합문서,
        전체_사양별_트리,
        셀너비,
        두꺼운테두리,
        도식화_시작행,
        도식화_시작열
    )

    # [14]
    트리_연결선_전체_그리기(
        결과통합문서,
        전체_사양별_트리
    )

    # [15]
    결과파일_저장(
        결과통합문서,
        결과파일
    )
