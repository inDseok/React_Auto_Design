import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import json
import os


# json_경로 = r"C:\Users\USER\Desktop\공정설계 자동화\app\sub_files\JSON 자식, 레벨 삭제\HL_STD_LHD_LD_parent.json"
json_경로 = r"C:\Users\USER\Desktop\HL_STD_LHD_LD 3.json"
결과파일 = r"C:\Users\USER\Desktop\서브단위 부품구성도.xlsx"


def 안전한_시트명(workbook, raw_title: str) -> str:
    base = (raw_title or "Sheet").strip() or "Sheet"
    invalid_chars = set('[]:*?/\\')
    sanitized = "".join("_" if ch in invalid_chars else ch for ch in base)[:31] or "Sheet"

    if sanitized not in workbook.sheetnames:
        return sanitized

    suffix = 2
    while True:
        candidate = f"{sanitized[:28]}_{suffix}"[:31]
        if candidate not in workbook.sheetnames:
            return candidate
        suffix += 1


# ================================================================
# 1-2. 결과 파일 생성 (테스트용)
# ================================================================
def 결과파일_초기화():
    """
    도식화 결과를 담을 Workbook 생성
    - 기본 시트 제거
    """

    wb = openpyxl.Workbook()

    # 기본 sheet 제거
    기본시트 = wb.active
    wb.remove(기본시트)

    return wb


# ================================================================
# 2-1. 레벨 자동 적용
# ================================================================
def 부모기반_레벨적용(bom_json: dict) -> dict:
    """
    부모 참조형 BOM JSON에 대해
    '부모' 정보를 이용해 '레벨'을 자동 계산하여 업데이트

    JSON -> JSON
    """

    for 사양명, 항목목록 in bom_json.items():

        # 1️⃣ 품명 → 노드 매핑
        노드맵 = {항목["노드ID"]: 항목 for 항목 in 항목목록}

        # 2️⃣ 레벨 계산 함수 (재귀)
        def 레벨계산(노드):
            # 이미 레벨이 있으면 그대로 사용
            if "레벨" in 노드:
                return 노드["레벨"]

            부모품명 = 노드.get("부모")

            # 루트 노드
            if 부모품명 is None:
                노드["레벨"] = 0
                return 0

            # 부모 노드 조회
            부모노드 = 노드맵.get(부모품명)
            if 부모노드 is None:
                raise ValueError(f"부모 노드를 찾을 수 없음: {부모품명}")

            레벨 = 레벨계산(부모노드) + 1
            노드["레벨"] = 레벨
            return 레벨

        # 3️⃣ 전체 노드에 대해 레벨 계산
        for 항목 in 항목목록:
            레벨계산(항목)

    return bom_json

# def 신규JSON_부모참조형_변환(raw_json: dict) -> dict:
#     meta = raw_json["meta"]
#     nodes = raw_json["nodes"]

#     사양명 = meta["spec_name"]

#     결과 = []

#     for idx, node in enumerate(nodes, start=1):
#         결과.append({
#             "행번호": idx,                     # 기존 호환용
#             "품명": node["id"],
#             "order": node.get("order", 0),
#             "부모": node.get("parent_name"),
#             "자식ID": node.get("name"),
#             # type 제외
#             "수량": int(node.get("qty", 0)),
#             "재질": node.get("material") or "",
            
#             "품번": node.get("part_no", ""),
#             "수량": int(node.get("qty", 0))# ⭐ 추가
#         })

#     return {사양명: 결과}

def 신규JSON_부모참조형_변환(raw_json: dict) -> dict:
    meta = raw_json["meta"]
    nodes = raw_json["nodes"]

    사양명 = meta["spec_name"]
    결과 = []

    for idx, node in enumerate(nodes, start=1):
        결과.append({
            "type" : node["type"],
            "행번호": idx,
            "품명": node["id"],                 # 표시용
            "노드ID": node["name"],              # 관계용 (고유)
            "부모": node.get("parent_name"),     # 관계용 (부모 노드ID)
            "order": node.get("order", 0),       # ★ 필수
            "품번": node.get("part_no", ""),
            "수량": int(node.get("qty", 0)),
            "재질": node.get("material") or ""
        })

    return {사양명: 결과}


# ================================================================
# 2-2. 트리변환
# ================================================================
def 부모참조형_JSON_트리변환(bom_json: dict) -> dict:

    결과 = {}

    for 사양명, 항목목록 in bom_json.items():

        # 1️⃣ 노드 복사 + 자식 리스트 초기화
        노드맵 = {}
        for 항목 in 항목목록:
            노드 = {
                "행번호": 항목["행번호"],
                "type" : 항목["type"],
                "노드ID": 항목["노드ID"],      # 🔑 관계용
                "품명": 항목["품명"],          # 표시용
                "품번": 항목["품번"],
                "레벨": 항목["레벨"],
                "수량": 항목["수량"],
                "재질": 항목.get("재질", ""),
                "order": 항목.get("order", 0),
                "자식": []
            }
            노드맵[항목["노드ID"]] = 노드

        # 2️⃣ 루트 노드 수집
        루트노드들 = []

        # 3️⃣ 부모 → 자식 연결
        for 항목 in 항목목록:
            현재노드 = 노드맵[항목["노드ID"]]
            부모ID = 항목.get("부모")

            if 부모ID is None:
                루트노드들.append(현재노드)
            else:
                부모노드 = 노드맵.get(부모ID)
                if 부모노드 is None:
                    raise ValueError(f"부모 노드 없음: {부모ID}")
                부모노드["자식"].append(현재노드)

        # 4️⃣ 자식 정렬 (order → 행번호)
        for 노드 in 노드맵.values():
            노드["자식"].sort(key=lambda x: (x.get("order", 0), x["행번호"]))

        결과[사양명] = 루트노드들

    return 결과




# -------------------------------------------------------
# 📌 양식 박스를 그리는 함수
# -------------------------------------------------------
def 양식박스_그리기(ws, 시작행, 시작열, 품명, 품번, 수량, 재질, type):
    """
    ws(워크시트)에 하나의 부품 박스를 생성함.
    박스 구성:
    - 왼쪽 큰 이미지 영역 (8행 × 6열)
    - 오른쪽 상단 ‘부품명’
    - 양산처
    - QT'Y / EA
    - 재질
    """

    r = 시작행
    c = 시작열

    # ------------------------------------------
    # 0) 왼쪽 이미지 큰 박스 (8행, 6칸 병합) 
    # ------------------------------------------

    m_start_row = r
    m_start_column = c
    m_end_row = r + 7
    m_end_column = c + 5   

    #병합
    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
                end_row=m_end_row, end_column=m_end_column)


    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리

    # ------------------------------------------
    # 1-1) 오른쪽 상단 - '부품명'
    # ------------------------------------------
    m_start_row = r
    m_start_column = c + 6
    m_end_row = r + 1
    m_end_column = c + 7

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
                end_row=m_end_row, end_column=m_end_column)

    cell = ws.cell(m_start_row, m_start_column)
    cell.value = "부품명"
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리
            ws.cell(rr, cc).fill = 회색채움  
    
    # ------------------------------------------
    # 1-2) 오른쪽 상단 - 부품명
    # ------------------------------------------
    m_start_row = r
    m_start_column = c + 8
    m_end_row = r + 1
    m_end_column = c + 11

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
                end_row=m_end_row, end_column=m_end_column)

    cell = ws.cell(m_start_row, m_start_column)
    cell.value = 품명
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리  

    # ------------------------------------------
    # 2-1) 품번
    # ------------------------------------------
    m_start_row = r+2
    m_start_column = c + 6
    m_end_row = r + 3
    m_end_column = c + 7

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
                end_row=m_end_row, end_column=m_end_column)

    cell = ws.cell(m_start_row, m_start_column)
    cell.value = "품번"
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리
            ws.cell(rr, cc).fill = 회색채움  

    # ------------------------------------------
    # 2-2) 품번값
    # ------------------------------------------
    m_start_row = r+2
    m_start_column = c + 8
    m_end_row = r + 3
    m_end_column = c + 11

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
            end_row=m_end_row, end_column=m_end_column)
    
    cell = ws.cell(m_start_row, m_start_column)
    cell.value = f"{품번}"
    cell.number_format = "@"       # ← ★ 텍스트로 인식되도록 지정
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리  



    # ------------------------------------------
    # 3-1) 재질
    # ------------------------------------------
    m_start_row = r+4
    m_start_column = c + 6
    m_end_row = r + 5
    m_end_column = c + 7

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
                end_row=m_end_row, end_column=m_end_column)

    cell = ws.cell(m_start_row, m_start_column)
    cell.value = "재질"
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리
            ws.cell(rr, cc).fill = 회색채움  

    # ------------------------------------------
    # 3-2) 재질값
    # ------------------------------------------
    m_start_row = r+4
    m_start_column = c + 8
    m_end_row = r + 5
    m_end_column = c + 11

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
            end_row=m_end_row, end_column=m_end_column)
    
    cell = ws.cell(m_start_row, m_start_column)
    cell.value = f"{재질}"
    cell.number_format = "@"       # ← ★ 텍스트로 인식되도록 지정
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리  





    # ------------------------------------------
    # 4-1 ) '수량'입력 (text)
    # ------------------------------------------
    m_start_row = r+6
    m_start_column = c + 6
    m_end_row = r + 7
    m_end_column = c + 7

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
            end_row=m_end_row, end_column=m_end_column)
    
    cell = ws.cell(m_start_row, m_start_column)
    cell.value = '수량'
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리
            ws.cell(rr, cc).fill = 회색채움  

    # ------------------------------------------
    # 4-2 ) 수량 입력 (value)
    # ------------------------------------------
    m_start_row = r + 6
    m_start_column = c + 8
    m_end_row = r + 7
    m_end_column = c + 10

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
                end_row=m_end_row, end_column=m_end_column)

    cell = ws.cell(m_start_row, m_start_column)
    cell.value = f"{수량}"
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리  



    # ------------------------------------------
    # 4-3 ) EA
    # ------------------------------------------
    # EA
    m_start_row = r + 6
    m_start_column = c + 11
    m_end_row = r + 7
    m_end_column = c + 11

    ws.merge_cells(start_row=m_start_row, start_column=m_start_column,
                end_row=m_end_row, end_column=m_end_column)

    cell = ws.cell(m_start_row, m_start_column)
    cell.value = "EA"
    cell.alignment = 가운데정렬

    # 병합된 범위 전체 테두리 적용
    for rr in range(m_start_row, m_end_row + 1):
        for cc in range(m_start_column, m_end_column + 1):
            ws.cell(rr, cc).border = 테두리  



    # ------------------------------------------
    # 5) 전체 박스 외곽 테두리 (4 LINE 최적화)                  ############박스_세로, 박스_간격도 함께 변경#######################
    # ------------------------------------------
    m_start_row = r 
    m_start_column = c 
    m_end_row = r + 전체박스_세로 - 1      # 총 8행 높이
    m_end_column = c + 전체박스_가로 - 1   # 총 12열 폭

    사용_테두리 = 빨간_두꺼운선 if type == "SUB" else 두꺼운선

    # 5-1) ★ 위쪽 라인
    for cc in range(m_start_column, m_end_column + 1):
        cell = ws.cell(m_start_row, cc)
        cell.border = Border(
            top=사용_테두리,
            left=cell.border.left,
            right=cell.border.right,
            bottom=cell.border.bottom
        )

    # 5-2) ★ 아래쪽 라인
    for cc in range(m_start_column, m_end_column + 1):
        cell = ws.cell(m_end_row, cc)
        cell.border = Border(
            bottom=사용_테두리,
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top
        )

    # 5-3) ★ 왼쪽 라인
    for rr in range(m_start_row, m_end_row + 1):
        cell = ws.cell(rr, m_start_column)
        cell.border = Border(
            left=사용_테두리,
            top=cell.border.top,
            bottom=cell.border.bottom,
            right=cell.border.right
        )

    # 5-4) ★ 오른쪽 라인
    for rr in range(m_start_row, m_end_row + 1):
        cell = ws.cell(rr, m_end_column)
        cell.border = Border(
            right=사용_테두리,
            top=cell.border.top,
            bottom=cell.border.bottom,
            left=cell.border.left
        )



    # 박스 하나의 높이는 총 8행 공간 사용 
    return r # 다음 박스가 시작될 행 반환


# -------------------------------------------------------
# 📌 트리 구조를 도식화하는 함수
# -------------------------------------------------------
# -------------------------------------------------------
# 📌 트리 구조를 도식화하는 함수 (최종 완성본)
# -------------------------------------------------------
def 트리_도식화(ws, 트리, 시작행, 시작열):

    def 출력_재귀(노드, 기준행, 이전형제레벨):
        """
        노드 1개 + 해당 노드의 모든 자식 subtree를 출력하고
        이 subtree가 사용한 마지막 행을 반환
        """

        # -------------------------------
        # 1. 기본 정보
        # -------------------------------
        현재레벨 = 노드["레벨"]

        # -------------------------------
        # 2. 출력 행 결정 규칙 (핵심)
        # -------------------------------
        if 이전형제레벨 is None:
            # 첫 형제 or 부모→첫자식
            출력행 = 기준행

        elif 현재레벨 == 이전형제레벨:
            # 같은 레벨 형제 → 아래로 이동
            출력행 = 기준행 + 박스_높이

        else:
            # 부모→자식 체인
            출력행 = 기준행

        # -------------------------------
        # 3. 출력 열 결정
        # -------------------------------
        박스열 = 시작열 + (현재레벨 * 레벨_이동칸)

        # -------------------------------
        # 4. 박스 출력
        # -------------------------------
        양식박스_그리기(
            ws,
            시작행=출력행,
            시작열=박스열,
            품명=노드["품명"],
            품번=노드["품번"],
            수량=노드["수량"],
            재질=노드["재질"],
            type=노드["type"]
        )

        # -------------------------------
        # 5. 좌표 저장 (연결선용)
        # -------------------------------
        노드["좌표"] = {
            "start_row": 출력행,
            "start_col": 박스열,
            "end_row": 출력행 + 전체박스_세로 - 1,
            "end_col": 박스열 + 전체박스_가로 - 1
        }

        # -------------------------------
        # 6. 자식 subtree 출력
        # -------------------------------
        다음기준행 = 출력행
        이전형제 = None

        for 자식 in sorted(노드["자식"], key=lambda x: x.get("order", 0)):
            다음기준행 = 출력_재귀(자식, 다음기준행, 이전형제)
            이전형제 = 자식["레벨"]

        # -------------------------------
        # 7. 이 subtree가 사용한 마지막 행 반환
        # -------------------------------
        return max(다음기준행, 출력행 + 박스_높이)

    # ===================================================
    # 루트 노드 출력 제어
    # ===================================================
    현재행 = 시작행
    이전루트레벨 = None

    for 루트노드 in sorted(트리, key=lambda x: x.get("order", 0)):
        현재행 = 출력_재귀(루트노드, 현재행, 이전루트레벨)
        이전루트레벨 = 루트노드["레벨"]







# ================================================================
# 13. 사양별 시트 생성 + 도식화 작성 (함수화)
# ================================================================
def 사양별_시트_생성_및_도식화(
    결과통합문서,
    사양별_트리,
    셀너비,
    두꺼운테두리,
    도식화_시작행,
    도식화_시작열,
    시트명_생성함수=None,
    제목_생성함수=None,
):
    """
    사양별 BOM 트리를 기반으로
    - 시트 생성
    - 시트 기본 설정
    - 상단 제목 생성
    - 트리 도식화 수행
    """

    for 사양명, 트리 in 사양별_트리.items():

        # --------------------------------------------------
        # 1) 시트 생성
        # --------------------------------------------------
        시트명 = 시트명_생성함수(사양명, 결과통합문서) if 시트명_생성함수 else str(사양명)[:31]
        ws = 결과통합문서.create_sheet(title=안전한_시트명(결과통합문서, 시트명))

        # --------------------------------------------------
        # 2) 열 너비 설정
        # --------------------------------------------------
        for col in range(1, 200):
            ws.column_dimensions[get_column_letter(col)].width = 셀너비

        # --------------------------------------------------
        # 3) 시트 기본 옵션
        # --------------------------------------------------
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 60

        # --------------------------------------------------
        # 4) 상단 제목 박스
        # --------------------------------------------------
        ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=24)

        title_cell = ws.cell(row=2, column=2)
        title_cell.value = 제목_생성함수(사양명) if 제목_생성함수 else f"{사양명} 조립단위"
        title_cell.font = Font(size=32, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # 병합 영역 전체 테두리
        for r in range(2, 6):
            for c in range(2, 25):
                ws.cell(row=r, column=c).border = 두꺼운테두리

        # --------------------------------------------------
        # 5) 트리 도식화
        # --------------------------------------------------
        트리_도식화(
            ws,
            트리,
            시작행=도식화_시작행,
            시작열=도식화_시작열
        )




# ================================================================
# 14. 박스 간 연결선 그리기
# ================================================================

# -------------------------------------------------------
# 가로선 / 세로선 그리기 기본 함수 (두꺼운선 사용)
# -------------------------------------------------------
def 가로선(ws, r, c1, c2):
    """행 r에서 c1~c2까지 가로선(top) 1줄만 긋기"""
    for c in range(c1, c2 + 1):
        cell = ws.cell(r, c)
        cell.border = Border(
            top=두꺼운선,
            left=cell.border.left,
            right=cell.border.right,
            bottom=cell.border.bottom
        )

def 세로선(ws, c, r1, r2):
    """열 c에서 r1~r2까지 세로선(left) 1줄만 긋기"""
    for r in range(r1, r2):
        cell = ws.cell(r, c)
        cell.border = Border(
            left=두꺼운선,
            top=cell.border.top,
            bottom=cell.border.bottom,
            right=cell.border.right
        )
# ================================================================
# 14-1. 부모–자식 연결선 그리기 (함수화)
# ================================================================
def 부모_자식_연결선_그리기(ws, 부모노드, 자식노드목록):
    """
    하나의 부모 노드와 그 자식 노드들을 연결
    """

    if not 자식노드목록:
        return

    # 박스 크기 정보
    박스_세로 = 전체박스_세로
    박스_간격 = 레벨_이동칸 - 전체박스_가로
    박스_높이 = 박스_세로 + 박스_간격

    # 부모 박스 중앙 좌표
    p_top = 부모노드["좌표"]["start_row"]
    p_right = 부모노드["좌표"]["end_col"]

    parent_center_r = p_top + (박스_세로 // 2)
    parent_center_c = p_right

    # 허브 위치
    hub_r = parent_center_r
    hub_c = parent_center_c + 박스_간격 - 1

    # STEP 1) 부모 → 허브
    가로선(ws, parent_center_r, parent_center_c + 1, hub_c)

    # STEP 2) 첫 번째 자식
    첫째 = 자식노드목록[0]
    c1_top = 첫째["좌표"]["start_row"]
    c1_left = 첫째["좌표"]["start_col"]

    child1_center_r = c1_top + (박스_세로 // 2)
    child1_center_c = c1_left

    가로선(ws, child1_center_r, hub_c, child1_center_c - 1)

    # STEP 3) 두 번째 이후 자식
    기준_r = child1_center_r

    for child in 자식노드목록[1:]:

        기준_r += 박스_높이

        ch_top = child["좌표"]["start_row"]
        ch_left = child["좌표"]["start_col"]

        child_center_r = ch_top + (박스_세로 // 2)
        child_center_c = ch_left

        세로선(ws, hub_c, hub_r, child_center_r)
        가로선(ws, child_center_r, hub_c, child_center_c - 1)


# ================================================================
# 14-2. 트리 전체 순회하며 연결선 그리기 (함수화)
# ================================================================
def 트리_연결선_전체_그리기(결과통합문서, 사양별_트리, 시트명_생성함수=None):
    """
    사양별 트리를 순회하며
    모든 부모–자식 연결선을 그림
    """

    for 사양명, 트리 in 사양별_트리.items():
        시트명 = 시트명_생성함수(사양명, 결과통합문서) if 시트명_생성함수 else str(사양명)[:31]
        ws = 결과통합문서[시트명]

        def 순회(node):
            부모_자식_연결선_그리기(ws, node, node["자식"])
            for child in node["자식"]:
                순회(child)

        for root in 트리:
            순회(root)



# ================================================================
# 15. 결과 파일 저장 (함수화)
# ================================================================
def 결과파일_저장(결과통합문서, 결과파일):
    """
    결과 워크북을 지정 경로에 저장
    """

    # 디렉토리 없으면 생성
    저장폴더 = os.path.dirname(결과파일)
    if 저장폴더 and not os.path.exists(저장폴더):
        os.makedirs(저장폴더)

    결과통합문서.save(결과파일)







def 엑셀_스타일_초기화():
    global 전체박스_가로, 전체박스_세로
    global 레벨_이동칸, 박스_높이
    global 두꺼운선, 얇은선, 테두리, 두꺼운테두리, 빨간_두꺼운선
    global 가운데정렬, 회색채움

    전체박스_가로 = 12
    전체박스_세로 = 8
    레벨_이동칸 = 16
    박스_높이 = 5

    두꺼운선 = Side(border_style="thick", color="000000")
    빨간_두꺼운선 = Side(style="thick", color="FF0000")
    얇은선 = Side(border_style="thin", color="000000")
    테두리 = Border(left=얇은선, right=얇은선, top=얇은선, bottom=얇은선)
    두꺼운테두리 = Border(
        left=두꺼운선,
        right=두꺼운선,
        top=두꺼운선,
        bottom=두꺼운선,
    )

    가운데정렬 = Alignment(horizontal="center", vertical="center")
    회색채움 = PatternFill(
        start_color="D9D9D9",
        end_color="D9D9D9",
        fill_type="solid",
    )


def build_tree_workbook_from_json(raw_json: dict, workbook=None, sheet_name_resolver=None, title_text_resolver=None):
    bom_json = 신규JSON_부모참조형_변환(raw_json)
    결과통합문서 = workbook or 결과파일_초기화()

    엑셀_스타일_초기화()

    결과_json = 부모기반_레벨적용(bom_json)
    전체_사양별_트리 = 부모참조형_JSON_트리변환(결과_json)

    사양별_시트_생성_및_도식화(
        결과통합문서,
        전체_사양별_트리,
        셀너비=4.2,
        두꺼운테두리=두꺼운테두리,
        도식화_시작행=8,
        도식화_시작열=3,
        시트명_생성함수=sheet_name_resolver,
        제목_생성함수=title_text_resolver,
    )

    트리_연결선_전체_그리기(
        결과통합문서,
        전체_사양별_트리,
        시트명_생성함수=sheet_name_resolver,
    )

    return 결과통합문서


def export_tree_excel_from_json(raw_json: dict, output_path: str):
    결과통합문서 = build_tree_workbook_from_json(raw_json)
    결과파일_저장(결과통합문서, output_path)

