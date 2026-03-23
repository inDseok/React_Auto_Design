from __future__ import annotations

from typing import Dict, List, Optional, Any
from functools import lru_cache

import json
import io
from uuid import uuid4

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse

from pathlib import Path

from backend.models import SubNodePatch, SubTree, MoveNodeRequest, SubNode, NodeType
from backend.Sub.bom_service import create_bom_run
from backend.Sub.utills import load_tree_json, save_tree_json, read_bom_meta, read_spec_meta
from backend.Sub.excel_loader import build_tree_from_sheet
from typing import Optional
from openpyxl import load_workbook
from backend.Sub.session_store import get_or_create_sid, refresh_session_state, save_session_state, SESSION_STATE

from backend.Sub.json_to_excel import export_tree_excel_from_json, 결과파일_초기화, build_tree_workbook_from_json
from backend.Assembly.json_to_excel import append_assembly_sheet_to_workbook
from pathlib import Path
from backend.Assembly.auto_match import load_db_rows, normalize_text, rf_score, jw_score

sub_router = APIRouter(prefix="/sub", tags=["SUB API"])


BASE_DIR = Path(__file__).resolve().parents[1]   # backend
DATA_DIR = BASE_DIR / "backend" /"data"
SESSION_STORE_PATH = DATA_DIR / "session_state.json"
ASSEMBLY_DB_PATH = BASE_DIR / "backend" / "작업시간분석표DB.xlsx"


@lru_cache(maxsize=4)
def _get_cached_db_rows(excel_path_str: str, mtime: float):
    return load_db_rows(Path(excel_path_str))


def _get_part_suggestions(query_values: List[str], limit: int = 5) -> List[Dict[str, Any]]:
    if not ASSEMBLY_DB_PATH.exists():
        raise HTTPException(status_code=500, detail=f"작업시간 분석표 DB 엑셀 없음: {ASSEMBLY_DB_PATH}")

    db_rows, db_choices = _get_cached_db_rows(
        str(ASSEMBLY_DB_PATH.resolve()),
        ASSEMBLY_DB_PATH.stat().st_mtime,
    )

    suggestions: Dict[str, Dict[str, Any]] = {}

    for query_raw in query_values:
        normalized_query = normalize_text(query_raw)
        if not normalized_query:
            continue

        for index, choice_norm in enumerate(db_choices):
            rf = rf_score(normalized_query, choice_norm)
            jw = jw_score(normalized_query, choice_norm)
            combined = round((rf * 0.45) + (jw * 0.55), 2)
            meta = db_rows[index]
            key = f"{meta['db_part_raw']}::{meta['sheet']}"

            previous = suggestions.get(key)
            candidate = {
                "query": query_raw,
                "db_part_raw": meta["db_part_raw"],
                "db_part_norm": meta["db_part_norm"],
                "sheet": meta["sheet"],
                "row_index": meta["row_index"],
                "score_rapidfuzz": round(rf, 2),
                "score_jaro_winkler": round(jw, 2),
                "score_combined": combined,
            }

            if previous is None or candidate["score_combined"] > previous["score_combined"]:
                suggestions[key] = candidate

    ranked = sorted(
        suggestions.values(),
        key=lambda item: (
            item["score_combined"],
            item["score_jaro_winkler"],
            item["score_rapidfuzz"],
        ),
        reverse=True,
    )

    return ranked[:limit]

@sub_router.post("/bom/upload")
async def upload_bom(file: UploadFile = File(...)):
    binary = await file.read()
    try:
        meta = create_bom_run(binary, file.filename)

        # 🔹 bom_filename 저장
        root = DATA_DIR / "bom_runs" / meta["bom_id"]
        bom_meta_path = root / "bom_meta.json"
        bom_meta_path.write_text(
            json.dumps(
                {
                    "bom_id": meta["bom_id"],
                    "bom_filename": file.filename,
                },
                ensure_ascii=False,
                indent=2
            ),
            encoding="utf-8"
        )

        return meta
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@sub_router.get("/bom/{bom_id}/specs")
def list_specs(bom_id: str):
    root = DATA_DIR / "bom_runs" / bom_id
    spec_meta = read_spec_meta(root)

    sheets = spec_meta.get("spec_info", {}).get("sheets", [])
    if isinstance(sheets, list) and sheets:
        specs = []
        for sheet in sheets:
            spec_items = sheet.get("specs", [])
            if not isinstance(spec_items, list):
                continue
            for item in spec_items:
                spec_name = item.get("spec_name")
                if spec_name:
                    specs.append(str(spec_name))

        if specs:
            return specs

    tree_excel = root / "tree.xlsx"

    if not tree_excel.exists():
        raise HTTPException(
            status_code=404,
            detail="tree.xlsx 파일이 없습니다."
        )

    try:
        wb = load_workbook(tree_excel, data_only=True)
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"엑셀 로드 실패: {e}"
        )

    # sheet 이름 = spec 목록
    specs = wb.sheetnames

    if not specs:
        raise HTTPException(
            status_code=400,
            detail="사양 시트가 없습니다."
        )

    return specs


@sub_router.get("/bom/{bom_id}/part-suggestions")
def get_part_suggestions(
    bom_id: str,
    spec: str,
    name: Optional[str] = None,
    part_no: Optional[str] = None,
    limit: int = 5,
):
    root_dir = DATA_DIR / "bom_runs" / bom_id
    json_path = root_dir / f"{spec}.json"

    if not json_path.exists():
        raise HTTPException(status_code=404, detail=f"트리 JSON이 없습니다: {json_path}")

    query_values = [value for value in [name, part_no] if value and str(value).strip()]
    if not query_values:
        return {"items": []}

    return {"items": _get_part_suggestions(query_values, max(1, min(limit, 10)))}

@sub_router.get("/bom/{bom_id}/tree", response_model=SubTree)
def get_tree(
    bom_id: str,
    request: Request,
    response: Response,
    spec: Optional[str] = None,
):
    sid = get_or_create_sid(request, response)
    refresh_session_state()
    state = SESSION_STATE.get(sid, {})

    resolved_spec = spec or state.get("spec")
    if not resolved_spec:
        raise HTTPException(status_code=400, detail="spec이 없습니다.")

    root_dir = DATA_DIR / "bom_runs" / bom_id
    tree_json_path = root_dir / f"{resolved_spec}.json"

    # 1. 캐시된 JSON 있으면 그대로 사용
    if tree_json_path.exists():
        tree = load_tree_json(root_dir, resolved_spec)
    else:
        # 2. 없으면 엑셀에서 생성
        tree_excel = root_dir / "tree.xlsx"
        if not tree_excel.exists():
            raise HTTPException(status_code=404, detail="tree.xlsx 파일이 없습니다.")
        
        bom_meta = read_bom_meta(root_dir)
        bom_filename = bom_meta.get("bom_filename")
        if not bom_filename:
            raise HTTPException(status_code=500, detail="bom_filename 없음")

        wb = load_workbook(tree_excel, data_only=True)
        if resolved_spec not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"시트 없음: {resolved_spec}")

        ws = wb[resolved_spec]

        tree = build_tree_from_sheet(
            ws=ws,
            bom_id=bom_id,
            bom_filename=bom_filename,
            spec_name=resolved_spec,
        )

        save_tree_json(root_dir, resolved_spec, tree)

    # 세션 갱신
    SESSION_STATE[sid] = {
        "bom_id": bom_id,
        "spec": resolved_spec,
        "selected_id": state.get("selected_id"),
    }
    save_session_state()

    return tree

def collect_descendants(nodes, parent_name: str):
    result = []
    queue = [parent_name]

    while queue:
        cur = queue.pop(0)
        children = [n for n in nodes if n.parent_name == cur]
        for child in children:
            result.append(child)
            queue.append(child.name)

    return result


@sub_router.patch("/bom/{bom_id}/node/{node_id:path}", response_model=SubTree)
def patch_node(
    bom_id: str,
    node_id: str,
    patch: SubNodePatch,
    request: Request,
    response: Response,
    spec: Optional[str] = None,
):
    sid = get_or_create_sid(request, response)
    refresh_session_state()
    state = SESSION_STATE.get(sid, {})

    resolved_spec = spec or state.get("spec")
    if not resolved_spec:
        raise HTTPException(status_code=400, detail="spec 없음")

    root_dir = DATA_DIR / "bom_runs" / bom_id

    tree = load_tree_json(root_dir, resolved_spec)
    nodes = tree.nodes

    # 1️⃣ 기존 id 로 노드 찾기
    target = next((n for n in nodes if n.id == node_id), None)
    if not target:
        raise HTTPException(status_code=404, detail="노드 없음")

    data = patch.dict(exclude_unset=True)

    # SUB 전파 여부 판단
    propagate_sub = False
    if "type" in data:
        if target.type != NodeType.SUB and data["type"] == NodeType.SUB:
            propagate_sub = True


    # ---------------------------
    # 2️⃣ id 변경 처리 (PK 변경)
    # ---------------------------
    if "id" in data and data["id"] and data["id"] != node_id:
        new_id = data["id"]

        # 같은 부모 밑에 중복 id 방지

        old_id = target.id
        target.id = new_id   # PK 변경

        # 3️⃣ 자식노드 parent_name 업데이트
        for n in nodes:
            if n.parent_name == old_id:
                n.parent_name = new_id

    # ---------------------------
    # 4️⃣ 나머지 필드 적용
    # ---------------------------
    for field, value in data.items():
        if field == "id":
            continue
        setattr(target, field, value)
    # 🔥 자식까지 SUB 전파
    if propagate_sub:
        descendants = collect_descendants(nodes, target.name)
        for child in descendants:
            child.type = NodeType.SUB

    save_tree_json(root_dir, resolved_spec, tree)

    return tree


@sub_router.patch("/bom/{bom_id}/move-node")
def move_node(
    bom_id: str,
    spec: str,
    req: MoveNodeRequest
):
    root_dir = DATA_DIR / "bom_runs" / bom_id
    tree = load_tree_json(root_dir, spec)

    nodes = tree.nodes

    # 1️⃣ 이동 대상 찾기 (id 기준)
    target = next((n for n in nodes if n.id == req.node_id), None)
    if not target:
        raise HTTPException(404, "node not found")

    # 2️⃣ new_parent_name → 부모 name으로 정규화
    raw_parent = req.new_parent_name

    if raw_parent in (None, "", "null"):
        parent_name = None
    else:
        parent_node = next(
            (
                n for n in nodes
                if n.id == raw_parent or n.name == raw_parent
            ),
            None,
        )

        if not parent_node:
            raise HTTPException(400, "parent node not found")

        parent_name = parent_node.name   # ⭐ 항상 name으로 저장

    # 3️⃣ 기존 부모/형제 목록 정리
    old_parent = target.parent_name

    # 4️⃣ 부모 name + order 변경
    target.parent_name = parent_name
    target.order = req.new_index or 0
    if parent_name is not None:
        parent = next(n for n in nodes if n.name == parent_name)
        if parent.type == NodeType.SUB:
            target.type = NodeType.SUB

    # 5️⃣ 새로운 부모 아래 형제 목록 수집
    siblings = [n for n in nodes if n.parent_name == parent_name and n.id != target.id]

    # ⭐ target도 포함해서 다시 정렬 대상으로 넣음
    siblings.append(target)

    # 6️⃣ order 기준 정렬 + 재부여
    siblings.sort(key=lambda x: (x.order or 0))

    for i, n in enumerate(siblings):
        n.order = i

    # 7️⃣ 저장
    save_tree_json(root_dir, spec, tree)

    return tree


from uuid import uuid4

@sub_router.post("/bom/{bom_id}/node", response_model=SubTree)
def create_node(
    bom_id: str,
    payload: dict,
    request: Request,
    response: Response,
):
    sid = get_or_create_sid(request, response)
    refresh_session_state()
    state = SESSION_STATE.get(sid, {})

    spec = state.get("spec")
    if not spec:
        raise HTTPException(status_code=400, detail="spec 없음")

    root_dir = DATA_DIR / "bom_runs" / bom_id
    tree = load_tree_json(root_dir, spec)
    nodes = tree.nodes

    # 1️⃣ 클라이언트에서 넘어온 parent_name (대부분 "부모 id")
    raw_parent = payload.get("parent_name")

    # 2️⃣ 저장할 때는 항상 "부모 name"으로 맞춘다
    if raw_parent in (None, "", "null"):
        parent_name = None
    else:
        # raw_parent가 id인지 name인지 섞여 있을 수 있으니 둘 다 시도
        parent_node = next(
            (
                n for n in nodes
                if n.id == raw_parent or n.name == raw_parent
            ),
            None,
        )
        if not parent_node:
            raise HTTPException(status_code=400, detail="parent 노드 없음")

        parent_name = parent_node.name  # ⭐ 최종적으로 name만 사용

    # 3️⃣ 새 노드 UI name
    name = payload.get("name") or f"{spec}:{uuid4()}"

    part_no   = payload.get("part_no")  or ""
    material  = payload.get("material") or ""
    qty       = payload.get("qty")      or 1
    if parent_name is not None:
        parent = next(n for n in nodes if n.name == parent_name)
        if parent.type == NodeType.SUB:
            node_type = NodeType.SUB

    node_type = payload.get("type")     or "PART"

    # 4️⃣ 내부 PK
    node_id = str(uuid4())

    # 5️⃣ 형제들 기준 order (부모 name 기준)
    siblings = [n for n in nodes if n.parent_name == parent_name]
    order = len(siblings)

    new_node = SubNode(
        id=node_id,
        parent_name=parent_name,  # ⭐ 여기에는 항상 "부모 name"을 저장
        order=order,
        type=node_type,
        name=name,
        part_no=part_no,
        material=material,
        qty=qty,
        inhouse=False,
        recommended_part_base=None,
        recommended_source_sheet=None,
        recommended_match_score=None,
    )

    nodes.append(new_node)

    save_tree_json(root_dir, spec, tree)

    return tree



@sub_router.delete("/bom/{bom_id}/node/{node_name:path}", response_model=SubTree)
def delete_node(
    bom_id: str,
    node_name: str,
    request: Request,
    response: Response,
):
    sid = get_or_create_sid(request, response)
    refresh_session_state()
    state = SESSION_STATE.get(sid, {})

    spec = state.get("spec")
    if not spec:
        raise HTTPException(status_code=400, detail="spec 없음")

    root_dir = DATA_DIR / "bom_runs" / bom_id

    tree = load_tree_json(root_dir, spec)
    nodes = tree.nodes

    # 1️⃣ 삭제할 노드 찾기 (UI name 기준)
    target = next((n for n in nodes if n.name == node_name), None)
    if not target:
        raise HTTPException(status_code=404, detail="노드 없음")

    # ⭐ 삭제 대상 name 목록
    to_delete = {target.name}
    visited = set()
    queue = [target.name]

    def collect_children(parent_name: str):
        return [n.name for n in nodes if n.parent_name == parent_name]

    # 2️⃣ parent_name 기준 BFS로 자식 모두 수집
    while queue:
        cur = queue.pop()
        if cur in visited:
            continue

        visited.add(cur)

        children = collect_children(cur)
        to_delete.update(children)
        queue.extend(children)

    # 3️⃣ name 기준 삭제
    tree.nodes = [n for n in nodes if n.name not in to_delete]

    save_tree_json(root_dir, spec, tree)

    return tree




from uuid import uuid4

@sub_router.post("/bom/{bom_id}/nodes", response_model=SubNode)
def add_node(
    bom_id: str,
    body: SubNode,               # ← 그대로 사용
    request: Request,
    response: Response,
):
    sid = get_or_create_sid(request, response)
    refresh_session_state()
    state = SESSION_STATE.get(sid, {})

    spec = state.get("spec")
    if not spec:
        raise HTTPException(status_code=400, detail="spec 없음")

    root_dir = DATA_DIR / "bom_runs" / bom_id

    # 1️⃣ JSON 트리 로드
    tree = load_tree_json(root_dir, spec)
    nodes = list(tree.nodes)

    # 2️⃣ 부모가 존재하는지 확인 (root는 None 가능)
    if body.parent_name is not None:
        parent_exists = any(n.id == body.parent_name for n in nodes)
        if not parent_exists:
            raise HTTPException(status_code=400, detail="부모 노드 없음")

    # 3️⃣ id는 서버에서 항상 새로 생성
    new_id = str(uuid4())

    # 4️⃣ order는 서버에서 자동 계산
    siblings = [
        n for n in nodes
        if n.parent_name == body.parent_name
    ]

    max_order = max(
        [n.order for n in siblings],
        default=-1
    )

    new_node = SubNode(
        id=new_id,                         # 🔥 덮어쓰기
        parent_name=body.parent_name,
        order=max_order + 1,               # 🔥 덮어쓰기
        name=new_id,
        type=body.type,
        part_no=body.part_no,
        material=body.material,
        qty=body.qty,
        inhouse = body.inhouse if body.inhouse is not None else False,
        recommended_part_base=None,
        recommended_source_sheet=None,
        recommended_match_score=None,
    )

    # 5️⃣ 추가
    nodes.append(new_node)

    # 6️⃣ 저장
    tree.nodes = nodes
    save_tree_json(root_dir, spec, tree)

    return new_node

@sub_router.get("/bom/{bom_id}/export_excel")
async def export_excel(bom_id: str, spec: str):

    root = DATA_DIR / "bom_runs" / bom_id
    json_path = root / f"{spec}.json"

    with open(json_path, "r", encoding="utf-8") as f:
        raw_json = json.load(f)

    output_path = root / f"{spec}_SUB구성도.xlsx"

    export_tree_excel_from_json(raw_json, str(output_path))

    return FileResponse(
        path=str(output_path),
        filename=f"{spec}_SUB구성도.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@sub_router.get("/bom/{bom_id}/export_excel_bundle")
async def export_excel_bundle(bom_id: str):
    root = DATA_DIR / "bom_runs" / bom_id
    if not root.exists():
        raise HTTPException(status_code=404, detail="BOM 폴더가 없습니다.")

    sub_json_paths = sorted(
        path for path in root.glob("*.json")
        if path.name not in {"bom_meta.json", "meta_spec.json"}
        and not path.name.endswith("_sequence.json")
        and not path.name.endswith("_assembly.json")
    )
    assembly_json_paths = sorted(root.glob("*_assembly.json"))

    if not sub_json_paths and not assembly_json_paths:
        raise HTTPException(status_code=404, detail="엑셀로 변환할 JSON 파일이 없습니다.")

    workbook = 결과파일_초기화()

    for json_path in sub_json_paths:
        with open(json_path, "r", encoding="utf-8") as file:
            raw_json = json.load(file)

        workbook = build_tree_workbook_from_json(
            raw_json,
            workbook=workbook,
            sheet_name_resolver=lambda _spec_name, _workbook: "1. sub단위 부품 구성도",
            title_text_resolver=lambda spec_name: f"{spec_name} 조립단위",
        )

    for assembly_json_path in assembly_json_paths:
        with open(assembly_json_path, "r", encoding="utf-8") as file:
            raw_json = json.load(file)

        spec_name = assembly_json_path.stem.replace("_assembly", "")
        append_assembly_sheet_to_workbook(raw_json, workbook, spec_name)

    output_buffer = io.BytesIO()
    workbook.save(output_buffer)
    output_buffer.seek(0)

    return StreamingResponse(
        output_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{bom_id}_excel_bundle.xlsx"'
        },
    )
