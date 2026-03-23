from __future__ import annotations

from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TITLE_FILL = PatternFill(fill_type="solid", start_color="E2E8F0", end_color="E2E8F0")
HEADER_FILL = PatternFill(fill_type="solid", start_color="CBD5E1", end_color="CBD5E1")
THIN_SIDE = Side(border_style="thin", color="94A3B8")
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER = Alignment(horizontal="center", vertical="center")

ASSEMBLY_COLUMNS = [
    ("부품 기준", "부품 기준"),
    ("요소작업", "요소작업"),
    ("OPTION", "OPTION"),
    ("작업자", "작업자"),
    ("no", "no"),
    ("동작요소", "동작요소"),
    ("반복횟수", "반복횟수"),
    ("SEC", "SEC"),
    ("TOTAL", "TOTAL"),
]

MERGE_TARGET_FIELDS = ["부품 기준", "요소작업", "OPTION"]
ROW_FILL_FIELDS = ["no"]
LEFT_ALIGN_FIELDS = {"요소작업", "동작요소"}
SUMMARY_FILL = PatternFill(fill_type="solid", start_color="BAE6FD", end_color="BAE6FD")
REQUIRED_FILL = PatternFill(fill_type="solid", start_color="FEF08A", end_color="FEF08A")
VALUE_FILL = PatternFill(fill_type="solid", start_color="BBF7D0", end_color="BBF7D0")
WASTE_FILL = PatternFill(fill_type="solid", start_color="FECACA", end_color="FECACA")


def _safe_sheet_title(workbook: Workbook, raw_title: str) -> str:
    base = (raw_title or "Assembly").strip() or "Assembly"
    invalid_chars = set('[]:*?/\\')
    sanitized = "".join("_" if ch in invalid_chars else ch for ch in base)[:31] or "Assembly"

    if sanitized not in workbook.sheetnames:
        return sanitized

    suffix = 2
    while True:
        candidate = f"{sanitized[:28]}_{suffix}"[:31]
        if candidate not in workbook.sheetnames:
            return candidate
        suffix += 1


def _normalize_value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def _to_float(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _normalize_text(value: Any) -> str:
    return str(value or "").strip()


def _resolve_row_fill(no_value: Any):
    normalized = _normalize_text(no_value)
    if "필비" in normalized:
        return REQUIRED_FILL
    if "가치" in normalized:
        return VALUE_FILL
    if "낭비" in normalized:
        return WASTE_FILL
    return None


def _build_effective_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    effective_rows = [dict(row) for row in rows if isinstance(row, dict)]
    fill_down_columns = ["부품 기준", "요소작업", "OPTION"]

    for column in fill_down_columns:
        last_value = None
        last_group_key = None

        for row in effective_rows:
            group_key = row.get("__groupKey")
            if group_key != last_group_key:
                last_group_key = group_key

            current_value = row.get(column)
            if current_value in (None, "") and last_value not in (None, ""):
                row[column] = last_value
            else:
                last_value = current_value

    return effective_rows


def _group_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    groups: List[Dict[str, Any]] = []
    current_group = None

    for row in rows:
        group_key = row.get("__groupKey") or f"group::{len(groups) + 1}"
        group_label = (
            row.get("__groupLabel")
            or row.get("__sequenceGroupLabel")
            or row.get("부품 기준")
            or "이름 없음"
        )

        if current_group is None or current_group["groupKey"] != group_key:
            current_group = {
                "groupKey": group_key,
                "groupLabel": str(group_label).strip() if group_label is not None else "이름 없음",
                "rows": [],
            }
            groups.append(current_group)

        current_group["rows"].append(row)

    return groups


def _merge_vertical_runs(
    ws,
    start_row: int,
    end_row: int,
    field_indexes: Dict[str, int],
    group_rows: List[Dict[str, Any]],
):
    if end_row < start_row or not group_rows:
        return

    part_col_idx = field_indexes["부품 기준"]
    part_ranges = []
    block_start_offset = 0

    for index in range(1, len(group_rows) + 1):
        current_key = _normalize_text(group_rows[index]["__partInstanceKey"]) if index < len(group_rows) else ""
        previous_key = _normalize_text(group_rows[index - 1]["__partInstanceKey"])

        if index < len(group_rows) and current_key == previous_key:
            continue

        block_start_row = start_row + block_start_offset
        block_end_row = start_row + index - 1
        part_ranges.append((block_start_row, block_end_row))

        part_value = ws.cell(row=block_start_row, column=part_col_idx).value
        if part_value not in (None, ""):
            if block_end_row > block_start_row:
                ws.merge_cells(
                    start_row=block_start_row,
                    start_column=part_col_idx,
                    end_row=block_end_row,
                    end_column=part_col_idx,
                )
            ws.cell(row=block_start_row, column=part_col_idx).alignment = CENTER

        block_start_offset = index

    for field in ("요소작업", "OPTION"):
        col_idx = field_indexes[field]
        for range_start, range_end in part_ranges:
            run_start = range_start
            run_value = ws.cell(row=range_start, column=col_idx).value

            for row_idx in range(range_start + 1, range_end + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value if row_idx <= range_end else None
                if row_idx <= range_end and cell_value == run_value and cell_value not in (None, ""):
                    continue

                current_end = row_idx - 1
                if run_value not in (None, ""):
                    if current_end - run_start >= 1:
                        ws.merge_cells(
                            start_row=run_start,
                            start_column=col_idx,
                            end_row=current_end,
                            end_column=col_idx,
                        )
                    ws.cell(row=run_start, column=col_idx).alignment = CENTER

                run_start = row_idx
                run_value = cell_value


def append_assembly_sheet_to_workbook(
    payload: Dict[str, Any],
    workbook: Workbook,
    spec_name: str,
):
    rows = payload.get("rows", []) if isinstance(payload, dict) else []
    effective_rows = _build_effective_rows(rows if isinstance(rows, list) else [])
    grouped_rows = _group_rows(effective_rows)

    sheet_title = _safe_sheet_title(workbook, "2. 작업시간 분석표")
    ws = workbook.create_sheet(title=sheet_title)

    column_count = len(ASSEMBLY_COLUMNS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "작업시간 분석표"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = CENTER
    title_cell.fill = TITLE_FILL
    title_cell.border = CELL_BORDER

    field_indexes = {}
    for col_idx, (_, header) in enumerate(ASSEMBLY_COLUMNS, start=1):
        field_indexes[ASSEMBLY_COLUMNS[col_idx - 1][0]] = col_idx
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = CENTER
        cell.fill = HEADER_FILL
        cell.border = CELL_BORDER

    current_row = 4
    last_column = len(ASSEMBLY_COLUMNS)

    for group in grouped_rows:
        group_start_row = current_row
        group_sec_total = 0.0
        group_total = 0.0

        for row in group["rows"]:
            row_fill = _resolve_row_fill(row.get("no"))
            for col_idx, (field, _) in enumerate(ASSEMBLY_COLUMNS, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = _normalize_value(row.get(field, ""))
                cell.border = CELL_BORDER
                cell.alignment = Alignment(
                    horizontal="left" if field in LEFT_ALIGN_FIELDS else "center",
                    vertical="center",
                    wrap_text=True,
                )
                if row_fill is not None and field in ROW_FILL_FIELDS:
                    cell.fill = row_fill

            sec_value = _to_float(row.get("SEC"))
            total_value = _to_float(row.get("TOTAL"))
            if total_value <= 0:
                total_value = sec_value * _to_float(row.get("반복횟수"))
            group_sec_total += sec_value
            group_total += total_value
            current_row += 1

        _merge_vertical_runs(
            ws,
            group_start_row,
            current_row - 1,
            field_indexes,
            group["rows"],
        )

        summary_label_end_col = max(1, field_indexes["반복횟수"] - 1)
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=summary_label_end_col,
        )
        summary_label_cell = ws.cell(row=current_row, column=1)
        summary_label_cell.value = group["groupLabel"]
        summary_label_cell.font = Font(bold=True)
        summary_label_cell.alignment = Alignment(horizontal="center", vertical="center")
        summary_label_cell.fill = SUMMARY_FILL
        summary_label_cell.border = CELL_BORDER

        summary_sec_cell = ws.cell(row=current_row, column=field_indexes["SEC"])
        summary_sec_cell.value = round(group_sec_total, 2)
        summary_sec_cell.font = Font(bold=True)
        summary_sec_cell.alignment = CENTER
        summary_sec_cell.fill = SUMMARY_FILL
        summary_sec_cell.border = CELL_BORDER

        summary_total_cell = ws.cell(row=current_row, column=field_indexes["TOTAL"])
        summary_total_cell.value = round(group_total, 2)
        summary_total_cell.font = Font(bold=True)
        summary_total_cell.alignment = CENTER
        summary_total_cell.fill = SUMMARY_FILL
        summary_total_cell.border = CELL_BORDER

        for col_idx in range(2, last_column):
            ws.cell(row=current_row, column=col_idx).fill = SUMMARY_FILL
            ws.cell(row=current_row, column=col_idx).border = CELL_BORDER

        current_row += 1

    widths = {
        "부품 기준": 28,
        "요소작업": 28,
        "OPTION": 34,
        "작업자": 10,
        "no": 10,
        "동작요소": 60,
        "반복횟수": 12,
        "SEC": 12,
        "TOTAL": 12,
    }

    for col_idx, (field, _) in enumerate(ASSEMBLY_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(field, 18)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(column_count)}{max(3, ws.max_row)}"
    ws.sheet_view.showGridLines = False

    return workbook
