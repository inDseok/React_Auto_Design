# backend/assembly/excel_db.py

from openpyxl import load_workbook
from pathlib import Path

EXCEL_PATH = Path("backend/작업시간분석표DB.xlsx")

def load_workbook_readonly():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError("작업시간분석표DB.xlsx 파일이 없습니다.")
    return load_workbook(EXCEL_PATH, data_only=True, read_only=True)
