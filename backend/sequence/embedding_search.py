from __future__ import annotations

import json
import logging
import os
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import numpy as np
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

EMBEDDING_MODEL = os.getenv("SEQUENCE_EMBEDDING_MODEL", "BAAI/bge-m3").strip() or "BAAI/bge-m3"
EMBEDDING_DEVICE = os.getenv("SEQUENCE_EMBEDDING_DEVICE", "auto").strip().lower() or "auto"
EMBEDDING_MAX_LENGTH = int(os.getenv("SEQUENCE_EMBEDDING_MAX_LENGTH", "512"))
EMBEDDING_SSL_VERIFY = os.getenv("SEQUENCE_EMBEDDING_SSL_VERIFY", "false").strip().lower() not in {
    "0",
    "false",
    "no",
    "off",
}
EMBEDDING_CACHE_DIR = Path(
    os.getenv("SEQUENCE_EMBEDDING_CACHE_DIR", "backend/data/hf_cache")
).resolve()
EMBEDDING_CACHE_DIR.mkdir(parents=True, exist_ok=True)
os.environ.setdefault("HF_HOME", str(EMBEDDING_CACHE_DIR))
os.environ.setdefault("HUGGINGFACE_HUB_CACHE", str(EMBEDDING_CACHE_DIR / "hub"))
os.environ.setdefault("HF_HUB_DISABLE_SYMLINKS_WARNING", "1")
os.environ.setdefault("HF_HUB_DISABLE_XET", "1")

EMBEDDING_DATA_DIR = Path(os.getenv("SEQUENCE_EMBEDDING_DATA_DIR", "backend/sequence/data")).resolve()
EMBEDDING_DATA_DIR.mkdir(parents=True, exist_ok=True)
EMBEDDING_INDEX_STEM = os.getenv("SEQUENCE_EMBEDDING_INDEX_STEM", "global_sequence_embedding_index").strip() or "global_sequence_embedding_index"
EMBEDDING_INDEX_VECTORS_PATH = EMBEDDING_DATA_DIR / f"{EMBEDDING_INDEX_STEM}.npz"
EMBEDDING_INDEX_META_PATH = EMBEDDING_DATA_DIR / f"{EMBEDDING_INDEX_STEM}.json"
EXCEL_DB_PATH = Path(os.getenv("SEQUENCE_EMBEDDING_SOURCE_XLSX", "backend/작업시간분석표DB.xlsx")).resolve()
EMBEDDING_INDEX_VERSION = 3


def _clean_text(value: Any) -> str:
    return " ".join(str(value or "").split()).strip()


def _normalize_key(value: Any) -> str:
    return (
        _clean_text(value)
        .upper()
        .replace("-", " ")
        .replace("_", " ")
        .replace("(", " ")
        .replace(")", " ")
        .replace("[", " ")
        .replace("]", " ")
    )


def _read_value(item: Any, key: str, default: Any = None) -> Any:
    if isinstance(item, dict):
        return item.get(key, default)
    return getattr(item, key, default)


def _join_fields(*values: Any) -> str:
    return " | ".join([_clean_text(value) for value in values if _clean_text(value)])


def _is_process_like_part_base(part_base: Any, process_label: Any = None) -> bool:
    normalized_part = _normalize_key(part_base)
    normalized_process = _normalize_key(process_label)

    if not normalized_part or set(normalized_part) <= {"-"}:
        return True
    if normalized_part in {"AAA", "SUB ASSY 결합".upper()}:
        return True
    if normalized_process and normalized_part == normalized_process:
        return True

    process_keywords = (
        "AIR BLOWING",
        "BLOWING",
        "BAR CODE",
        "BARCODE",
        "BAR-CODE",
        "HAND HEAT STACKING",
        "HEAT STACKING",
        "설비 작동",
        "설비 지그에서 취출",
        "설비 스위치",
        "스위치 ON",
        "부품안착동작",
        "부품 안착 동작",
        "부품취출동작",
        "부품 취출 동작",
        "포장지제거동작",
        "포장지 제거 동작",
        "검사동작",
        "검사 동작",
        "조립동작",
        "조립 동작",
        "동작",
        "취출작업",
        "취출 작업",
        "안착작업",
        "안착 작업",
        "결합작업",
        "결합 작업",
        "체결작업",
        "체결 작업",
        "연결작업",
        "연결 작업",
        "로딩작업",
        "로딩 작업",
        "UNLOADING",
        "LOADING",
        "WORK",
        "공정",
        "작업",
    )
    return any(_normalize_key(keyword) in normalized_part for keyword in process_keywords)


def _read_sequence_sheet_headers(ws, header_row: int) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is not None and _clean_text(value):
            headers[_clean_text(value)] = col
    return headers


def _resolve_sequence_sheet_columns(headers: Dict[str, int]) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    part_col = None
    process_col = None
    option_col = None
    for key, column in headers.items():
        normalized = key.replace(" ", "")
        if "부품" in normalized and part_col is None:
            part_col = column
        if ("요소작업" in normalized or "공정" in normalized) and process_col is None:
            process_col = column
        if "OPTION" in key.upper() and option_col is None:
            option_col = column
    return part_col, process_col, option_col


def _iter_excel_sequence_rows(excel_path: Path) -> List[Dict[str, str]]:
    if not excel_path.exists():
        return []
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    rows: List[Dict[str, str]] = []

    for worksheet in workbook.worksheets:
        headers = _read_sequence_sheet_headers(worksheet, 2)
        part_col, process_col, option_col = _resolve_sequence_sheet_columns(headers)
        if not part_col:
            continue

        current_part = ""
        current_process = ""
        current_option = ""

        for row in range(3, worksheet.max_row + 1):
            raw_part = worksheet.cell(row=row, column=part_col).value
            raw_process = worksheet.cell(row=row, column=process_col).value if process_col else None
            raw_option = worksheet.cell(row=row, column=option_col).value if option_col else None

            if _clean_text(raw_part):
                current_part = _clean_text(raw_part)
            if _clean_text(raw_process):
                current_process = _clean_text(raw_process)
            if _clean_text(raw_option):
                current_option = _clean_text(raw_option)

            if not current_part and not current_process:
                continue

            rows.append(
                {
                    "partBase": current_part,
                    "processLabel": current_process,
                    "option": current_option,
                    "sourceSheet": worksheet.title,
                }
            )

    return rows


def _build_global_embedding_documents(excel_path: Path) -> Dict[str, List[Dict[str, Any]]]:
    rows = _iter_excel_sequence_rows(excel_path)
    part_documents: List[Dict[str, Any]] = []
    process_documents: List[Dict[str, Any]] = []
    seen_parts = set()
    seen_processes = set()

    for row in rows:
        part_base = _clean_text(row.get("partBase"))
        process_label = _clean_text(row.get("processLabel"))
        option_value = _clean_text(row.get("option"))
        source_sheet = _clean_text(row.get("sourceSheet"))

        if part_base and not _is_process_like_part_base(part_base, process_label):
            part_key = (_normalize_key(part_base), source_sheet)
            if part_key not in seen_parts:
                seen_parts.add(part_key)
                part_documents.append(
                    {
                        "kind": "PART",
                        "partBase": part_base,
                        "sourceSheet": source_sheet,
                        "text": _join_fields("PART", part_base, source_sheet),
                    }
                )

        if process_label:
            process_key = (_normalize_key(process_label), _normalize_key(part_base), source_sheet)
            if process_key not in seen_processes:
                seen_processes.add(process_key)
                process_documents.append(
                    {
                        "kind": "PROCESS",
                        "processKey": process_label,
                        "label": process_label,
                        "partBase": part_base,
                        "sourceSheet": source_sheet,
                        "option": option_value,
                        "text": _join_fields("PROCESS", process_label, part_base, source_sheet),
                    }
                )

    return {
        "parts": part_documents,
        "processes": process_documents,
    }


@lru_cache(maxsize=1)
def _load_embedding_model() -> Tuple[Any, Any, Any]:
    try:
        import torch
        from transformers import AutoModel, AutoTokenizer
    except Exception as exc:
        raise RuntimeError("transformers/torch 패키지가 필요합니다.") from exc

    if not EMBEDDING_SSL_VERIFY:
        try:
            import requests
            import urllib3
            from huggingface_hub import configure_http_backend

            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

            def backend_factory() -> requests.Session:
                session = requests.Session()
                session.verify = False
                return session

            configure_http_backend(backend_factory=backend_factory)
        except Exception as exc:
            logger.warning("Failed to configure HuggingFace SSL verification override: %s", exc)

    resolved_device = "cuda" if EMBEDDING_DEVICE == "auto" else EMBEDDING_DEVICE
    if resolved_device == "cuda":
        try:
            import torch

            if not torch.cuda.is_available():
                resolved_device = "cpu"
        except Exception:
            resolved_device = "cpu"
    if resolved_device == "auto":
        resolved_device = "cpu"

    tokenizer = AutoTokenizer.from_pretrained(EMBEDDING_MODEL)
    model = AutoModel.from_pretrained(EMBEDDING_MODEL)
    model.to(resolved_device)
    model.eval()
    return tokenizer, model, torch


def _mean_pool(last_hidden_state: Any, attention_mask: Any) -> Any:
    input_mask_expanded = attention_mask.unsqueeze(-1).expand(last_hidden_state.size()).float()
    return (last_hidden_state * input_mask_expanded).sum(1) / input_mask_expanded.sum(1).clamp(min=1e-9)


def _embed_texts(texts: Sequence[str]) -> np.ndarray:
    tokenizer, model, torch = _load_embedding_model()
    vectors: List[np.ndarray] = []
    batch_size = int(os.getenv("SEQUENCE_EMBEDDING_BATCH_SIZE", "16"))

    for start in range(0, len(texts), batch_size):
        batch = list(texts[start : start + batch_size])
        encoded = tokenizer(
            batch,
            padding=True,
            truncation=True,
            max_length=EMBEDDING_MAX_LENGTH,
            return_tensors="pt",
        )
        target_device = next(model.parameters()).device
        encoded = {key: value.to(target_device) for key, value in encoded.items()}
        with torch.inference_mode():
            output = model(**encoded)
            pooled = _mean_pool(output.last_hidden_state, encoded["attention_mask"])
            pooled = torch.nn.functional.normalize(pooled, p=2, dim=1)
        vectors.append(pooled.cpu().numpy())

    if not vectors:
        return np.empty((0, 0), dtype=np.float32)
    return np.vstack(vectors).astype(np.float32)


def build_and_save_global_embedding_index(
    *,
    excel_path: Optional[Path] = None,
    vectors_path: Optional[Path] = None,
    meta_path: Optional[Path] = None,
) -> Dict[str, Any]:
    resolved_excel_path = (excel_path or EXCEL_DB_PATH).resolve()
    resolved_vectors_path = (vectors_path or EMBEDDING_INDEX_VECTORS_PATH).resolve()
    resolved_meta_path = (meta_path or EMBEDDING_INDEX_META_PATH).resolve()

    documents = _build_global_embedding_documents(resolved_excel_path)
    part_texts = [str(item.get("text") or "") for item in documents["parts"]]
    process_texts = [str(item.get("text") or "") for item in documents["processes"]]
    part_vectors = _embed_texts(part_texts) if part_texts else np.empty((0, 0), dtype=np.float32)
    process_vectors = _embed_texts(process_texts) if process_texts else np.empty((0, 0), dtype=np.float32)

    resolved_vectors_path.parent.mkdir(parents=True, exist_ok=True)
    resolved_meta_path.parent.mkdir(parents=True, exist_ok=True)
    np.savez_compressed(
        resolved_vectors_path,
        part_vectors=part_vectors,
        process_vectors=process_vectors,
    )
    resolved_meta_path.write_text(
        json.dumps(
            {
                "indexVersion": EMBEDDING_INDEX_VERSION,
                "sourceExcelPath": str(resolved_excel_path),
                "sourceExcelMtime": resolved_excel_path.stat().st_mtime if resolved_excel_path.exists() else 0.0,
                "parts": documents["parts"],
                "processes": documents["processes"],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    return {
        "parts": documents["parts"],
        "processes": documents["processes"],
        "part_vectors": part_vectors,
        "process_vectors": process_vectors,
    }


@lru_cache(maxsize=1)
def load_or_build_global_embedding_index() -> Dict[str, Any]:
    if (
        not EMBEDDING_INDEX_VECTORS_PATH.exists()
        or not EMBEDDING_INDEX_META_PATH.exists()
        or (EXCEL_DB_PATH.exists() and EMBEDDING_INDEX_META_PATH.stat().st_mtime < EXCEL_DB_PATH.stat().st_mtime)
    ):
        return build_and_save_global_embedding_index()

    metadata = json.loads(EMBEDDING_INDEX_META_PATH.read_text(encoding="utf-8"))
    if int(metadata.get("indexVersion") or 0) != EMBEDDING_INDEX_VERSION:
        return build_and_save_global_embedding_index()
    vectors_payload = np.load(EMBEDDING_INDEX_VECTORS_PATH)
    return {
        "parts": list(metadata.get("parts") or []),
        "processes": list(metadata.get("processes") or []),
        "part_vectors": np.asarray(vectors_payload["part_vectors"], dtype=np.float32),
        "process_vectors": np.asarray(vectors_payload["process_vectors"], dtype=np.float32),
    }


def _rank_global_vectors(query: str, vectors: np.ndarray, metadata: List[Dict[str, Any]], *, limit: int) -> List[Dict[str, Any]]:
    if not query or vectors.size == 0 or not metadata:
        return []

    query_matrix = _embed_texts([query])
    if query_matrix.shape[0] == 0:
        return []

    scores = vectors @ query_matrix[0]
    ranked_indices = np.argsort(-scores)[:limit]
    return [
        {
            **dict(metadata[int(index)]),
            "embeddingScore": float(scores[int(index)]),
        }
        for index in ranked_indices
    ]


def _candidate_part_keys(part: Any) -> List[str]:
    keys = [
        _normalize_key(_read_value(part, "partBase")),
        _normalize_key(_read_value(part, "partName")),
        _normalize_key(_read_value(part, "partId")),
        _normalize_key(_read_value(part, "nodeName")),
    ]
    return [key for key in keys if key]


def _match_candidate_parts(
    global_part_hits: List[Dict[str, Any]],
    candidate_parts: Sequence[Any],
    *,
    limit: int,
) -> List[Dict[str, Any]]:
    if not global_part_hits or not candidate_parts:
        return []

    by_key: Dict[str, List[Any]] = {}
    for part in candidate_parts or []:
        for key in _candidate_part_keys(part):
            by_key.setdefault(key, []).append(part)

    ranked: List[Dict[str, Any]] = []
    seen = set()
    for hit in global_part_hits:
        hit_key = _normalize_key(hit.get("partBase"))
        if not hit_key:
            continue
        candidates = list(by_key.get(hit_key) or [])
        if not candidates:
            continue
        preferred_sheet = _clean_text(hit.get("sourceSheet"))
        candidates.sort(
            key=lambda item: (
                0 if _clean_text(_read_value(item, "sourceSheet")) == preferred_sheet else 1,
                str(_read_value(item, "nodeName") or _read_value(item, "partBase") or ""),
            )
        )
        for candidate in candidates:
            dedupe_key = (
                _clean_text(_read_value(candidate, "nodeName")),
                _clean_text(_read_value(candidate, "partBase")),
                _clean_text(_read_value(candidate, "sourceSheet")),
            )
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            ranked.append(
                {
                    "kind": "PART",
                    "item": candidate,
                    "embeddingScore": float(hit.get("embeddingScore") or 0),
                    "text": hit.get("text") or "",
                }
            )
            if len(ranked) >= limit:
                return ranked
    return ranked


def _match_process_templates(
    global_process_hits: List[Dict[str, Any]],
    process_templates: Sequence[Any],
    *,
    limit: int,
) -> List[Dict[str, Any]]:
    if not global_process_hits or not process_templates:
        return []

    by_key: Dict[Tuple[str, str, str], Any] = {}
    for process in process_templates or []:
        key = (
            _clean_text(_read_value(process, "processKey")),
            _clean_text(_read_value(process, "partBase")),
            _clean_text(_read_value(process, "sourceSheet")),
        )
        if key[0]:
            by_key[key] = process

    ranked: List[Dict[str, Any]] = []
    seen = set()
    for hit in global_process_hits:
        exact_key = (
            _clean_text(hit.get("processKey")),
            _clean_text(hit.get("partBase")),
            _clean_text(hit.get("sourceSheet")),
        )
        matched = by_key.get(exact_key)
        if matched is None:
            fallback = next(
                (
                    process
                    for process in process_templates or []
                    if _clean_text(_read_value(process, "processKey")) == exact_key[0]
                ),
                None,
            )
            matched = fallback
        if matched is None:
            continue
        dedupe_key = (
            _clean_text(_read_value(matched, "processKey")),
            _clean_text(_read_value(matched, "partBase")),
            _clean_text(_read_value(matched, "sourceSheet")),
        )
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        ranked.append(
            {
                "kind": "PROCESS",
                "item": matched,
                "embeddingScore": float(hit.get("embeddingScore") or 0),
                "text": hit.get("text") or "",
            }
        )
        if len(ranked) >= limit:
            return ranked
    return ranked


def search_chat_candidates_with_bge_m3(
    message: str,
    *,
    candidate_parts: Sequence[Any],
    process_templates: Sequence[Any],
    part_limit: int,
    process_limit: int,
) -> Dict[str, List[Dict[str, Any]]]:
    index = load_or_build_global_embedding_index()
    global_part_hits = _rank_global_vectors(
        message,
        index.get("part_vectors"),
        list(index.get("parts") or []),
        limit=max(part_limit * 4, 40),
    )
    global_process_hits = _rank_global_vectors(
        message,
        index.get("process_vectors"),
        list(index.get("processes") or []),
        limit=max(process_limit * 4, 80),
    )

    return {
        "parts": _match_candidate_parts(global_part_hits, candidate_parts, limit=part_limit),
        "processes": _match_process_templates(global_process_hits, process_templates, limit=process_limit),
    }
