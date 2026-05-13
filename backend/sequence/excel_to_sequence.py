from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


PROCESS_KEYWORDS = (
    "T/SCREW",
    "SCREW",
    "체결",
    "작동",
    "리딩",
    "마킹",
    "연결작업",
    "취출작업",
    "압입",
    "도포",
    "BLOW",
    "BLOWING",
    "검사",
    "버튼",
    "커넥터",
    "에어",
)


def infer_node_type(label: str, process_name: str, option_text: str) -> str:
    haystack = f"{label} {process_name} {option_text}".upper()
    if any(keyword.upper() in haystack for keyword in PROCESS_KEYWORDS):
        return "PROCESS"
    return "PART"


def make_process_key(label: str) -> str:
    return label.strip()


def convert_excel_to_sequence(
    source_path: Path,
    *,
    bom_id: str = "excel-import",
    spec: str = "Sheet1",
    sheet_name: Optional[str] = None,
) -> Dict[str, Any]:
    workbook = load_workbook(source_path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    nodes: List[Dict[str, Any]] = []
    edges: List[Dict[str, Any]] = []
    groups: List[Dict[str, Any]] = []
    worker_groups: List[Dict[str, Any]] = []

    current_group_label = "그룹 1"
    current_group_node_ids: List[str] = []
    current_group_index = 1
    group_y = 80
    node_index = 1
    last_node_id: Optional[str] = None

    def flush_group() -> None:
        nonlocal current_group_node_ids, current_group_label, group_y, current_group_index, last_node_id
        if not current_group_node_ids:
            return
        groups.append(
            {
                "id": f"grp-excel-{current_group_index}",
                "label": current_group_label,
                "nodeIds": current_group_node_ids[:],
                "skippedAutoEdgeIds": [],
            }
        )
        current_group_index += 1
        current_group_node_ids = []
        group_y += 220
        last_node_id = None

    for row in worksheet.iter_rows(values_only=True):
        label = str(row[0]).strip() if len(row) > 0 and row[0] not in (None, "") else ""
        process_name = str(row[1]).strip() if len(row) > 1 and row[1] not in (None, "") else ""
        option_text = str(row[2]).strip() if len(row) > 2 and row[2] not in (None, "") else ""
        worker = str(row[3]).strip() if len(row) > 3 and row[3] not in (None, "") else "1"

        if label:
            flush_group()
            current_group_label = label

        if not label:
            continue

        node_id = f"N-EXCEL-{node_index}"
        node_type = infer_node_type(label, process_name, option_text)
        x = 60 + (len(current_group_node_ids) * 230)
        y = group_y

        if node_type == "PART":
            node_data = {
                "partId": label,
                "partName": label,
                "nodeName": label,
                "inhouse": True,
                "partBase": label,
                "sourceSheet": worksheet.title,
                "option": option_text,
                "statusLabel": process_name,
                "label": label,
                "worker": worker,
            }
        else:
            node_data = {
                "processKey": make_process_key(label),
                "processType": "STANDARD",
                "label": label,
                "partBase": process_name or label,
                "sourceSheet": worksheet.title,
                "option": option_text,
                "statusLabel": process_name,
                "worker": worker,
            }

        nodes.append(
            {
                "id": node_id,
                "type": node_type,
                "position": {"x": x, "y": y},
                "data": node_data,
            }
        )
        current_group_node_ids.append(node_id)

        if last_node_id:
            edges.append(
                {
                    "id": f"E-{last_node_id}-{node_id}",
                    "source": last_node_id,
                    "target": node_id,
                    "type": "smoothstep",
                    "sourceHandle": "out",
                    "targetHandle": "in",
                    "data": {},
                }
        )

        last_node_id = node_id
        node_index += 1

    flush_group()

    return {
        "bomId": bom_id,
        "spec": spec,
        "nodes": nodes,
        "edges": edges,
        "groups": groups,
        "workerGroups": worker_groups,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source")
    parser.add_argument("output")
    parser.add_argument("--bom-id", default="excel-import")
    parser.add_argument("--spec", default="Sheet1")
    parser.add_argument("--sheet-name", default=None)
    args = parser.parse_args()

    payload = convert_excel_to_sequence(
        Path(args.source),
        bom_id=args.bom_id,
        spec=args.spec,
        sheet_name=args.sheet_name,
    )
    Path(args.output).write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
