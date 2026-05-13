from __future__ import annotations

import json
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl import load_workbook

from backend.Assembly.auto_match import (
    COMBINED_THRESHOLD,
    combined_score,
    jw_score,
    load_db_rows,
    match_one_best,
    rf_score,
)

from .models import GraphIndex, SequenceStep, WindowDocument

PROCESS_HINT_KEYWORDS = (
    "로딩",
    "언로딩",
    "안착",
    "취출",
    "이동",
    "검사",
    "체결",
    "조립",
    "작동",
    "연결",
    "도포",
    "삽입",
    "SCREW",
    "BLOWING",
    "LOADING",
    "UNLOADING",
    "INSPECTION",
    "TEST",
)

PART_STOPWORDS = (
    "이동",
    "검사",
    "설비 작동",
)

PROCESS_STOPWORDS = (
    "설비 작동",
    "이동",
)

EXCEL_DB_PATH = Path(__file__).resolve().parent.parent / "작업시간분석표DB.xlsx"
PROCESS_COMBINED_THRESHOLD = 82.0


def _iter_sequence_files(sequence_dir: Path) -> Iterable[Path]:
    return sorted(path for path in sequence_dir.glob("*_sequence.json") if path.is_file())


def _sort_nodes(nodes: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return sorted(
        nodes,
        key=lambda node: (
            float((node.get("position") or {}).get("y", 0) or 0),
            float((node.get("position") or {}).get("x", 0) or 0),
            str(node.get("id") or ""),
        ),
    )


def _normalize_text(value: Any) -> str:
    return " ".join(str(value or "").split()).strip()


@lru_cache(maxsize=1)
def _get_db_rows_and_choices() -> tuple[list[dict[str, Any]], list[str]]:
    if not EXCEL_DB_PATH.exists():
        return [], []
    return load_db_rows(EXCEL_DB_PATH)


@lru_cache(maxsize=1)
def _get_exact_db_part_map() -> dict[str, str]:
    db_rows, _ = _get_db_rows_and_choices()
    exact_map: dict[str, str] = {}
    for row in db_rows:
        normalized = _normalize_token_text(row.get("db_part_raw"))
        raw = _normalize_text(row.get("db_part_raw"))
        if normalized and raw and normalized not in exact_map:
            exact_map[normalized] = raw
    return exact_map


@lru_cache(maxsize=1)
def _get_db_process_labels() -> list[str]:
    if not EXCEL_DB_PATH.exists():
        return []

    workbook = load_workbook(EXCEL_DB_PATH, read_only=True, data_only=True)
    labels: list[str] = []
    seen = set()

    for worksheet in workbook.worksheets:
        headers: dict[str, int] = {}
        for col in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=2, column=col).value
            if value:
                headers[str(value).strip()] = col

        process_col = None
        for header, col in headers.items():
            normalized = str(header).replace(" ", "")
            if "요소작업" in normalized or "공정" in normalized:
                process_col = col
                break

        if not process_col:
            continue

        for row in range(3, worksheet.max_row + 1):
            value = worksheet.cell(row=row, column=process_col).value
            raw = _normalize_text(value)
            if not raw or raw in seen:
                continue
            seen.add(raw)
            labels.append(raw)

    return labels


@lru_cache(maxsize=1)
def _get_exact_db_process_map() -> dict[str, str]:
    exact_map: dict[str, str] = {}
    for raw in _get_db_process_labels():
        normalized = _normalize_token_text(raw)
        if normalized and normalized not in exact_map:
            exact_map[normalized] = raw
    return exact_map


def _normalize_token_text(value: Any) -> str:
    text = _normalize_text(value).upper()
    if not text:
        return ""
    for old, new in (("-", " "), ("_", " "), ("(", " "), (")", " "), ("/", " ")):
        text = text.replace(old, new)
    return " ".join(text.split()).strip()


def _matches_stopword(value: str, candidates: Sequence[str]) -> bool:
    normalized = _normalize_token_text(value)
    if not normalized:
        return False
    return any(_normalize_token_text(candidate) in normalized for candidate in candidates if candidate)


def _classify_step_type(node: Dict[str, Any]) -> str:
    node_type = _normalize_text(node.get("type")).upper()
    data = node.get("data") or {}

    if node_type == "PROCESS":
        return "PROCESS"
    if node_type == "PART":
        return "PART"

    process_key = _normalize_text(data.get("processKey"))
    if process_key:
        return "PROCESS"

    part_base = _normalize_text(data.get("partBase"))
    if part_base:
        return "PART"

    process_like_texts = [
        _normalize_text(data.get("label")),
        _normalize_text(data.get("nodeName")),
        _normalize_text(data.get("partName")),
        _normalize_text(data.get("partId")),
    ]
    normalized_texts = [_normalize_token_text(value) for value in process_like_texts if _normalize_text(value)]
    if any(
        keyword in normalized_text
        for normalized_text in normalized_texts
        for keyword in (_normalize_token_text(item) for item in PROCESS_HINT_KEYWORDS)
    ):
        return "PROCESS"

    return "PART"


def _build_part_identity(data: Dict[str, Any]) -> tuple[str, str]:
    raw_candidates = [
        _normalize_text(data.get("partBase")),
        _normalize_text(data.get("partId")),
        _normalize_text(data.get("nodeName")),
        _normalize_text(data.get("partName")),
        _normalize_text(data.get("label")),
    ]
    raw_candidates = [value for value in raw_candidates if value]
    if not raw_candidates:
        return "", ""

    exact_map = _get_exact_db_part_map()
    for candidate in raw_candidates:
        normalized_candidate = _normalize_token_text(candidate)
        canonical = exact_map.get(normalized_candidate)
        if canonical:
            return canonical, canonical

    db_rows, db_choices = _get_db_rows_and_choices()
    if db_rows and db_choices:
        for candidate in raw_candidates:
            best = match_one_best(candidate, db_rows, db_choices)
            if best and float(best.get("score_combined") or 0) >= COMBINED_THRESHOLD:
                canonical = _normalize_text(best.get("db_part_raw"))
                if canonical:
                    return canonical, canonical

    return "", ""


def _build_process_identity(data: Dict[str, Any]) -> tuple[str, str]:
    raw_candidates = [
        _normalize_text(data.get("processKey")),
        _normalize_text(data.get("label")),
        _normalize_text(data.get("nodeName")),
        _normalize_text(data.get("statusLabel")),
    ]
    raw_candidates = [value for value in raw_candidates if value]
    if not raw_candidates:
        return "", ""

    exact_map = _get_exact_db_process_map()
    for candidate in raw_candidates:
        normalized_candidate = _normalize_token_text(candidate)
        canonical = exact_map.get(normalized_candidate)
        if canonical:
            return canonical, canonical

    process_labels = _get_db_process_labels()
    if process_labels:
        for candidate in raw_candidates:
            candidate_normalized = _normalize_token_text(candidate)
            best_label = ""
            best_score = -1.0
            for process_label in process_labels:
                process_normalized = _normalize_token_text(process_label)
                rf = rf_score(candidate_normalized, process_normalized)
                jw = jw_score(candidate_normalized, process_normalized)
                score = combined_score(rf, jw)
                if score > best_score:
                    best_score = score
                    best_label = process_label
            if best_label and best_score >= PROCESS_COMBINED_THRESHOLD:
                canonical = _normalize_text(best_label)
                return canonical, canonical

    return "", ""


def _should_keep_step(step_type: str, key: str, label: str) -> bool:
    if not key:
        return False
    if step_type == "PART" and _matches_stopword(f"{key} {label}", PART_STOPWORDS):
        return False
    if step_type == "PROCESS" and _matches_stopword(f"{key} {label}", PROCESS_STOPWORDS):
        return False
    return True


def _step_from_node(node: Dict[str, Any], index: int) -> SequenceStep:
    node_type = _classify_step_type(node)
    data = node.get("data") or {}
    reason = _normalize_text(data.get("statusLabel") or data.get("option"))

    if node_type == "PROCESS":
        process_key, label = _build_process_identity(data)
        return SequenceStep(
            type="PROCESS",
            key=process_key,
            label=label,
            reason=reason,
            source_id=_normalize_text(node.get("id")),
            index=index,
        )

    part_id, label = _build_part_identity(data)
    return SequenceStep(
        type="PART",
        key=part_id,
        label=label,
        reason=reason,
        source_id=_normalize_text(node.get("id")),
        index=index,
    )


def _extract_steps(sequence_payload: Dict[str, Any]) -> List[SequenceStep]:
    sorted_nodes = _sort_nodes(sequence_payload.get("nodes") or [])
    steps: List[SequenceStep] = []
    for index, node in enumerate(sorted_nodes):
        step = _step_from_node(node, index)
        if _should_keep_step(step.type, step.key, step.label):
            steps.append(step)
    return steps


def _build_part_windows(
    steps: Sequence[SequenceStep],
    *,
    window_size: int = 3,
    stride: int = 2,
) -> List[List[SequenceStep]]:
    part_indexes = [index for index, step in enumerate(steps) if step.type == "PART" and step.key]
    windows: List[List[SequenceStep]] = []

    if len(part_indexes) < window_size:
        return windows

    for start in range(0, len(part_indexes) - window_size + 1, stride):
        selected_part_indexes = part_indexes[start : start + window_size]
        if len(selected_part_indexes) < window_size:
            continue
        left = selected_part_indexes[0]
        right = selected_part_indexes[-1]
        snippet = [step for step in steps[left : right + 1] if step.key]
        if len([step for step in snippet if step.type == "PART"]) == window_size:
            windows.append(snippet)
    return windows


def _transition_pairs(snippet: Sequence[SequenceStep]) -> List[str]:
    transitions: List[str] = []
    for left, right in zip(snippet, snippet[1:]):
        transitions.append(f"{left.type}:{left.key} -> {right.type}:{right.key}")
    return transitions


def _add_count(bucket: Dict[str, Dict[str, int]], left_key: str, right_key: str) -> None:
    bucket.setdefault(left_key, {})
    bucket[left_key][right_key] = bucket[left_key].get(right_key, 0) + 1


def _build_documents_for_file(sequence_path: Path) -> List[WindowDocument]:
    payload = json.loads(sequence_path.read_text(encoding="utf-8"))
    steps = _extract_steps(payload)
    windows = _build_part_windows(steps)
    documents: List[WindowDocument] = []

    for window_index, snippet in enumerate(windows, start=1):
        part_steps = [step for step in snippet if step.type == "PART"]
        process_steps = [step for step in snippet if step.type == "PROCESS"]
        source_name = sequence_path.name.replace("_sequence.json", "")
        documents.append(
            WindowDocument(
                doc_id=f"{source_name}-w{window_index}",
                source_file=str(sequence_path),
                source_name=source_name,
                window_index=window_index,
                anchor_part_ids=[step.key for step in part_steps],
                anchor_labels=[step.label for step in part_steps],
                process_labels=[step.key for step in process_steps],
                snippet=[step.to_dict() for step in snippet],
                transitions=_transition_pairs(snippet),
            )
        )

    return documents


def build_index_from_sequence_dir(sequence_dir: Path) -> GraphIndex:
    index = GraphIndex()

    for sequence_path in _iter_sequence_files(sequence_dir):
        documents = _build_documents_for_file(sequence_path)
        index.documents.extend(documents)

        for document in documents:
            for transition in document.transitions:
                index.transition_counts[transition] = index.transition_counts.get(transition, 0) + 1

            snippet_steps = [SequenceStep(**item) for item in document.snippet]
            for left, right in zip(snippet_steps, snippet_steps[1:]):
                if left.type == "PART" and right.type == "PROCESS":
                    _add_count(index.part_to_process_counts, left.key, right.key)
                if left.type == "PROCESS" and right.type == "PART":
                    _add_count(index.process_to_part_counts, left.key, right.key)

    return index


def write_index(index: GraphIndex, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(index.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
