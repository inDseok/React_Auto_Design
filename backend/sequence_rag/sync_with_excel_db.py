from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Sequence, Set

from openpyxl import load_workbook

from .neo4j_export import build_cypher_from_index


STOP_TOKENS = {
    "LH",
    "RH",
    "STD",
    "ECE",
    "LHD",
    "RHD",
    "LD",
    "HD",
    "EC",
    "LEFT",
    "RIGHT",
    "TYPE",
    "TYP",
    "ASSY",
    "S/A",
}

SYN_MAP = {
    "BPR": "BUMPER",
    "BUMPER": "BUMPER",
    "BRKT": "BRACKET",
    "BRACKET": "BRACKET",
    "HSG": "HOUSING",
    "HOUSING": "HOUSING",
    "INR": "INNER",
    "INNER": "INNER",
    "OTR": "MAIN",
    "OUTER": "MAIN",
    "EXTN": "EXTENSION",
    "EXT": "EXTENSION",
    "EXTENSION": "EXTENSION",
    "WIRG": "WIRING",
    "WIRING": "WIRING",
    "WIRE": "WIRING",
    "TURN": "T/SIG",
    "SIGNAL": "",
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip()
    if not text:
        return ""

    text = re.sub(r"\[[^\]]*\]", " ", text)
    text = text.replace("-", " ").replace("_", " ")
    text = text.upper()
    text = re.sub(r"[^A-Z0-9/ ]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()

    tokens: List[str] = []
    for token in text.split(" "):
        if not token:
            continue
        if token in STOP_TOKENS:
            break
        token = SYN_MAP.get(token, token)
        if token:
            tokens.append(token)

    return " ".join(tokens).strip()


def load_allowed_keys(excel_path: Path) -> Set[str]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    allowed: Set[str] = set()

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            raw = str((row[0] if row else "") or "").strip()
            if not raw:
                continue
            normalized = normalize_text(raw)
            if normalized:
                allowed.add(normalized)

    return allowed


def _transition_pairs(snippet: Sequence[Dict[str, Any]]) -> List[str]:
    transitions: List[str] = []
    for left, right in zip(snippet, snippet[1:]):
        left_type = str(left.get("type") or "").upper()
        right_type = str(right.get("type") or "").upper()
        left_key = str(left.get("key") or "").strip()
        right_key = str(right.get("key") or "").strip()
        if left_key and right_key:
            transitions.append(f"{left_type}:{left_key} -> {right_type}:{right_key}")
    return transitions


def _add_count(bucket: Dict[str, Dict[str, int]], left_key: str, right_key: str) -> None:
    bucket.setdefault(left_key, {})
    bucket[left_key][right_key] = bucket[left_key].get(right_key, 0) + 1


def filter_index_payload(
    index_payload: Dict[str, Any],
    *,
    allowed_keys: Set[str],
    min_steps: int = 1,
    require_part: bool = True,
) -> Dict[str, Any]:
    source_buckets: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for document in index_payload.get("documents", []):
        source_name = str(document.get("source_name") or "")
        source_buckets[source_name].append(document)

    filtered_documents: List[Dict[str, Any]] = []
    transition_counts: Dict[str, int] = {}
    part_to_process_counts: Dict[str, Dict[str, int]] = {}
    process_to_part_counts: Dict[str, Dict[str, int]] = {}

    for source_name, documents in sorted(source_buckets.items()):
        window_index = 1
        for document in documents:
            snippet = document.get("snippet") or []
            filtered_snippet = [
                step
                for step in snippet
                if normalize_text(step.get("key") or "") in allowed_keys
            ]

            if len(filtered_snippet) < min_steps:
                continue

            if require_part and not any(
                str(step.get("type") or "").upper() == "PART"
                for step in filtered_snippet
            ):
                continue

            part_steps = [
                step for step in filtered_snippet if str(step.get("type") or "").upper() == "PART"
            ]
            process_steps = [
                step for step in filtered_snippet if str(step.get("type") or "").upper() == "PROCESS"
            ]
            transitions = _transition_pairs(filtered_snippet)

            rebuilt_document = {
                **document,
                "doc_id": f"{source_name}-w{window_index}",
                "window_index": window_index,
                "anchor_part_ids": [str(step.get("key") or "") for step in part_steps if step.get("key")],
                "anchor_labels": [str(step.get("label") or "") for step in part_steps if step.get("label")],
                "process_labels": [str(step.get("key") or "") for step in process_steps if step.get("key")],
                "snippet": filtered_snippet,
                "transitions": transitions,
            }
            filtered_documents.append(rebuilt_document)
            window_index += 1

            for transition in transitions:
                transition_counts[transition] = transition_counts.get(transition, 0) + 1

            for left, right in zip(filtered_snippet, filtered_snippet[1:]):
                left_type = str(left.get("type") or "").upper()
                right_type = str(right.get("type") or "").upper()
                left_key = str(left.get("key") or "").strip()
                right_key = str(right.get("key") or "").strip()
                if not left_key or not right_key:
                    continue
                if left_type == "PART" and right_type == "PROCESS":
                    _add_count(part_to_process_counts, left_key, right_key)
                if left_type == "PROCESS" and right_type == "PART":
                    _add_count(process_to_part_counts, left_key, right_key)

    return {
        "documents": filtered_documents,
        "transitionCounts": transition_counts,
        "partToProcessCounts": part_to_process_counts,
        "processToPartCounts": process_to_part_counts,
    }


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--excel-db",
        default=str(base_dir.parent / "작업시간분석표DB.xlsx"),
    )
    parser.add_argument(
        "--index",
        default=str(base_dir / "data" / "graph_index.json"),
    )
    parser.add_argument(
        "--output-index",
        default=str(base_dir / "data" / "graph_index.json"),
    )
    parser.add_argument(
        "--output-cypher",
        default=str(base_dir / "data" / "neo4j_import.cypher"),
    )
    parser.add_argument("--min-steps", type=int, default=1)
    parser.add_argument(
        "--allow-process-only",
        action="store_true",
        help="Keep windows that no longer contain a part after filtering.",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel_db)
    index_path = Path(args.index)
    output_index_path = Path(args.output_index)
    output_cypher_path = Path(args.output_cypher)

    payload = json.loads(index_path.read_text(encoding="utf-8"))
    original_documents = len(payload.get("documents") or [])
    allowed_keys = load_allowed_keys(excel_path)
    filtered_payload = filter_index_payload(
        payload,
        allowed_keys=allowed_keys,
        min_steps=max(args.min_steps, 1),
        require_part=not args.allow_process_only,
    )

    output_index_path.parent.mkdir(parents=True, exist_ok=True)
    output_index_path.write_text(
        json.dumps(filtered_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    cypher = build_cypher_from_index(filtered_payload)
    output_cypher_path.parent.mkdir(parents=True, exist_ok=True)
    output_cypher_path.write_text(cypher, encoding="utf-8")

    print(f"allowed_keys={len(allowed_keys)}")
    print(f"documents_before={original_documents}")
    print(f"documents_after={len(filtered_payload.get('documents') or [])}")
    print(f"graph_index_written={output_index_path}")
    print(f"cypher_written={output_cypher_path}")


if __name__ == "__main__":
    main()
