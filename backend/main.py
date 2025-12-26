from __future__ import annotations

from backend.session_excel import get_or_create_sid, SessionState
from typing import Dict, List, Optional, Any

import json
from uuid import uuid4

from fastapi import FastAPI, HTTPException, Request, Response, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles

from pydantic import BaseModel
from pathlib import Path

from backend.models import SubNodePatch, SubTree, MoveNodeRequest
from backend.bom_service import create_bom_run, DATA_DIR
from backend.utills import find_node_by_id, load_tree_json, save_tree_json
from backend.excel_loader import build_tree_from_sheet
from fastapi import Cookie
from typing import Optional
from openpyxl import load_workbook


app = FastAPI()

templates = Jinja2Templates(directory="frontend/template")
app.mount("/static", StaticFiles(directory="frontend/static"), name="static")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def load_session_state():
    if not SESSION_STORE_PATH.exists():
        return {}
    try:
        return json.loads(SESSION_STORE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}

def save_session_state():
    SESSION_STORE_PATH.write_text(
        json.dumps(SESSION_STATE, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

DATA_DIR = Path("backend/data")

SESSION_STORE_PATH = DATA_DIR / "session_state.json"
SESSION_STATE: Dict[str, Dict[str, Optional[str]]] = load_session_state()

def read_bom_meta(root: Path) -> dict:
    meta_path = root / "bom_meta.json"
    if not meta_path.exists():
        return {}
    try:
        return json.loads(meta_path.read_text(encoding="utf-8"))
    except Exception:
        return {}

@app.get("/favicon.ico", include_in_schema=False)
async def favicon():
    return Response(status_code=204)


@app.post("/api/state", response_model=SessionState)
def set_state(payload: dict, request: Request, response: Response):
    sid = get_or_create_sid(request, response)

    # 기존 상태 가져오기 (없으면 빈 dict)
    prev = SESSION_STATE.get(sid, {})
    print("[SESSION_STATE]", SESSION_STATE)
    # merge 규칙:
    # payload에 있는 값만 갱신
    # payload에 없으면 기존 값 유지
    next_state = {
        "bom_id": payload["bom_id"] if "bom_id" in payload else prev.get("bom_id"),
        "spec": payload["spec"] if "spec" in payload else prev.get("spec"),
        "selected_id": payload["selected_id"] if "selected_id" in payload else prev.get("selected_id"),
    }

    SESSION_STATE[sid] = next_state
    save_session_state()

    return SessionState(**next_state)


@app.get("/api/state", response_model=SessionState)
def get_state(request: Request, response: Response):
    sid = get_or_create_sid(request, response)
    state = SESSION_STATE.get(sid, {})
    print("[SESSION_STATE]", SESSION_STATE)
    return SessionState(
        bom_id=state.get("bom_id"),
        spec=state.get("spec"),
        selected_id=state.get("selected_id"),
    )



@app.post("/api/bom/upload")
async def upload_bom(file: UploadFile = File(...)):
    binary = await file.read()
    try:
        meta = create_bom_run(binary, file.filename)

        # 🔹 bom_filename 저장
        root = DATA_DIR / "bom_runs" / meta["bom_id"]
        bom_meta_path = root / "bom_meta.json"
        bom_meta_path.write_text(
            json.dumps(
                {
                    "bom_id": meta["bom_id"],
                    "bom_filename": file.filename,
                },
                ensure_ascii=False,
                indent=2
            ),
            encoding="utf-8"
        )

        return meta
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/bom/{bom_id}/specs")
def list_specs(bom_id: str):
    root = DATA_DIR / "bom_runs" / bom_id
    tree_excel = root / "tree.xlsx"

    if not tree_excel.exists():
        raise HTTPException(
            status_code=404,
            detail="tree.xlsx 파일이 없습니다."
        )

    try:
        wb = load_workbook(tree_excel, data_only=True)
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"엑셀 로드 실패: {e}"
        )

    # sheet 이름 = spec 목록
    specs = wb.sheetnames

    if not specs:
        raise HTTPException(
            status_code=400,
            detail="사양 시트가 없습니다."
        )

    return specs

@app.get("/api/bom/{bom_id}/tree", response_model=SubTree)
def get_tree(
    bom_id: str,
    request: Request,
    response: Response,
    spec: Optional[str] = None,
):
    sid = get_or_create_sid(request, response)
    state = SESSION_STATE.get(sid, {})

    resolved_spec = spec or state.get("spec")
    if not resolved_spec:
        raise HTTPException(status_code=400, detail="spec이 없습니다.")

    root_dir = DATA_DIR / "bom_runs" / bom_id
    tree_json_path = root_dir / f"{resolved_spec}.json"

    # 1. 캐시된 JSON 있으면 그대로 사용
    if tree_json_path.exists():
        tree = load_tree_json(root_dir, resolved_spec)
    else:
        # 2. 없으면 엑셀에서 생성
        tree_excel = root_dir / "tree.xlsx"
        if not tree_excel.exists():
            raise HTTPException(status_code=404, detail="tree.xlsx 파일이 없습니다.")

        bom_meta = read_bom_meta(root_dir)
        bom_filename = bom_meta.get("bom_filename")
        if not bom_filename:
            raise HTTPException(status_code=500, detail="bom_filename 없음")

        wb = load_workbook(tree_excel, data_only=True)
        if resolved_spec not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"시트 없음: {resolved_spec}")

        ws = wb[resolved_spec]

        tree = build_tree_from_sheet(
            ws=ws,
            bom_id=bom_id,
            bom_filename=bom_filename,
            spec_name=resolved_spec,
        )

        save_tree_json(root_dir, resolved_spec, tree)

    # 세션 갱신
    SESSION_STATE[sid] = {
        "bom_id": bom_id,
        "spec": resolved_spec,
        "selected_id": state.get("selected_id"),
    }
    save_session_state()

    return tree



@app.patch("/api/bom/{bom_id}/node/{node_id}", response_model=SubTree)
def patch_node(
    bom_id: str,
    node_id: str,
    patch: SubNodePatch,
    request: Request,
    response: Response,
):
    sid = get_or_create_sid(request, response)
    state = SESSION_STATE.get(sid, {})

    spec = state.get("spec")
    if not spec:
        raise HTTPException(status_code=400, detail="spec 없음")

    root_dir = DATA_DIR / "bom_runs" / bom_id

    tree = load_tree_json(root_dir, spec)
    nodes = tree.nodes   # ✅ 핵심 수정

    # flat 구조에서 노드 찾기
    target = next((n for n in nodes if n.id == node_id), None)
    if not target:
        raise HTTPException(status_code=404, detail="노드 없음")

    # patch 적용 (Pydantic 객체 기준)
    for field, value in patch.dict(exclude_unset=True).items():
        setattr(target, field, value)

    save_tree_json(root_dir, spec, tree)

    return tree

@app.patch("/api/bom/{bom_id}/move-node")
def move_node(
    bom_id: str,
    spec: str,                 # ⭐ 추가
    req: MoveNodeRequest
):
    root_dir = DATA_DIR / "bom_runs" / bom_id
    tree = load_tree_json(root_dir, spec)   # ⭐ spec 포함

    nodes = tree.nodes
    print("DEBUG nodes example =", nodes[:3])
    
    target = next((n for n in nodes if n.id == req.node_id), None)
    if not target:
        raise HTTPException(404, "node not found")

    # 4. parent + order 변경
    target.parent_id = req.new_parent_id
    target.order = req.new_index

    # 5. 같은 부모 아래 order 재정렬 (권장)
    siblings = [n for n in nodes if n.parent_id == req.new_parent_id]
    siblings.sort(key=lambda x: (n.order or 0))

    for i, n in enumerate(siblings):
        n.order = i

    # 6. JSON 저장
    save_tree_json(root_dir, spec, tree)

    return tree
