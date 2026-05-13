from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel


router = APIRouter(prefix="/lob", tags=["lob"])


class WorkerLobRow(BaseModel):
    worker: str
    valueTime: float
    requiredTime: float
    wasteTime: float
    movementTime: float


class WorkerLobExportRequest(BaseModel):
    spec: Optional[str] = None
    tactTime: float = 0
    neckTime: float = 0
    expectedCycleTime: float = 0
    workerRows: List[WorkerLobRow]


class TactInputs(BaseModel):
    workDaysPerYear: Optional[float] = 0
    dailyAvailableMinutes: Optional[float] = 0
    plannedStopMinutes: Optional[float] = 0
    realAvailableMinutes: Optional[float] = 0
    annualVehicleTarget: Optional[float] = 0
    quantityPerVehicle: Optional[float] = 0
    lineCount: Optional[float] = 0


class EquipmentLobRow(BaseModel):
    name: Optional[str] = ""
    investmentCost: Optional[Any] = ""
    equipmentTimeInput: Optional[Any] = 0
    manualTimeInput: Optional[Any] = 0
    reviewChecked: Optional[bool] = False
    improvementNote: Optional[str] = ""


THIN = Side(border_style="thin", color="94A3B8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill(fill_type="solid", start_color="DBEAFE", end_color="DBEAFE")
CARD_FILL = PatternFill(fill_type="solid", start_color="EFF6FF", end_color="EFF6FF")


def _safe_sheet_title(workbook: Workbook, raw_title: str) -> str:
    base = (raw_title or "Sheet").strip() or "Sheet"
    invalid_chars = set('[]:*?/\\')
    sanitized = "".join("_" if ch in invalid_chars else ch for ch in base)[:31] or "Sheet"

    if sanitized not in workbook.sheetnames:
        return sanitized

    suffix = 2
    while True:
        candidate = f"{sanitized[:28]}_{suffix}"[:31]
        if candidate not in workbook.sheetnames:
            return candidate
        suffix += 1


def _to_number(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _get_expected_ct_divisor(annual_vehicle_target) -> float:
    annual_vehicle_target = _to_number(annual_vehicle_target)
    if 1 <= annual_vehicle_target <= 2000:
        return 0.7
    if annual_vehicle_target <= 5000:
        return 0.75
    if annual_vehicle_target <= 10000:
        return 0.83
    if annual_vehicle_target <= 50000:
        return 0.85
    if annual_vehicle_target <= 100000:
        return 0.86
    if annual_vehicle_target <= 150000:
        return 0.89
    if annual_vehicle_target >= 150001:
        return 0.9
    return 0.0


def _calculate_tact_metrics(tact_inputs: Optional[Dict[str, Any]]) -> Dict[str, float]:
    tact_inputs = tact_inputs or {}
    work_days = _to_number(tact_inputs.get("workDaysPerYear"))
    daily_available = _to_number(tact_inputs.get("dailyAvailableMinutes"))
    planned_stop = _to_number(tact_inputs.get("plannedStopMinutes"))
    real_available = _to_number(tact_inputs.get("realAvailableMinutes"))
    annual_vehicles = _to_number(tact_inputs.get("annualVehicleTarget"))
    qty_per_vehicle = _to_number(tact_inputs.get("quantityPerVehicle"))
    line_count = _to_number(tact_inputs.get("lineCount"))

    computed_real_available = max(daily_available - planned_stop, 0)
    annual_required_quantity = annual_vehicles * qty_per_vehicle
    daily_required_quantity = annual_required_quantity / work_days if work_days > 0 else 0
    line_tact_minutes = (
        (real_available / daily_required_quantity) * line_count
        if daily_required_quantity > 0
        else 0
    )

    return {
        "workDaysPerYear": work_days,
        "dailyAvailableMinutes": daily_available,
        "plannedStopMinutes": planned_stop,
        "computedRealAvailableMinutes": computed_real_available,
        "realAvailableMinutes": real_available,
        "annualVehicleTarget": annual_vehicles,
        "quantityPerVehicle": qty_per_vehicle,
        "lineCount": line_count,
        "annualRequiredQuantity": annual_required_quantity,
        "dailyRequiredQuantity": daily_required_quantity,
        "lineTactMinutes": line_tact_minutes,
        "lineTactSeconds": line_tact_minutes * 60,
    }


def _build_process_design_metrics(rows, tact_inputs: Optional[Dict[str, Any]]) -> Dict[str, float]:
    worker_rows = _build_worker_rows_from_assembly_rows(rows)
    tact_metrics = _calculate_tact_metrics(tact_inputs)
    total_manual_time = sum(
        row["valueTime"] + row["requiredTime"] + row["wasteTime"] + row["movementTime"]
        for row in worker_rows
    )
    neck_time = max(
        (
            row["valueTime"] + row["requiredTime"] + row["wasteTime"] + row["movementTime"]
            for row in worker_rows
        ),
        default=0,
    )
    worker_count = len(worker_rows)
    divisor = _get_expected_ct_divisor(tact_metrics["annualVehicleTarget"])
    expected_cycle_time = neck_time / divisor if divisor > 0 and neck_time > 0 else 0
    expected_uph = 3600 / expected_cycle_time if expected_cycle_time > 0 else 0
    standard_uph = 3600 / neck_time if neck_time > 0 else 0
    expected_upmh = expected_uph / worker_count if worker_count > 0 else 0
    minimum_workers = (
        total_manual_time / tact_metrics["lineTactSeconds"]
        if tact_metrics["lineTactSeconds"] > 0
        else 0
    )
    total_worker_labor_sum = total_manual_time
    max_worker_labor_sum = neck_time
    lob_percent = (
        (total_worker_labor_sum / (max_worker_labor_sum * worker_count)) * 100
        if max_worker_labor_sum > 0 and worker_count > 0
        else 0
    )
    load_hours = (
        tact_metrics["dailyRequiredQuantity"] / expected_uph
        if expected_uph > 0
        else 0
    )
    daily_operating_hours = tact_metrics["realAvailableMinutes"] / 60
    load_rate_percent = (
        (load_hours / daily_operating_hours) * 100
        if daily_operating_hours > 0
        else 0
    )

    return {
        **tact_metrics,
        "totalManualTime": total_manual_time,
        "minimumWorkers": minimum_workers,
        "workerCount": worker_count,
        "neckTime": neck_time,
        "expectedCycleTime": expected_cycle_time,
        "efficiencyPercent": (neck_time / expected_cycle_time * 100) if expected_cycle_time > 0 else 0,
        "standardUph": standard_uph,
        "expectedUph": expected_uph,
        "expectedUpmh": expected_upmh,
        "lobPercent": lob_percent,
        "loadHours": load_hours,
        "loadRatePercent": load_rate_percent,
        "dailyLineCapacity": expected_uph * daily_operating_hours,
    }


def _style_table_cell(cell, *, header: bool = False):
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if header:
        cell.font = Font(bold=True)
        cell.fill = HEAD_FILL


def append_tact_time_sheet_to_workbook(
    workbook: Workbook,
    spec: Optional[str],
    tact_inputs: Optional[Dict[str, Any]],
):
    metrics = _calculate_tact_metrics(tact_inputs)
    ws = workbook.create_sheet(title=_safe_sheet_title(workbook, "4. Tact Time"))
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14

    ws["A1"] = "Tact Time 분석"
    ws["A1"].font = Font(size=18, bold=True)
    if spec:
        ws["A2"] = f"사양: {spec}"
        ws["A2"].font = Font(size=11, bold=True)

    rows = [
        ("1년 근무 횟수", metrics["workDaysPerYear"], "일"),
        ("1일 총 가용시간", metrics["dailyAvailableMinutes"], "분"),
        ("계획 정지시간", metrics["plannedStopMinutes"], "분"),
        ("실가용시간", metrics["realAvailableMinutes"], "분"),
        ("1년 생산 대수", metrics["annualVehicleTarget"], "대"),
        ("대당 환산 개수", metrics["quantityPerVehicle"], "개"),
        ("총 라인 수", metrics["lineCount"], "라인"),
        ("연간 요구 수량", metrics["annualRequiredQuantity"], "개"),
        ("일일 요구 수량", metrics["dailyRequiredQuantity"], "개/일"),
        ("라인당 Tact Time", metrics["lineTactMinutes"], "분"),
        ("라인당 Tact Time", metrics["lineTactSeconds"], "초"),
    ]

    header_row = 4
    for col, label in enumerate(("항목", "값", "단위"), start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        _style_table_cell(cell, header=True)

    for row_idx, (label, value, unit) in enumerate(rows, start=header_row + 1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=round(value, 2))
        ws.cell(row=row_idx, column=3, value=unit)
        for col in range(1, 4):
            _style_table_cell(ws.cell(row=row_idx, column=col))

    return workbook


def append_equipment_lob_sheet_to_workbook(
    workbook: Workbook,
    spec: Optional[str],
    equipment_rows: Optional[List[Dict[str, Any]]],
    target_cycle_time: float = 0,
):
    normalized_rows = []
    for index, row in enumerate(equipment_rows or [], start=1):
        equipment_time = _to_number(row.get("equipmentTimeInput"))
        manual_time = _to_number(row.get("manualTimeInput"))
        name = str(row.get("name") or "").strip() or f"설비 {index}"
        total_time = equipment_time + manual_time
        if not any((name.strip(), equipment_time, manual_time, row.get("investmentCost"), row.get("improvementNote"))):
            continue
        normalized_rows.append({
            "name": name,
            "investmentCost": row.get("investmentCost") or "",
            "equipmentTime": equipment_time,
            "manualTime": manual_time,
            "totalTime": total_time,
            "reviewChecked": bool(row.get("reviewChecked")),
            "improvementNote": row.get("improvementNote") or "",
        })

    if not normalized_rows:
        return workbook

    ws = workbook.create_sheet(title=_safe_sheet_title(workbook, "5. 설비 LOB"))
    ws.sheet_view.showGridLines = False
    widths = [20, 16, 14, 14, 14, 12, 32]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws["A1"] = "설비 LOB 분석표"
    ws["A1"].font = Font(size=18, bold=True)
    if spec:
        ws["A2"] = f"사양: {spec}"
        ws["A2"].font = Font(size=11, bold=True)
    ws["A3"] = "설비 목표 C/T"
    ws["B3"] = round(target_cycle_time, 2)
    ws["B3"].number_format = "0.00"

    headers = ["설비/공정", "투자비", "장비 시간", "수작업 시간", "합계", "재검토", "개선 방향"]
    header_row = 5
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        _style_table_cell(cell, header=True)

    for row_idx, row in enumerate(normalized_rows, start=header_row + 1):
        values = [
            row["name"],
            row["investmentCost"],
            row["equipmentTime"],
            row["manualTime"],
            row["totalTime"],
            "필요" if row["reviewChecked"] else "",
            row["improvementNote"],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            _style_table_cell(cell)
            if col in {3, 4, 5}:
                cell.number_format = "0.00"
            if col == 6 and row["reviewChecked"]:
                cell.fill = PatternFill(fill_type="solid", start_color="FEE2E2", end_color="FEE2E2")

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "설비 시간 구성"
    chart.y_axis.title = "sec"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.x_axis.numFmt = "General"
    chart.x_axis.lblOffset = 100
    data = Reference(
        ws,
        min_col=3,
        max_col=4,
        min_row=header_row,
        max_row=header_row + len(normalized_rows),
    )
    categories = Reference(
        ws,
        min_col=1,
        max_col=1,
        min_row=header_row + 1,
        max_row=header_row + len(normalized_rows),
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 10
    chart.width = 22
    chart.legend.position = "b"
    ws.add_chart(chart, "I5")

    return workbook


def append_process_design_sheet_to_workbook(
    workbook: Workbook,
    spec: Optional[str],
    rows,
    tact_inputs: Optional[Dict[str, Any]],
):
    metrics = _build_process_design_metrics(rows, tact_inputs)
    ws = workbook.create_sheet(title=_safe_sheet_title(workbook, "6. 공정설계표"))
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18

    ws["A1"] = "공정설계표"
    ws["A1"].font = Font(size=18, bold=True)
    if spec:
        ws["A2"] = f"사양: {spec}"
        ws["A2"].font = Font(size=11, bold=True)

    table_rows = [
        (1, "라인 Tact Time", "실가용시간 / 일일 요구 수량 x 라인 수", "sec", metrics["lineTactSeconds"]),
        (2, "작업자 총공수 (ΣHT)", "작업자별 가치·필비·낭비 공수 합계", "sec", metrics["totalManualTime"]),
        (3, "최소(이론) 작업자 수 (n)", "작업자 총공수 / Tact Time", "명", metrics["minimumWorkers"]),
        (4, "운영 작업자 수 (N)", "작업자 배정 수", "명", metrics["workerCount"]),
        (5, "Neck Time", "작업자별 공수 합계 중 최대값", "sec", metrics["neckTime"]),
        (6, "예상 Cycle Time", "Neck Time / 연간 생산대수별 기준율", "sec", metrics["expectedCycleTime"]),
        (7, "실동율", "Neck Time / 예상 Cycle Time", "%", metrics["efficiencyPercent"]),
        (8, "표준 UPH", "3600 / Neck Time", "대/h", metrics["standardUph"]),
        (9, "예상 UPH", "3600 / 예상 Cycle Time", "대/h", metrics["expectedUph"]),
        (10, "예상 UPMH", "예상 UPH / 운영 작업자 수", "대/MH", metrics["expectedUpmh"]),
        (11, "LOB", "공수 합계 총합 / (최대 공수 합계 x 작업자 수)", "%", metrics["lobPercent"]),
        (12, "부하시간", "일일 요구 수량 / 예상 UPH", "h", metrics["loadHours"]),
        (13, "부하율", "부하시간 / 일일 실가용시간", "%", metrics["loadRatePercent"]),
        (14, "일일 라인 생산능력", "예상 UPH x 일일 실가용시간", "대/일", metrics["dailyLineCapacity"]),
    ]

    header_row = 4
    for col, label in enumerate(("NO.", "항목", "세부 기준", "단위", spec or "값"), start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        _style_table_cell(cell, header=True)

    for row_idx, row in enumerate(table_rows, start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=round(value, 2) if isinstance(value, float) else value)
            _style_table_cell(cell)
            if col_idx == 5:
                cell.number_format = "0.00"

    return workbook


def _build_worker_rows_from_assembly_rows(rows):
    worker_map = {}

    for row in rows:
        category = str(row.get("no") or "").strip()
        if category not in {"가치", "필비", "낭비"}:
            continue

        worker = str(row.get("작업자") or "").strip() or "미지정"
        total_value = _to_number(row.get("TOTAL"))
        if total_value <= 0:
            total_value = _to_number(row.get("SEC")) * _to_number(row.get("반복횟수") or 1)

        if worker not in worker_map:
            worker_map[worker] = {
                "worker": worker,
                "valueTime": 0.0,
                "requiredTime": 0.0,
                "wasteTime": 0.0,
                "movementTime": 0.0,
            }

        if category == "가치":
            worker_map[worker]["valueTime"] += total_value
        elif category == "필비":
            worker_map[worker]["requiredTime"] += total_value
        elif category == "낭비":
            worker_map[worker]["wasteTime"] += total_value

    return sorted(
        worker_map.values(),
        key=lambda item: (float("inf"), item["worker"])
        if not any(char.isdigit() for char in item["worker"])
        else (int("".join(filter(str.isdigit, item["worker"]))), item["worker"]),
    )


def append_worker_lob_sheet_to_workbook(
    workbook: Workbook,
    spec: Optional[str],
    rows,
    tact_time: float = 0,
    neck_time: Optional[float] = None,
    expected_cycle_time: Optional[float] = None,
    annual_vehicle_target: Optional[float] = None,
):
    worker_rows = _build_worker_rows_from_assembly_rows(rows)
    if not worker_rows:
        return workbook

    computed_neck_time = max(
        (
            row["valueTime"] + row["requiredTime"] + row["wasteTime"] + row["movementTime"]
            for row in worker_rows
        ),
        default=0,
    )
    resolved_neck_time = computed_neck_time if neck_time is None else neck_time
    if expected_cycle_time is None:
        divisor = _get_expected_ct_divisor(annual_vehicle_target)
        resolved_expected_cycle_time = (
            (resolved_neck_time / divisor) if divisor > 0 and resolved_neck_time > 0 else resolved_neck_time
        )
    else:
        resolved_expected_cycle_time = expected_cycle_time

    ws = workbook.create_sheet(title=_safe_sheet_title(workbook, "3. 작업자 LOB"))
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 18
    for col_idx in range(2, len(worker_rows) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    total_col_idx = len(worker_rows) + 2
    lob_col_idx = len(worker_rows) + 3
    lob_end_col_idx = len(worker_rows) + 4
    ws.column_dimensions[get_column_letter(total_col_idx)].width = 14
    ws.column_dimensions[get_column_letter(lob_col_idx)].width = 12
    ws.column_dimensions[get_column_letter(lob_end_col_idx)].width = 12

    ws["A1"] = "작업자 LOB 분석표"
    ws["A1"].font = Font(size=18, bold=True)
    if spec:
        ws["A2"] = f"사양: {spec}"
        ws["A2"].font = Font(size=11, bold=True)

    summary_items = [
        ("Tact Time", tact_time if tact_time > 0 else ""),
        ("Neck Time", resolved_neck_time),
        ("예상 Cycle Time", resolved_expected_cycle_time),
    ]

    summary_row = 4
    for index, (label, value) in enumerate(summary_items):
        row = summary_row + index
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=round(value, 2) if isinstance(value, (int, float)) else value)
        for col in (1, 2):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.fill = CARD_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 1:
                cell.font = Font(bold=True)

    table_header_row = 9
    headers = ["항목", *[row["worker"] for row in worker_rows], "전체 합계"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=table_header_row, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.fill = HEAD_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(
        start_row=table_header_row,
        start_column=lob_col_idx,
        end_row=table_header_row,
        end_column=lob_end_col_idx,
    )
    lob_header_cell = ws.cell(row=table_header_row, column=lob_col_idx, value="LOB")
    lob_header_cell.font = Font(bold=True)
    lob_header_cell.fill = HEAD_FILL
    lob_header_cell.border = BORDER
    lob_header_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=table_header_row, column=lob_end_col_idx).border = BORDER

    row_defs = [
        ("가치", "valueTime", False),
        ("필비", "requiredTime", False),
        ("낭비", "wasteTime", False),
        ("이동시간", "movementTime", False),
        ("공수 합계", "laborSum", False),
        ("실동율", "efficiency", True),
    ]

    def resolve_row_value(worker_row, key):
        if key == "laborSum":
            return (
                worker_row["valueTime"]
                + worker_row["requiredTime"]
                + worker_row["wasteTime"]
                + worker_row["movementTime"]
            )
        if key == "efficiency":
            labor_sum = resolve_row_value(worker_row, "laborSum")
            return (worker_row["valueTime"] / labor_sum) if labor_sum > 0 else 0
        return worker_row[key]

    data_start_row = table_header_row + 1
    for row_offset, (label, key, is_percent) in enumerate(row_defs):
        row_idx = data_start_row + row_offset
        ws.cell(row=row_idx, column=1, value=label)

        for col_offset, worker_row in enumerate(worker_rows, start=2):
            value = resolve_row_value(worker_row, key)
            cell = ws.cell(row=row_idx, column=col_offset, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0.0%" if is_percent else "0.00"

        total_cell = ws.cell(row=row_idx, column=len(worker_rows) + 2)
        if key == "efficiency":
            total_cell.value = (
                f"=AVERAGE(B{row_idx}:{get_column_letter(len(worker_rows) + 1)}{row_idx})"
            )
            total_cell.number_format = "0.0%"
        else:
            total_cell.value = f"=SUM(B{row_idx}:{get_column_letter(len(worker_rows) + 1)}{row_idx})"
            total_cell.number_format = "0.00"
        total_cell.border = BORDER
        total_cell.alignment = Alignment(horizontal="center", vertical="center")

        label_cell = ws.cell(row=row_idx, column=1)
        label_cell.border = BORDER
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        label_cell.font = Font(bold=True)

    total_row = data_start_row + len(row_defs) - 1
    for row_idx in range(data_start_row, total_row + 1):
        for col_idx in range(1, lob_end_col_idx + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if row_idx in {data_start_row + 4, data_start_row + 5} or col_idx == len(worker_rows) + 2:
                cell.fill = HEAD_FILL

    ws.merge_cells(
        start_row=data_start_row,
        start_column=lob_col_idx,
        end_row=total_row,
        end_column=lob_end_col_idx,
    )
    lob_cell = ws.cell(row=data_start_row, column=lob_col_idx)
    labor_sum_row = data_start_row + 4
    labor_sum_start_col = get_column_letter(2)
    labor_sum_end_col = get_column_letter(len(worker_rows) + 1)
    lob_cell.value = (
        f"=IFERROR(SUM({labor_sum_start_col}{labor_sum_row}:{labor_sum_end_col}{labor_sum_row})/"
        f"(MAX({labor_sum_start_col}{labor_sum_row}:{labor_sum_end_col}{labor_sum_row})*{len(worker_rows)}),0)"
    )
    lob_cell.number_format = "0.0%"
    lob_cell.font = Font(size=24, bold=True, color="1D4ED8")
    lob_cell.fill = CARD_FILL
    lob_cell.border = BORDER
    lob_cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx in range(data_start_row, total_row + 1):
        for col_idx in range(lob_col_idx, lob_end_col_idx + 1):
            ws.cell(row=row_idx, column=col_idx).border = BORDER
            if row_idx != data_start_row or col_idx != lob_col_idx:
                ws.cell(row=row_idx, column=col_idx).fill = CARD_FILL

    helper_defs = [
        ("가치", "valueTime"),
        ("필비", "requiredTime"),
        ("낭비", "wasteTime"),
        ("이동시간", "movementTime"),
    ]
    helper_start_col = max(20, lob_end_col_idx + 2)
    helper_header_row = 1
    ws.cell(row=helper_header_row, column=helper_start_col, value="작업자")
    helper_font = Font(color="FFFFFF", size=1)
    helper_fill = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    helper_border = Border()
    ws.cell(row=helper_header_row, column=helper_start_col).font = helper_font
    ws.cell(row=helper_header_row, column=helper_start_col).fill = helper_fill
    ws.cell(row=helper_header_row, column=helper_start_col).border = helper_border
    for offset, (label, _key) in enumerate(helper_defs, start=1):
        cell = ws.cell(row=helper_header_row, column=helper_start_col + offset, value=label)
        cell.font = helper_font
        cell.fill = helper_fill
        cell.border = helper_border

    main_table_row_by_key = {
        key: data_start_row + row_offset
        for row_offset, (_label, key, _is_percent) in enumerate(row_defs)
    }

    for row_idx, worker_row in enumerate(worker_rows, start=2):
        worker_col_idx = row_idx
        worker_header_ref = f"{get_column_letter(worker_col_idx)}{table_header_row}"
        label_cell = ws.cell(row=row_idx, column=helper_start_col, value=f"={worker_header_ref}&\"번\"")
        label_cell.font = helper_font
        label_cell.fill = helper_fill
        label_cell.border = helper_border
        for offset, (_label, key) in enumerate(helper_defs, start=1):
            source_row_idx = main_table_row_by_key[key]
            source_ref = f"{get_column_letter(worker_col_idx)}{source_row_idx}"
            cell = ws.cell(row=row_idx, column=helper_start_col + offset, value=f"={source_ref}")
            cell.number_format = "0.00"
            cell.font = helper_font
            cell.fill = helper_fill
            cell.border = helper_border

    for col_idx in range(helper_start_col, helper_start_col + len(helper_defs) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 0.2

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "작업자 LOB 그래프"
    chart.y_axis.title = "sec"
    chart.x_axis.title = "작업자 번호"
    chart.legend.position = "b"
    chart.legend.overlay = False
    chart.height = 9
    chart.width = 24
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.y_axis.numFmt = "0.00"
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"

    data = Reference(
        ws,
        min_col=helper_start_col + 1,
        max_col=helper_start_col + len(helper_defs),
        min_row=1,
        max_row=1 + len(worker_rows),
    )
    categories = Reference(
        ws,
        min_col=helper_start_col,
        max_col=helper_start_col,
        min_row=2,
        max_row=1 + len(worker_rows),
    )
    chart.add_data(data, titles_from_data=True, from_rows=False)
    chart.set_categories(categories)
    series_colors = ["22C55E", "FACC15", "EF4444", "94A3B8"]
    for series, color in zip(chart.ser, series_colors):
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.solidFill = color
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.position = "ctr"
    ws.add_chart(chart, "I4")
    workbook.active = workbook.index(ws)

    return workbook


@router.post("/worker/export_excel")
def export_worker_lob_excel(payload: WorkerLobExportRequest):
    if not payload.workerRows:
        raise HTTPException(status_code=400, detail="작업자 LOB 데이터가 없습니다.")

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    append_worker_lob_sheet_to_workbook(
        wb,
        payload.spec,
        [row.model_dump() for row in payload.workerRows],
        payload.tactTime,
        payload.neckTime,
        payload.expectedCycleTime,
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{payload.spec or 'worker_lob'}_worker_lob.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
