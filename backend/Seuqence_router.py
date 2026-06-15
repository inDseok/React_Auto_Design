import logging
import time
from datetime import datetime

from fastapi import APIRouter, HTTPException, Query
from fastapi.encoders import jsonable_encoder
from pathlib import Path
import json
import os
from openpyxl import load_workbook
from typing import Any, List, Dict, Optional, Set, Tuple
from functools import lru_cache
import re
from types import SimpleNamespace
from backend.Assembly.auto_match import (
    combined_score,
    load_db_rows,
    match_one_best,
    jw_score,
    COMBINED_THRESHOLD,
    normalize_text,
    rf_score,
    TOPK,
)
from backend.sequence.ai_provider import generate_sequence_chat_recommendations
from backend.sequence.ai_service import generate_sequence_ai_draft
from backend.sequence.embedding_search import search_chat_candidates_with_bge_m3
from backend.sequence.schema import (
    SequenceAIDraftRequest,
    SequenceDebugPrintRequest,
    SequenceChatPerPartRequest,
    SequenceChatPerPartResponse,
    SequenceChatRequest,
    SequenceChatResponse,
    SequenceNextProcessRecommendationRequest,
    SequenceNextProcessRecommendationResponse,
    SequenceSaveRequest,
)
from backend.sequence_rag import retrieve_references_from_neo4j
from backend.sequence_rag.neo4j_retriever import retrieve_expanded_nodes_from_neo4j
from backend.sequence_rag.retriever import recommend_windows, summarize_next_process_candidates
from backend.sequence_rag.runtime import get_or_build_index

router = APIRouter(
    prefix="/sequence",
    tags=["sequence"]
)

logger = logging.getLogger(__name__)

DATA_DIR = Path("backend")
SEQUENCE_DEBUG_LOG_PATH = DATA_DIR / "logs" / "sequence_debug_log.txt"


def _write_json_atomic(path: Path, payload: Dict) -> None:
    temp_path = path.with_suffix(f"{path.suffix}.tmp")
    content = json.dumps(payload, ensure_ascii=False, indent=2)

    try:
        temp_path.write_text(content, encoding="utf-8")
        os.replace(temp_path, path)
    except PermissionError:
        path.write_text(content, encoding="utf-8")
        if temp_path.exists():
            temp_path.unlink()


def _append_sequence_debug_log(stage: str, payload: Optional[Dict[str, Any]] = None) -> None:
    SEQUENCE_DEBUG_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    safe_payload = payload or {}
    line = json.dumps(
        {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "stage": str(stage or "").strip() or "UNKNOWN",
            "payload": safe_payload,
        },
        ensure_ascii=False,
    )
    with SEQUENCE_DEBUG_LOG_PATH.open("a", encoding="utf-8") as fp:
        fp.write(line + "\n")


@lru_cache(maxsize=4)
def _get_cached_db_rows(excel_path_str: str, mtime: float):
    return load_db_rows(Path(excel_path_str))


@lru_cache(maxsize=4)
def _get_cached_db_match_data(excel_path_str: str, mtime: float):
    db_rows, db_choices = _get_cached_db_rows(excel_path_str, mtime)
    exact_rows_by_norm: Dict[str, List[Dict[str, Any]]] = {}

    for row in db_rows:
        norm = str(row.get("db_part_norm") or "").strip()
        if not norm:
            continue
        exact_rows_by_norm.setdefault(norm, []).append(row)

    return db_rows, db_choices, exact_rows_by_norm


def _format_tree_label(node: Dict) -> str:
    for key in ("id", "part_no", "name"):
        value = node.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return "ROOT"


def _build_tree_path_map(nodes: List[Dict]) -> Dict[str, List[str]]:
    node_by_name = {}
    for node in nodes:
        node_name = str(node.get("name") or "").strip()
        if node_name:
            node_by_name[node_name] = node

    path_cache: Dict[str, List[str]] = {}

    def resolve_path(node_name: str) -> List[str]:
        if not node_name:
            return []
        if node_name in path_cache:
            return path_cache[node_name]

        node = node_by_name.get(node_name)
        if not node:
            path_cache[node_name] = []
            return []

        parent_name = str(node.get("parent_name") or "").strip()
        parent_path = resolve_path(parent_name) if parent_name else []
        current_label = _format_tree_label(node)
        path_cache[node_name] = [*parent_path, current_label]
        return path_cache[node_name]

    for node_name in node_by_name:
        resolve_path(node_name)

    return path_cache


def _build_match_score_from_candidate(candidate: Dict[str, Any], source: str = "auto-match") -> Dict[str, Any]:
    return {
        "combined": float(candidate["score_combined"]),
        "rapidfuzz": float(candidate["score_rapidfuzz"]),
        "jaro_winkler": float(candidate["score_jw"]),
        "source": source,
    }


@lru_cache(maxsize=32)
def _get_cached_sequence_part_candidates(
    bom_id: str,
    spec: str,
    include_all_parts: bool,
    tree_mtime: float,
    excel_mtime: float,
) -> Dict[str, Any]:
    root_dir = DATA_DIR / "data" / "bom_runs" / bom_id
    json_path = root_dir / f"{spec}.json"
    excel_path = DATA_DIR / "작업시간분석표DB.xlsx"

    try:
        tree = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as e:
        raise HTTPException(500, f"JSON 로드 실패: {str(e)}")

    nodes = tree.get("nodes", [])
    db_rows, db_choices, exact_rows_by_norm = _get_cached_db_match_data(
        str(excel_path.resolve()),
        excel_mtime,
    )

    parts = []
    query_cache: Dict[str, Optional[Dict[str, Any]]] = {}
    tree_path_map = _build_tree_path_map(nodes)

    for n in nodes:
        if n.get("type") != "PART":
            continue
        if not include_all_parts and n.get("inhouse") is not True:
            continue

        part_id = n.get("id")
        part_name = part_id or n.get("name")
        node_name = str(n.get("name") or "").strip()
        recommended_part_base = n.get("recommended_part_base")
        recommended_source_sheet = n.get("recommended_source_sheet")
        recommended_match_score = n.get("recommended_match_score")

        best = None
        if recommended_part_base and recommended_source_sheet:
            part_base = recommended_part_base
            source_sheet = recommended_source_sheet
            match_score = recommended_match_score or {
                "source": "manual-recommendation",
            }
        else:
            for query_raw in [part_name, part_id]:
                if not query_raw:
                    continue

                query_text = str(query_raw).strip()
                if not query_text:
                    continue

                if query_text in query_cache:
                    candidate = query_cache[query_text]
                else:
                    normalized_query = normalize_text(query_text)
                    exact_matches = exact_rows_by_norm.get(normalized_query, [])

                    if len(exact_matches) == 1:
                        exact_match = exact_matches[0]
                        candidate = {
                            "json_id_raw": query_text,
                            "json_id_norm": normalized_query,
                            "db_part_raw": exact_match["db_part_raw"],
                            "db_part_norm": exact_match["db_part_norm"],
                            "score_rapidfuzz": 100.0,
                            "score_jw": 100.0,
                            "score_combined": 100.0,
                            "sheet": exact_match["sheet"],
                            "row_index": exact_match["row_index"],
                        }
                    else:
                        candidate = match_one_best(
                            query_raw=query_text,
                            db_rows=db_rows,
                            db_choices=db_choices,
                            topk=TOPK,
                        )

                    query_cache[query_text] = candidate

                if candidate and candidate["score_combined"] >= COMBINED_THRESHOLD:
                    best = candidate
                    break

        if recommended_part_base and recommended_source_sheet:
            pass
        elif best and best["score_combined"] >= COMBINED_THRESHOLD:
            part_base = best["db_part_raw"]
            source_sheet = best["sheet"]
            match_score = _build_match_score_from_candidate(best)
        else:
            part_base = None
            source_sheet = None
            match_score = None

        parts.append({
            "partId": part_id,
            "partName": part_name,
            "inhouse": n.get("inhouse") is True,
            "treePath": tree_path_map.get(node_name, []),
            "parentName": n.get("parent_name"),
            "nodeName": node_name,
            "partBase": part_base,
            "sourceSheet": source_sheet,
            "matchScore": match_score,
            "parentId": n.get("parent_id"),
            "order": n.get("order"),
        })

    return {
        "bomId": bom_id,
        "spec": spec,
        "source": "sub-tree + auto-match" if include_all_parts else "sub-tree inhouse + auto-match",
        "count": len(parts),
        "parts": parts,
    }


def _load_sequence_part_candidates(
    bom_id: str,
    spec: str,
    *,
    include_all_parts: bool = False,
) -> Dict[str, Any]:
    root_dir = DATA_DIR / "data"/ "bom_runs" / bom_id
    json_path = root_dir / f"{spec}.json"
    excel_path = DATA_DIR / "작업시간분석표DB.xlsx"

    if not excel_path.exists():
        raise HTTPException(
            status_code=500,
            detail=f"작업시간 분석표 DB 엑셀 없음: {excel_path}"
        )

    if not json_path.exists():
        return {
            "bomId": bom_id,
            "spec": spec,
            "source": "manual-sequence" if include_all_parts else "manual-sequence-inhouse",
            "count": 0,
            "parts": [],
        }
    return _get_cached_sequence_part_candidates(
        bom_id,
        spec,
        include_all_parts,
        json_path.stat().st_mtime,
        excel_path.stat().st_mtime,
    )


def _filter_sequence_part_candidates_to_db_mapped(payload: Dict[str, Any]) -> Dict[str, Any]:
    parts = [
        part
        for part in list(payload.get("parts") or [])
        if str(part.get("partBase") or "").strip() and str(part.get("sourceSheet") or "").strip()
    ]
    return {
        **payload,
        "source": "작업시간분석표 매핑 부품",
        "count": len(parts),
        "parts": parts,
    }


@router.get("/inhouse-parts")
def get_inhouse_parts(bomId: str, spec: str):
    """
    Sequence 구성용 inhouse PART 목록 반환
    + 작업시간 DB 기준 partBase / sourceSheet 자동 매칭
    (Assembly row 생성은 하지 않음)
    """
    return _load_sequence_part_candidates(bomId, spec, include_all_parts=False)


@router.get("/parts")
def get_sequence_parts(bomId: str, spec: str):
    """
    Sequence 채팅 추천용 전체 PART 목록 반환
    - 상위 조립체(inhouse=false) 포함
    """
    payload = _load_sequence_part_candidates(bomId, spec, include_all_parts=True)
    return _filter_sequence_part_candidates_to_db_mapped(payload)



EXCEL_DB_PATH = Path("backend/작업시간분석표DB.xlsx")


def _read_sequence_sheet_headers(ws, header_row: int) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is not None and str(value).strip():
            headers[str(value).strip()] = col
    return headers


def _resolve_sequence_sheet_columns(headers: Dict[str, int]) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    part_col = None
    process_col = None
    option_col = None

    for key, column in headers.items():
        normalized = key.replace(" ", "")
        if "부품" in normalized and part_col is None:
            part_col = column
        if ("요소작업" in normalized or "공정" in normalized) and process_col is None:
            process_col = column
        if "OPTION" in key.upper() and option_col is None:
            option_col = column

    return part_col, process_col, option_col


def _resolve_sequence_action_columns(headers: Dict[str, int]) -> Dict[str, Optional[int]]:
    columns: Dict[str, Optional[int]] = {
        "part": None,
        "process": None,
        "option": None,
        "worker": None,
        "category": None,
        "action": None,
    }

    for key, column in headers.items():
        normalized = key.replace(" ", "")
        upper_key = key.upper()

        if columns["part"] is None and "부품" in normalized:
            columns["part"] = column
        if columns["process"] is None and ("요소작업" in normalized or "공정" in normalized):
            columns["process"] = column
        if columns["option"] is None and "OPTION" in upper_key:
            columns["option"] = column
        if columns["worker"] is None and "작업자" in normalized:
            columns["worker"] = column
        if columns["category"] is None and normalized == "no":
            columns["category"] = column
        if columns["action"] is None and "동작요소" in normalized:
            columns["action"] = column

    return columns


def _iter_sequence_sheet_rows(ws, header_row: int):
    headers = _read_sequence_sheet_headers(ws, header_row)
    part_col, process_col, option_col = _resolve_sequence_sheet_columns(headers)

    if not part_col:
        return

    current_part = None
    current_process = None
    current_option = None

    for row in range(header_row + 1, ws.max_row + 1):
        raw_part = ws.cell(row=row, column=part_col).value
        raw_process = ws.cell(row=row, column=process_col).value if process_col else None
        raw_option = ws.cell(row=row, column=option_col).value if option_col else None

        if raw_part is not None and str(raw_part).strip():
            current_part = str(raw_part).strip()
        if raw_process is not None and str(raw_process).strip():
            current_process = str(raw_process).strip()
        if raw_option is not None and str(raw_option).strip():
            current_option = str(raw_option).strip()

        option_value = current_option or ""
        if not current_part and not current_process and not option_value:
            continue

        yield {
            "row": row,
            "partBase": current_part or "",
            "processLabel": current_process or "",
            "option": option_value,
        }


def _iter_sequence_action_rows(ws, header_row: int):
    headers = _read_sequence_sheet_headers(ws, header_row)
    columns = _resolve_sequence_action_columns(headers)

    if not columns["part"]:
        return

    current_part = None
    current_process = None
    current_option = None

    for row in range(header_row + 1, ws.max_row + 1):
        raw_part = ws.cell(row=row, column=columns["part"]).value
        raw_process = (
            ws.cell(row=row, column=columns["process"]).value
            if columns["process"]
            else None
        )
        raw_option = (
            ws.cell(row=row, column=columns["option"]).value
            if columns["option"]
            else None
        )
        raw_worker = (
            ws.cell(row=row, column=columns["worker"]).value
            if columns["worker"]
            else None
        )
        raw_category = (
            ws.cell(row=row, column=columns["category"]).value
            if columns["category"]
            else None
        )
        raw_action = (
            ws.cell(row=row, column=columns["action"]).value
            if columns["action"]
            else None
        )

        if raw_part is not None and str(raw_part).strip():
            current_part = str(raw_part).strip()
        if raw_process is not None and str(raw_process).strip():
            current_process = str(raw_process).strip()
        if raw_option is not None and str(raw_option).strip():
            current_option = str(raw_option).strip()

        action_value = str(raw_action).strip() if raw_action is not None and str(raw_action).strip() else ""
        if not current_part and not current_process and not current_option and not action_value:
            continue

        yield {
            "row": row,
            "partBase": current_part or "",
            "processLabel": current_process or "",
            "option": current_option or "",
            "worker": str(raw_worker).strip() if raw_worker is not None and str(raw_worker).strip() else "",
            "category": str(raw_category).strip() if raw_category is not None and str(raw_category).strip() else "",
            "actionElement": action_value,
        }


def collect_options_for_part(ws, part_col: int, option_col: int, header_row: int, part_base: str):
    options = set()

    for item in _iter_sequence_sheet_rows(ws, header_row):
        if item.get("partBase") != part_base:
            continue
        option_value = item.get("option")
        if option_value:
            options.add(option_value)

    return sorted(options)


@lru_cache(maxsize=4)
def _get_cached_sequence_option_rows(excel_path_str: str, mtime: float) -> Dict[str, Any]:
    wb = load_workbook(Path(excel_path_str), data_only=True)
    rows_by_sheet: Dict[str, List[Dict[str, Any]]] = {}
    action_rows_by_sheet: Dict[str, List[Dict[str, Any]]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_by_sheet[sheet_name] = list(_iter_sequence_sheet_rows(ws, 2))
        action_rows_by_sheet[sheet_name] = list(_iter_sequence_action_rows(ws, 2))

    return {
        "sheetNames": list(wb.sheetnames),
        "rowsBySheet": rows_by_sheet,
        "actionRowsBySheet": action_rows_by_sheet,
    }


def _get_sequence_option_index() -> Dict[str, Any]:
    if not EXCEL_DB_PATH.exists():
        return {
            "sheetNames": [],
            "rowsBySheet": {},
            "actionRowsBySheet": {},
        }

    return _get_cached_sequence_option_rows(
        str(EXCEL_DB_PATH.resolve()),
        EXCEL_DB_PATH.stat().st_mtime,
    )


def _resolve_option_sheet_names(option_index: Dict[str, Any], source_sheet: Optional[str]) -> List[str]:
    requested_sheet = (source_sheet or "").strip()
    sheet_names = list(option_index.get("sheetNames") or [])
    if requested_sheet and requested_sheet in sheet_names:
        return [requested_sheet]
    return sheet_names


def _load_action_elements_from_excel(
    part_base: str,
    option_value: str,
    *,
    process_label: str = "",
    source_sheet: str = "",
) -> Dict[str, Any]:
    normalized_part = _normalize_relation_key(part_base)
    normalized_process = _normalize_relation_key(process_label)
    normalized_option = _normalize_relation_key(option_value)

    if not normalized_part or not normalized_option:
        return {
            "matchedSheets": [],
            "rows": [],
            "resolvedOption": "",
            "matchStrategy": "none",
        }

    option_index = _get_sequence_option_index()
    sheet_names = _resolve_option_sheet_names(option_index, source_sheet)
    action_rows_by_sheet = option_index.get("actionRowsBySheet") or {}
    def collect_matches(
        *,
        target_option: str = "",
        require_option: bool = True,
        require_process: bool = True,
    ) -> Tuple[List[str], List[Dict[str, Any]], str]:
        local_matched_sheets: List[str] = []
        local_matched_rows: List[Dict[str, Any]] = []
        resolved_option = ""
        normalized_target_option = _normalize_relation_key(target_option) if require_option else ""
        if require_option and not normalized_target_option:
            return local_matched_sheets, local_matched_rows, resolved_option

        for sheet_name in sheet_names:
            sheet_rows = action_rows_by_sheet.get(sheet_name, [])
            current_sheet_matches = []
            current_sheet_options: List[str] = []

            for item in sheet_rows:
                if _normalize_relation_key(item.get("partBase")) != normalized_part:
                    continue
                if require_option and _normalize_relation_key(item.get("option")) != normalized_target_option:
                    continue
                if require_process and normalized_process and not _process_labels_match(item.get("processLabel"), process_label):
                    continue

                option_text = str(item.get("option") or "").strip()
                if option_text:
                    current_sheet_options.append(option_text)
                current_sheet_matches.append(
                    {
                        "row": item.get("row"),
                        "partBase": item.get("partBase", ""),
                        "processLabel": item.get("processLabel", ""),
                        "option": item.get("option", ""),
                        "worker": item.get("worker", ""),
                        "category": item.get("category", ""),
                        "actionElement": item.get("actionElement", ""),
                        "sheetName": sheet_name,
                    }
                )

            if current_sheet_matches:
                local_matched_sheets.append(sheet_name)
                local_matched_rows.extend(current_sheet_matches)
                if not resolved_option:
                    distinct_options = list(dict.fromkeys([opt for opt in current_sheet_options if opt]))
                    if distinct_options:
                        resolved_option = distinct_options[0]

        return local_matched_sheets, local_matched_rows, resolved_option

    matched_sheets, matched_rows, resolved_option = collect_matches(
        target_option=option_value,
        require_option=True,
        require_process=True,
    )
    if matched_rows:
        return {
            "matchedSheets": matched_sheets,
            "rows": matched_rows,
            "resolvedOption": resolved_option or str(option_value or "").strip(),
            "matchStrategy": "exact",
        }

    if normalized_process:
        matched_sheets, matched_rows, resolved_option = collect_matches(
            require_option=False,
            require_process=True,
        )
        distinct_options = list(
            dict.fromkeys(
                str(item.get("option") or "").strip()
                for item in matched_rows
                if str(item.get("option") or "").strip()
            )
        )
        if matched_rows and len(distinct_options) == 1:
            return {
                "matchedSheets": matched_sheets,
                "rows": matched_rows,
                "resolvedOption": resolved_option or distinct_options[0],
                "matchStrategy": "process-only",
            }

    matched_sheets = []
    matched_rows = []
    resolved_option = ""
    fallback_option = ""
    fallback_options = _load_options_from_excel(part_base, source_sheet)
    distinct_fallback_options = list(dict.fromkeys([str(option or "").strip() for option in fallback_options if str(option or "").strip()]))
    if len(distinct_fallback_options) == 1:
        fallback_option = distinct_fallback_options[0]
        if _normalize_relation_key(fallback_option) != normalized_option:
            matched_sheets, matched_rows, resolved_option = collect_matches(
                target_option=fallback_option,
                require_option=True,
                require_process=False,
            )

    if matched_rows:
        return {
            "matchedSheets": matched_sheets,
            "rows": matched_rows,
            "resolvedOption": resolved_option or fallback_option,
            "matchStrategy": "single-option",
        }

    return {
        "matchedSheets": [],
        "rows": [],
        "resolvedOption": "",
        "matchStrategy": "none",
    }


def _process_labels_match(left: Any, right: Any) -> bool:
    normalized_left = _normalize_relation_key(left)
    normalized_right = _normalize_relation_key(right)
    if not normalized_left or not normalized_right:
        return False
    return (
        normalized_left == normalized_right
        or normalized_left in normalized_right
        or normalized_right in normalized_left
    )


_CHAT_STOP_TOKENS = {
    "추천",
    "해주세요",
    "해줘",
    "구성",
    "자동",
    "시퀀스",
    "공정",
    "부품",
    "옵션",
    "작업",
    "중심",
    "위주",
    "포함",
}

_CHAT_TOKEN_SYNONYMS = {
    "메인": ["MAIN"],
    "렌즈": ["LENS", "OTR LENS", "INR LENS", "ASPHERIC LENS", "TIR LENS"],
    "메인렌즈": ["MAIN LENS", "OTR LENS", "OTR", "MAIN"],
    "메인 렌즈": ["MAIN LENS", "OTR LENS", "OTR", "MAIN"],
    "히트싱크": ["HEAT SINK", "H/S"],
    "히트 싱크": ["HEAT SINK", "H/S"],
    "HEAT": ["HEAT SINK"],
    "SINK": ["HEAT SINK"],
    "H/S": ["HEAT SINK", "H/S"],
    "램": ["LAM"],
    "람": ["LAM"],
    "LAM": ["LAM"],
    "더스트": ["DUST", "DUST COVER", "DUST CAP"],
    "커버": ["COVER", "DUST COVER"],
    "캡": ["CAP", "DUST CAP"],
    "더스트커버": ["DUST", "COVER", "DUST COVER", "DUST CAP", "DUST CAP COVER"],
    "더스트 커버": ["DUST", "COVER", "DUST COVER", "DUST CAP", "DUST CAP COVER"],
    "더스트캡": ["DUST", "CAP", "DUST CAP", "DUST COVER", "DUST CAP COVER"],
    "더스트 캡": ["DUST", "CAP", "DUST CAP", "DUST COVER", "DUST CAP COVER"],
    "베젤": ["BEZEL"],
    "범퍼": ["BUMPER"],
    "브라켓": ["BRACKET", "BRKT"],
    "프로젝션": ["PROJECTION", "PROJ", "UNIT"],
    "유닛": ["UNIT"],
    "하우징": ["HOUSING", "MAIN HOUSING", "HOUSING S/A"],
    "모듈": ["MODULE", "모듈", "LDM", "LED DRIVE MODULE", "주광 LED 모듈 ASS'Y", "광모듈"],
    "MODULE": ["MODULE", "모듈", "LDM", "LED DRIVE MODULE", "주광 LED 모듈 ASS'Y", "광모듈"],
    "광모듈": ["주광 LED 모듈 ASS'Y"],
    "광 모듈": ["주광 LED 모듈 ASS'Y"],
    "바코드": ["BAR", "CODE", "BAR CODE", "BAR-CODE"],
    "스캔": ["SCAN", "BAR", "CODE"],
    "가압": ["가압", "압입", "조립", "결합", "PRESS"],
    "체결": ["SCREW", "T/SCREW"],
    "스크류": ["SCREW", "T/SCREW"],
    "볼트": ["BOLT", "SCREW"],
    "안착": ["안착", "로딩", "지그 안착", "LOADING"],
    "와이어": ["WIRE", "WIRING"],
    "커넥터": ["CONNECTOR", "연결"],
    "연결": ["CONNECTOR", "WIRE", "WIRING"],
    "포장": ["WRAP", "박싱", "PACKING"],
    "랩핑": ["WRAP"],
    "박싱": ["BOX", "박싱", "PACKING"],
    "검사": ["검사", "INSPECTION"],
    "안착": ["안착", "LOADING"],
    "취출": ["취출", "UNLOAD"],
    "에어": ["AIR", "BLOWING"],
    "에어블로잉": ["에어 블로잉", "블로잉", "AIR", "BLOWING", "AIR BLOWING"],
    "에어블로윙": ["에어 블로윙", "블로윙", "AIR", "BLOWING", "AIR BLOWING"],
    "블로잉": ["BLOWING", "AIR"],
    "설비작동": ["설비 작동", "작동","동작", "스위치", "ON", "OPERATE"],
    "플라스틱": ["PLASTIC", "플라스틱류"],
    "PCB": ["PCB"],
    "기판": ["PCB", "기판"],
    "비철": ["비철", "METAL", "비철금속류"],
    "금속": ["METAL", "비철금속류"],
}

_CHAT_KOREAN_PARTICLE_SUFFIXES = (
    "과",
    "와",
    "을",
    "를",
    "은",
    "는",
    "이",
    "가",
    "도",
    "만",
    "에",
    "에서",
    "으로",
    "로",
    "한테",
    "께",
    "랑",
    "이랑",
    "하고",
)

_PROCESS_MATERIAL_HINTS = {
    "plastic": ("PLASTIC", "플라스틱"),
    "pcb": ("PCB", "기판"),
    "metal": ("비철", "금속", "METAL"),
}

_PROCESS_MESSAGE_HINTS = {
    "가압": ("가압", "압입", "조립", "결합", "PRESS"),
    "체결": ("체결", "SCREW", "T/SCREW", "드라이버", "볼트", "나사"),
    "안착": ("안착", "로딩", "지그 안착", "LOADING"),
    "취출": ("취출", "취출 작업", "부품 취출", "단품 취출", "UNLOAD", "EXTRACT", "TAKE OUT"),
    "블로잉": ("블로잉", "블로윙", "에어블로잉", "에어블로윙", "BLOWING", "AIR BLOWING", "에어 블로잉", "에어 블로윙"),
    "작동": ("작동", "설비 작동", "스위치", "ON", "동작", "OPERATE"),
    "바코드": ("BAR", "CODE", "BAR CODE", "BAR-CODE", "스캔"),
}

_CHAT_FASTENER_HINTS = (
    "SCREW",
    "T/SCREW",
    "M/SCREW",
    "BOLT",
    "NUT",
    "WASHER",
    "CLIP",
)

_CHAT_PART_FAMILY_HINTS = {
    "HEAT SINK": ("히트싱크", "히트 싱크", "HEAT SINK", "H/S"),
    "BEZEL": ("베젤", "BEZEL"),
    "LENS": ("렌즈", "LENS"),
    "HOUSING": ("하우징", "HOUSING", "HSG"),
    "BRACKET": ("브라켓", "BRACKET", "BRKT"),
    "BUMPER": ("범퍼", "BUMPER"),
    "COVER": ("커버", "COVER"),
    "CAP": ("캡", "CAP"),
    "MODULE": ("모듈", "MODULE", "광모듈", "광 모듈", "LDM", "LED DRIVE MODULE", "주광 LED 모듈"),
    "LDM": ("광모듈", "광 모듈", "LDM", "LED DRIVE MODULE", "주광 LED 모듈"),
    "PCB": ("PCB", "기판"),
}

_CHAT_PART_CANONICAL_TARGETS = {
    "광모듈": "주광 LED 모듈 ASS'Y",
    "광 모듈": "주광 LED 모듈 ASS'Y",
}


def _normalize_chat_text(value: Any) -> str:
    normalized = str(value or "").strip().upper()
    if not normalized:
        return ""
    normalized = normalized.replace("_", " ").replace("-", " ")
    normalized = re.sub(r"[^0-9A-Z가-힣/ ]+", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def _tokenize_chat_text(value: Any) -> List[str]:
    normalized = _normalize_chat_text(value)
    if not normalized:
        return []
    tokens: List[str] = []
    base_variants: List[str] = []
    for token in normalized.split(" "):
        if not token or len(token) < 2 or token in _CHAT_STOP_TOKENS:
            continue
        token_variants = [token]
        for suffix in _CHAT_KOREAN_PARTICLE_SUFFIXES:
            if token.endswith(suffix) and len(token) > len(suffix) + 1:
                stripped = token[: -len(suffix)].strip()
                if stripped and stripped not in token_variants:
                    token_variants.append(stripped)

        for variant in token_variants:
            if not variant or len(variant) < 2 or variant in _CHAT_STOP_TOKENS:
                continue
            base_variants.append(variant)
            tokens.append(variant)
            for synonym in _CHAT_TOKEN_SYNONYMS.get(variant, []):
                synonym_token = _normalize_chat_text(synonym)
                if synonym_token:
                    tokens.append(synonym_token)

    for index in range(len(base_variants) - 1):
        combined = f"{base_variants[index]} {base_variants[index + 1]}".strip()
        if not combined or combined in _CHAT_STOP_TOKENS:
            continue
        tokens.append(combined)
        for synonym in _CHAT_TOKEN_SYNONYMS.get(combined, []):
            synonym_token = _normalize_chat_text(synonym)
            if synonym_token:
                tokens.append(synonym_token)
    return list(dict.fromkeys(tokens))


def _score_candidate(message_tokens: List[str], *fields: Any) -> Tuple[float, List[str]]:
    haystack = " ".join(_normalize_chat_text(field) for field in fields if field).strip()
    if not haystack:
        return 0.0, []

    matched_tokens: List[str] = []
    score = 0.0
    for token in message_tokens:
        if token in haystack:
            matched_tokens.append(token)
            score += 2.0 if len(token) >= 4 else 1.0

    if not score and haystack:
        if any(keyword in haystack for keyword in ("SCREW", "체결")) and any(
            keyword in " ".join(message_tokens) for keyword in ("SCREW", "체결", "볼트", "나사")
        ):
            matched_tokens.append("체결")
            score += 1.5
        if any(keyword in haystack for keyword in ("BAR CODE", "BAR/CODE", "BARCODE", "스캔")) and any(
            keyword in " ".join(message_tokens) for keyword in ("BAR", "CODE", "스캔", "바코드")
        ):
            matched_tokens.append("바코드")
            score += 1.5
        if any(keyword in haystack for keyword in ("포장", "WRAP", "박싱")) and any(
            keyword in " ".join(message_tokens) for keyword in ("포장", "랩핑", "WRAP", "박싱")
        ):
            matched_tokens.append("포장")
            score += 1.5

    return score, matched_tokens


def _requested_part_families_from_message(message_tokens: List[str]) -> set:
    token_blob = " ".join(_normalize_chat_text(token) for token in message_tokens if token).strip()
    if not token_blob:
        return set()

    requested = set()
    for family, hints in _CHAT_PART_FAMILY_HINTS.items():
        if any(_normalize_chat_text(hint) in token_blob for hint in hints):
            requested.add(family)
    return requested


def _message_requests_bumper_bracket(message_tokens: List[str]) -> bool:
    token_blob = " ".join(_normalize_chat_text(token) for token in message_tokens if token).strip()
    if not token_blob:
        return False
    return any(
        phrase in token_blob
        for phrase in (
            "범퍼브라켓",
            "범퍼 브라켓",
            "BUMPER BRACKET",
            "BUMPER BRKT",
        )
    )


def _part_matches_requested_families(
    part: Any,
    requested_families: set,
    message_tokens: Optional[List[str]] = None,
) -> bool:
    if not requested_families:
        return True

    if isinstance(part, dict):
        values = (
            part.get("partBase"),
            part.get("contextPartBase"),
            part.get("partName"),
            part.get("partId"),
            part.get("nodeName"),
            part.get("label"),
            part.get("displayLabel"),
            part.get("operationLabel"),
            part.get("processKey"),
            part.get("sourceSheet"),
        )
    else:
        values = (
            getattr(part, "partBase", None),
            getattr(part, "contextPartBase", None),
            getattr(part, "partName", None),
            getattr(part, "partId", None),
            getattr(part, "nodeName", None),
            getattr(part, "label", None),
            getattr(part, "displayLabel", None),
            getattr(part, "operationLabel", None),
            getattr(part, "processKey", None),
            getattr(part, "sourceSheet", None),
        )

    haystack = " ".join(_normalize_chat_text(value) for value in values if value).strip()
    if not haystack:
        return False
    detected_families = _detect_part_families_in_values(values)

    if _message_requests_bumper_bracket(message_tokens or []):
        if {"BUMPER", "BRACKET"}.issubset(detected_families):
            return True
        if "BUMPER" in detected_families or "BRACKET" in detected_families:
            return False

    for family in requested_families:
        if family == "HOUSING" and ("HOUSING" in haystack or "HSG" in haystack or "하우징" in haystack):
            return True
        if family == "BRACKET" and ("BRACKET" in haystack or "BRKT" in haystack or "브라켓" in haystack):
            return True
        if family == "MODULE" and any(
            hint in haystack
            for hint in ("MODULE", "모듈", "LDM", "LED DRIVE MODULE", "주광 LED 모듈", "광모듈", "광 모듈")
        ):
            return True
        if family in haystack:
            return True
    return False


def _filter_part_matches_by_requested_families(
    part_matches: List[Dict[str, Any]],
    requested_families: set,
    message_tokens: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    if not requested_families:
        return part_matches
    return [
        item
        for item in (part_matches or [])
        if _part_matches_requested_families(item, requested_families, message_tokens)
    ]


def _read_candidate_values(item: Any, keys: Tuple[str, ...]) -> Tuple[Any, ...]:
    if isinstance(item, dict):
        return tuple(item.get(key) for key in keys)
    return tuple(getattr(item, key, None) for key in keys)


def _detect_part_families_in_values(values: Tuple[Any, ...]) -> set:
    haystack = " ".join(_normalize_chat_text(value) for value in values if value).strip()
    if not haystack:
        return set()

    detected = set()
    for family, hints in _CHAT_PART_FAMILY_HINTS.items():
        if any(_normalize_chat_text(hint) in haystack for hint in hints):
            detected.add(family)
    return detected


def _requested_process_actions_from_message(message_tokens: List[str]) -> set[str]:
    token_blob = " ".join(_normalize_chat_text(token) for token in message_tokens if token).strip()
    if not token_blob:
        return set()

    requested = set()
    for keyword, hints in _PROCESS_MESSAGE_HINTS.items():
        normalized_hints = [_normalize_chat_text(keyword), *(_normalize_chat_text(hint) for hint in hints)]
        if any(hint and hint in token_blob for hint in normalized_hints):
            requested.add(keyword)
    return requested


def _process_matches_requested_action(process: Any, message_tokens: List[str]) -> bool:
    requested_actions = _requested_process_actions_from_message(message_tokens)
    if not requested_actions:
        return False

    values = _read_candidate_values(
        process,
        (
            "processKey",
            "label",
            "displayLabel",
            "operationLabel",
            "partBase",
            "contextPartBase",
            "sourceSheet",
        ),
    )
    process_haystack = " ".join(_normalize_chat_text(value) for value in values if value).strip()
    if not process_haystack:
        return False

    for keyword in requested_actions:
        hints = _PROCESS_MESSAGE_HINTS.get(keyword, ())
        if any(
            _normalize_chat_text(hint) in process_haystack for hint in hints
        ):
            return True

    requested_materials = _extract_requested_materials(message_tokens)
    if requested_materials and _detect_process_material(values[0], " ".join(str(value or "") for value in values[1:])):
        return True

    return False


def _process_matches_action_keyword(process: Any, keyword: str) -> bool:
    hints = _PROCESS_MESSAGE_HINTS.get(keyword, ())
    if not hints:
        return False

    values = _read_candidate_values(
        process,
        (
            "processKey",
            "label",
            "displayLabel",
            "operationLabel",
            "partBase",
            "contextPartBase",
            "sourceSheet",
        ),
    )
    process_haystack = " ".join(_normalize_chat_text(value) for value in values if value).strip()
    if not process_haystack:
        return False
    return any(_normalize_chat_text(hint) in process_haystack for hint in hints)


def _boost_requested_action_process_matches(
    process_matches: List[Dict[str, Any]],
    *,
    effective_process_templates: List[Dict[str, Any]],
    message_tokens: List[str],
    limit: int,
) -> List[Dict[str, Any]]:
    requested_actions = _requested_process_actions_from_message(message_tokens)
    if not requested_actions:
        return process_matches
    suppress_fastener_candidates = _message_prefers_connector_connection(message_tokens)

    boosted = list(process_matches or [])
    supplemental_candidates = _recommend_processes_from_message_fallback(
        message_tokens=message_tokens,
        effective_process_templates=effective_process_templates,
        limit=max(limit * 4, 20),
    )
    if suppress_fastener_candidates:
        boosted = [
            item for item in boosted
            if not _is_fastener_like_process_candidate(item)
        ]
        supplemental_candidates = [
            item for item in supplemental_candidates
            if not _is_fastener_like_process_candidate(item)
        ]
    supplemental_candidates = _dedupe_process_family_matches(
        [
            *supplemental_candidates,
            *[
                {
                    "processKey": str(process.get("processKey") or "").strip(),
                    "label": process.get("label"),
                    "displayLabel": _resolve_process_display_label(process),
                    "operationLabel": process.get("label"),
                    "partBase": process.get("partBase"),
                    "contextPartBase": process.get("partBase"),
                    "sourceSheet": process.get("sourceSheet"),
                    "score": 1.0,
                    "reason": "요청 액션 보강",
                }
                for process in effective_process_templates
                if str(process.get("processKey") or "").strip()
                and not (
                    suppress_fastener_candidates
                    and _is_fastener_like_process_candidate(process)
                )
                and any(
                    _normalize_chat_text(hint) in " ".join(
                        _normalize_chat_text(value)
                        for value in (
                            process.get("processKey"),
                            process.get("label"),
                            process.get("partBase"),
                            process.get("sourceSheet"),
                        )
                        if value
                    )
                    for keyword in requested_actions
                    for hint in _PROCESS_MESSAGE_HINTS.get(keyword, ())
                )
            ],
        ]
    )

    seen_keys = {
        _normalize_relation_key(item.get("processKey") or item.get("label") or "")
        for item in boosted
        if str(item.get("processKey") or item.get("label") or "").strip()
    }
    prioritized: List[Dict[str, Any]] = []

    for keyword in requested_actions:
        if any(_process_matches_action_keyword(item, keyword) for item in boosted):
            continue
        candidate = next(
            (
                item
                for item in supplemental_candidates
                if _process_matches_action_keyword(item, keyword)
                and _normalize_relation_key(item.get("processKey") or item.get("label") or "") not in seen_keys
            ),
            None,
        )
        if candidate is None:
            continue
        candidate_key = _normalize_relation_key(candidate.get("processKey") or candidate.get("label") or "")
        if candidate_key:
            seen_keys.add(candidate_key)
        prioritized.append(candidate)

    if not prioritized:
        return process_matches

    return [*prioritized, *boosted][:limit]


def _prioritize_requested_action_part_base_matches(
    process_matches: List[Dict[str, Any]],
    *,
    message_tokens: List[str],
    limit: int,
) -> List[Dict[str, Any]]:
    requested_actions = _requested_process_actions_from_message(message_tokens)
    if not requested_actions:
        return process_matches

    def action_part_base_score(item: Dict[str, Any]) -> tuple[int, float, str]:
        part_base = str(item.get("partBase") or item.get("contextPartBase") or "").strip()
        process_score = float(item.get("score") or 0)
        part_base_match = 0
        label_match = 0

        for keyword in requested_actions:
            hints = _PROCESS_MESSAGE_HINTS.get(keyword, ())
            normalized_hints = [_normalize_chat_text(keyword), *(_normalize_chat_text(hint) for hint in hints)]
            if any(hint and hint in _normalize_chat_text(part_base) for hint in normalized_hints):
                part_base_match = 1
            if _process_matches_action_keyword(item, keyword):
                label_match = 1

        return (part_base_match, label_match, process_score, str(part_base or item.get("label") or item.get("processKey") or ""))

    ranked = sorted(
        list(process_matches or []),
        key=lambda item: (
            -action_part_base_score(item)[0],
            -action_part_base_score(item)[1],
            -action_part_base_score(item)[2],
            action_part_base_score(item)[3],
        ),
    )

    # 사용자가 액션을 명시했고, 그 액션이 partBase 자체에 반영된 후보가 있으면
    # 해당 후보군만 남겨서 "부품 기준" 추천 의도를 우선한다.
    part_base_matched = [item for item in ranked if action_part_base_score(item)[0] > 0]
    if part_base_matched:
        return part_base_matched[:limit]

    return ranked[:limit]


def _build_part_message_candidates(part: Dict[str, Any]) -> List[str]:
    candidates: List[str] = []
    values = [
        part.get("displayLabel"),
        part.get("partBase"),
        part.get("partName"),
        part.get("partId"),
        part.get("nodeName"),
    ]
    for value in values:
        normalized = _normalize_chat_text(value)
        if normalized:
            candidates.append(normalized)

    family = _part_family_key(
        part.get("displayLabel")
        or part.get("partBase")
        or part.get("partName")
        or part.get("partId")
        or part.get("nodeName")
        or ""
    )
    for hint in _CHAT_PART_FAMILY_HINTS.get(family, ()):
        normalized = _normalize_chat_text(hint)
        if normalized:
            candidates.append(normalized)

    return list(dict.fromkeys(candidates))


def _find_part_message_order_index(part: Dict[str, Any], message: str) -> float:
    normalized_message = _normalize_chat_text(message)
    if not normalized_message:
        return float("inf")

    best_index = float("inf")
    for candidate in _build_part_message_candidates(part):
        index = normalized_message.find(candidate)
        if index >= 0 and index < best_index:
            best_index = index
    return best_index


def _build_requested_extraction_process_match(
    effective_process_templates: List[Dict[str, Any]],
) -> Optional[Dict[str, Any]]:
    ranked_candidates: List[Tuple[int, Dict[str, Any]]] = []

    for process in effective_process_templates or []:
        if not _process_matches_action_keyword(process, "취출"):
            continue

        part_base = str(process.get("partBase") or "").strip()
        source_sheet = str(process.get("sourceSheet") or "").strip()
        label = str(process.get("label") or process.get("processKey") or "").strip()
        normalized_part_base = _normalize_chat_text(part_base)
        priority = 0

        if "부품(단품)류 취출작업" in part_base or "부품 단품류 취출작업" in part_base:
            priority += 100
        if "취출" in normalized_part_base:
            priority += 50
        if source_sheet == "공통 DB":
            priority += 20
        if source_sheet == "표준 동작":
            priority += 10

        ranked_candidates.append(
            (
                priority,
                {
                    "processKey": str(process.get("processKey") or "").strip(),
                    "label": process.get("label"),
                    "displayLabel": _resolve_process_display_label(process),
                    "operationLabel": process.get("label"),
                    "partBase": process.get("partBase"),
                    "contextPartBase": process.get("partBase"),
                    "sourceSheet": process.get("sourceSheet"),
                    "score": 9999.0,
                    "reason": "취출 요청 우선 추천",
                },
            )
        )

    if not ranked_candidates:
        return None

    ranked_candidates.sort(
        key=lambda item: (
            -item[0],
            str(item[1].get("partBase") or ""),
            str(item[1].get("label") or item[1].get("processKey") or ""),
        )
    )
    return ranked_candidates[0][1]


def _filter_process_matches_by_requested_context(
    process_matches: List[Dict[str, Any]],
    requested_families: set,
    message_tokens: List[str],
) -> List[Dict[str, Any]]:
    requested_actions = _requested_process_actions_from_message(message_tokens)
    requested_fastener = _message_requests_fastener(message_tokens)
    requested_fastener_types = _requested_fastener_types_from_message(message_tokens)
    if not requested_families and not requested_actions:
        return process_matches

    filtered: List[Dict[str, Any]] = []
    for item in process_matches or []:
        if requested_fastener and _fastener_candidate_matches_requested_types(item, requested_fastener_types):
            filtered.append(item)
            continue
        context_families = _detect_part_families_in_values(
            _read_candidate_values(
                item,
                ("partBase", "contextPartBase", "sourceSheet"),
            )
        )
        if requested_families and context_families and not context_families.intersection(requested_families):
            continue
        part_match = _part_matches_requested_families(item, requested_families, message_tokens)
        action_match = _process_matches_requested_action(item, message_tokens)
        if part_match or action_match:
            filtered.append(item)
    return filtered


def _load_options_from_excel(part_base: str, source_sheet: Optional[str]) -> List[str]:
    part_base = str(part_base or "").strip()
    if not part_base or not EXCEL_DB_PATH.exists():
        return []

    option_index = _get_sequence_option_index()
    rows_by_sheet = option_index.get("rowsBySheet") or {}
    sheet_names = _resolve_option_sheet_names(option_index, source_sheet)

    options: List[str] = []
    for sheet_name in sheet_names:
        for item in rows_by_sheet.get(sheet_name, []):
            if item.get("partBase") != part_base:
                continue

            option_value = item.get("option")
            if option_value:
                options.append(option_value)

    return list(dict.fromkeys(options))


def _load_process_options_from_excel(
    part_base: str,
    process_label: str,
    source_sheet: Optional[str],
) -> List[str]:
    part_base = str(part_base or "").strip()
    process_label = str(process_label or "").strip()
    if not part_base or not process_label or not EXCEL_DB_PATH.exists():
        return []

    option_index = _get_sequence_option_index()
    rows_by_sheet = option_index.get("rowsBySheet") or {}
    sheet_names = _resolve_option_sheet_names(option_index, source_sheet)

    options: List[str] = []
    for sheet_name in sheet_names:
        part_level_options: List[str] = []
        for item in rows_by_sheet.get(sheet_name, []):
            if item.get("partBase") != part_base:
                continue

            option_value = item.get("option")
            if option_value:
                part_level_options.append(option_value)

            if not _process_labels_match(item.get("processLabel"), process_label):
                continue

            if option_value:
                options.append(option_value)

        if not options and part_level_options:
            options.extend(part_level_options)
        elif part_level_options:
            options.extend(part_level_options)

    return list(dict.fromkeys(options))


def _sequence_process_template_to_dict(template: Any) -> Dict[str, Any]:
    if isinstance(template, dict):
        return {
            "processKey": template.get("processKey"),
            "processType": template.get("processType") or "STANDARD",
            "label": template.get("label"),
            "sourceSheet": template.get("sourceSheet"),
            "partBase": template.get("partBase"),
        }

    return {
        "processKey": getattr(template, "processKey", None),
        "processType": getattr(template, "processType", None) or "STANDARD",
        "label": getattr(template, "label", None),
        "sourceSheet": getattr(template, "sourceSheet", None),
        "partBase": getattr(template, "partBase", None),
    }


@lru_cache(maxsize=4)
def _get_cached_process_templates(excel_path_str: str, mtime: float) -> Dict[str, Any]:
    excel_path = Path(excel_path_str)
    wb = load_workbook(excel_path, data_only=True)

    processes = []
    seen = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for item in _iter_sequence_sheet_rows(ws, 2):
            part_base = str(item.get("partBase") or "").strip()
            process_label = str(item.get("processLabel") or "").strip()
            if not part_base or not process_label:
                continue
            if not _is_valid_process_label(process_label, part_base):
                continue
            key = (sheet_name, part_base, process_label)

            if key in seen:
                continue
            seen.add(key)

            processes.append({
                "processKey": f"{sheet_name}:{part_base}:{process_label}",
                "processType": "STANDARD",
                "label": process_label,
                "sourceSheet": sheet_name,
                "partBase": part_base,
            })

    return {
        "source": "assembly-db",
        "processes": processes,
        "count": len(processes),
    }


def _resolve_effective_process_templates(req_templates: Optional[List[Any]] = None) -> List[Dict[str, Any]]:
    if req_templates:
        processes = [
            _sequence_process_template_to_dict(template)
            for template in req_templates
        ]
        return [
            process
            for process in processes
            if str(process.get("processKey") or "").strip()
            and str(process.get("label") or "").strip()
        ]

    return get_process_templates().get("processes", [])


def _build_chat_reply(
    message: str,
    part_matches: List[Dict[str, Any]],
    process_matches: List[Dict[str, Any]],
    option_matches: List[Dict[str, Any]],
) -> str:
    lines = [f"입력 내용을 바탕으로 추천을 정리했습니다: {message.strip()}"]

    if part_matches:
        lines.append(
            "추천 부품: "
            + ", ".join(
                item.get("partBase")
                or item.get("partName")
                or item.get("nodeName")
                or "이름 없음"
                for item in part_matches
            )
        )

    if process_matches:
        lines.append(
            "추천 공정: "
            + ", ".join(item.get("label") or item.get("processKey") or "이름 없음" for item in process_matches)
        )

    if option_matches:
        lines.append(
            "추천 옵션: "
            + "; ".join(
                f"{item.get('targetKey')}: {', '.join((item.get('options') or [])[:2])}"
                for item in option_matches
                if item.get("options")
            )
        )

    if not part_matches and not process_matches:
        lines.append("직접 겹치는 키워드는 적어서 현재 선택 부품과 전체 공정 템플릿 기준으로 좁혀서 추천했습니다.")

    return "\n".join(lines)


def _part_matches_lighting_module_alias(value: Any) -> bool:
    family = _part_family_key(value)
    if family == "LDM":
        return True
    normalized = _normalize_relation_key(value)
    return any(
        keyword in normalized
        for keyword in (
            "LDM",
            "LED DRIVE MODULE",
            "주광 LED 모듈",
            "광모듈",
            "광 모듈",
        )
    )


def _resolve_chat_part_display_label(part: Dict[str, Any], message_tokens: List[str]) -> Optional[str]:
    return None


def _resolve_chat_part_canonical_entry(
    part: Dict[str, Any],
    message_tokens: List[str],
    effective_process_templates: List[Dict[str, Any]],
) -> Optional[Dict[str, str]]:
    token_blob = " ".join(_normalize_chat_text(token) for token in message_tokens if token).strip()
    if not token_blob:
        return None

    part_family = _part_family_key(
        part.get("partBase") or part.get("partName") or part.get("partId") or part.get("nodeName") or ""
    )
    if part_family != "LDM":
        return None

    canonical_target = ""
    for alias, target in _CHAT_PART_CANONICAL_TARGETS.items():
        if _normalize_chat_text(alias) in token_blob:
            canonical_target = str(target or "").strip()
            break
    if not canonical_target:
        return None

    db_part_lookup = _build_db_part_lookup(effective_process_templates or [])
    matches = db_part_lookup.get("LDM") or []
    normalized_target = _normalize_relation_key(canonical_target)
    for item in matches:
        if _normalize_relation_key(item.get("partBase")) == normalized_target:
            return item
    return None


def _apply_chat_part_canonical_resolution(
    parts: List[Dict[str, Any]],
    message_tokens: List[str],
    effective_process_templates: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    resolved_parts: List[Dict[str, Any]] = []
    for part in parts or []:
        canonical_entry = _resolve_chat_part_canonical_entry(
            part,
            message_tokens,
            effective_process_templates,
        )
        if canonical_entry is None:
            resolved_parts.append(part)
            continue
        canonical_part_base = str(canonical_entry.get("partBase") or "").strip()
        canonical_source_sheet = str(canonical_entry.get("sourceSheet") or "").strip()
        resolved_parts.append(
            {
                **part,
                "nodeName": canonical_part_base or part.get("nodeName"),
                "partBase": canonical_part_base or part.get("partBase"),
                "partName": canonical_part_base or part.get("partName"),
                "partId": canonical_part_base or part.get("partId"),
                "sourceSheet": canonical_source_sheet or part.get("sourceSheet"),
            }
        )
    return resolved_parts


def _apply_chat_part_display_labels(
    parts: List[Dict[str, Any]],
    message_tokens: List[str],
) -> List[Dict[str, Any]]:
    labeled_parts: List[Dict[str, Any]] = []
    for part in parts or []:
        display_label = _resolve_chat_part_display_label(part, message_tokens)
        if display_label:
            labeled_parts.append(
                {
                    **part,
                    "displayLabel": display_label,
                }
            )
            continue
        labeled_parts.append(part)
    return labeled_parts


def _is_fastener_like_part(*values: Any) -> bool:
    haystack = " ".join(_normalize_chat_text(value) for value in values if value).strip()
    if not haystack:
        return False
    if re.search(r"\bT\s*/?\s*SCREW\b", haystack):
        return True
    return any(hint in haystack for hint in _CHAT_FASTENER_HINTS)


def _message_requests_fastener(message_tokens: List[str]) -> bool:
    token_blob = " ".join(_normalize_chat_text(token) for token in message_tokens if token).strip()
    if not token_blob:
        return False
    if _message_prefers_connector_connection(message_tokens):
        return False
    return any(
        hint in token_blob
        for hint in (
            "SCREW",
            "T/SCREW",
            "M/SCREW",
            "체결",
            "스크류",
            "볼트",
            "나사",
            "FASTENER",
        )
    )


def _requested_fastener_types_from_message(message_tokens: List[str]) -> set[str]:
    token_blob = " ".join(_normalize_chat_text(token) for token in message_tokens if token).strip()
    if not token_blob:
        return set()
    if _message_prefers_connector_connection(message_tokens):
        return set()

    requested: set[str] = set()
    if any(hint in token_blob for hint in ("SCREW", "T/SCREW", "M/SCREW", "스크류", "나사")):
        requested.add("SCREW")
    if any(hint in token_blob for hint in ("BOLT", "볼트")):
        requested.add("BOLT")
    if any(hint in token_blob for hint in ("NUT", "너트")):
        requested.add("NUT")
    if "체결" in token_blob and not requested:
        requested.add("GENERIC")
    return requested


def _message_requests_explicit_screw_text(message_text: str) -> bool:
    normalized = _normalize_chat_text(message_text)
    if not normalized:
        return False
    return any(
        hint in normalized
        for hint in ("스크류", "SCREW", "T/SCREW", "M/SCREW", "나사", "볼트")
    )


def _message_prefers_connector_connection(message_tokens: List[str]) -> bool:
    token_blob = " ".join(_normalize_chat_text(token) for token in message_tokens if token).strip()
    if not token_blob:
        return False
    if not any(
        hint in token_blob
        for hint in (
            "CONNECTOR",
            "커넥터",
            "연결",
            "WIRE",
            "WIRING",
            "커넥팅",
            "CONNECTING",
            "COUPLER",
            "SOCKET",
            "HARNESS",
            "하네스",
        )
    ):
        return False
    if any(hint in token_blob for hint in ("SCREW", "T/SCREW", "M/SCREW", "스크류", "볼트", "나사", "FASTENER")):
        return False
    return True


def _filter_out_fastener_for_connector_connection(
    process_matches: List[Dict[str, Any]],
    message_tokens: List[str],
) -> List[Dict[str, Any]]:
    if not _message_prefers_connector_connection(message_tokens):
        return list(process_matches or [])
    return [
        item
        for item in (process_matches or [])
        if not _is_fastener_like_process_candidate(item)
    ]


def _boost_connector_connection_process_matches(
    process_matches: List[Dict[str, Any]],
    *,
    effective_process_templates: List[Dict[str, Any]],
    message_tokens: List[str],
    limit: int,
) -> List[Dict[str, Any]]:
    if not _message_prefers_connector_connection(message_tokens):
        return list(process_matches or [])

    boosted = _filter_out_fastener_for_connector_connection(process_matches, message_tokens)
    connector_candidates: List[Dict[str, Any]] = []
    for template in effective_process_templates or []:
        haystack = " ".join(
            _normalize_chat_text(value)
            for value in (
                template.get("processKey"),
                template.get("label"),
                template.get("partBase"),
                template.get("sourceSheet"),
            )
            if value
        ).strip()
        if not haystack:
            continue
        if not any(
            hint in haystack
            for hint in (
                "CONNECTOR",
                "커넥터",
                "연결",
                "WIRE",
                "WIRING",
                "커넥팅",
                "CONNECTING",
                "COUPLER",
                "SOCKET",
                "HARNESS",
                "하네스",
            )
        ):
            continue
        connector_candidates.append(
            _build_process_match_from_template(
                template,
                score=999.0,
                reason="커넥터 연결 요청 보정",
            )
        )

    merged = _dedupe_process_family_matches([*connector_candidates, *boosted])
    return merged[:limit]


def _fastener_candidate_matches_requested_types(item: Any, requested_fastener_types: set[str]) -> bool:
    if not requested_fastener_types:
        return _is_fastener_like_process_candidate(item)

    if isinstance(item, dict):
        values = (
            item.get("processKey"),
            item.get("label"),
            item.get("displayLabel"),
            item.get("operationLabel"),
            item.get("partBase"),
            item.get("contextPartBase"),
        )
    else:
        values = (
            getattr(item, "processKey", None),
            getattr(item, "label", None),
            getattr(item, "displayLabel", None),
            getattr(item, "operationLabel", None),
            getattr(item, "partBase", None),
            getattr(item, "contextPartBase", None),
        )
    haystack = " ".join(_normalize_chat_text(value) for value in values if value).strip()
    if not haystack:
        return False

    if "GENERIC" in requested_fastener_types:
        return _is_fastener_like_process_candidate(item)
    if "SCREW" in requested_fastener_types and ("SCREW" in haystack or "T/SCREW" in haystack or "M/SCREW" in haystack):
        return True
    if "BOLT" in requested_fastener_types and "BOLT" in haystack:
        return True
    if "NUT" in requested_fastener_types and "NUT" in haystack:
        return True
    return False


def _is_fastener_like_process_candidate(item: Any) -> bool:
    if isinstance(item, dict):
        values = (
            item.get("processKey"),
            item.get("label"),
            item.get("displayLabel"),
            item.get("operationLabel"),
            item.get("partBase"),
            item.get("contextPartBase"),
        )
    else:
        values = (
            getattr(item, "processKey", None),
            getattr(item, "label", None),
            getattr(item, "displayLabel", None),
            getattr(item, "operationLabel", None),
            getattr(item, "partBase", None),
            getattr(item, "contextPartBase", None),
        )
    return _is_fastener_like_part(*values)


def _boost_fastener_process_matches(
    process_matches: List[Dict[str, Any]],
    *,
    effective_process_templates: List[Dict[str, Any]],
    requested_part_families: set,
    message_tokens: List[str],
    limit: int,
) -> List[Dict[str, Any]]:
    if not _message_requests_fastener(message_tokens):
        return process_matches
    requested_fastener_types = _requested_fastener_types_from_message(message_tokens)

    if any(_fastener_candidate_matches_requested_types(item, requested_fastener_types) for item in (process_matches or [])):
        return process_matches

    fastener_candidates: List[Dict[str, Any]] = []
    for template in effective_process_templates or []:
        if not _fastener_candidate_matches_requested_types(template, requested_fastener_types):
            continue
        if requested_part_families and "GENERIC" in requested_fastener_types:
            context_families = _detect_part_families_in_values(
                (
                    template.get("partBase"),
                    template.get("sourceSheet"),
                    template.get("label"),
                )
            )
            if context_families and not context_families.intersection(requested_part_families):
                continue
        fastener_candidates.append(
            _build_process_match_from_template(
                template,
                score=999.0,
                reason="스크류/체결 요청 보정",
            )
        )

    if not fastener_candidates:
        return process_matches

    merged = _dedupe_process_family_matches([*process_matches, *fastener_candidates])
    if any(_fastener_candidate_matches_requested_types(item, requested_fastener_types) for item in merged[:limit]):
        return merged

    # Ensure at least one fastener candidate survives the final slice.
    best_fastener = next(
        (item for item in merged if _fastener_candidate_matches_requested_types(item, requested_fastener_types)),
        None,
    )
    if not best_fastener:
        return merged
    head = [item for item in merged if not _is_fastener_like_process_candidate(item)][: max(limit - 1, 0)]
    return [*head, best_fastener]


def _candidate_overlaps_tokens(candidate_text: Any, tokens: List[str]) -> bool:
    haystack = _normalize_chat_text(candidate_text)
    if not haystack:
        return False
    return any(token and token in haystack for token in tokens)


def _extract_requested_materials(message_tokens: List[str]) -> set[str]:
    requested = set()
    token_text = " ".join(message_tokens)
    if any(keyword in token_text for keyword in ("PLASTIC", "플라스틱")):
        requested.add("plastic")
    if any(keyword in token_text for keyword in ("PCB", "기판")):
        requested.add("pcb")
    if any(keyword in token_text for keyword in ("비철", "금속", "METAL")):
        requested.add("metal")
    return requested


def _message_has_direct_candidate_overlap(
    message_tokens: List[str],
    candidate_parts: List[Any],
    process_templates: List[Dict[str, Any]],
) -> bool:
    if not message_tokens:
        return False

    for part in candidate_parts or []:
        score, _matched_tokens = _score_candidate(
            message_tokens,
            getattr(part, "partBase", None) if not isinstance(part, dict) else part.get("partBase"),
            getattr(part, "partName", None) if not isinstance(part, dict) else part.get("partName"),
            getattr(part, "partId", None) if not isinstance(part, dict) else part.get("partId"),
            getattr(part, "nodeName", None) if not isinstance(part, dict) else part.get("nodeName"),
        )
        if score > 0:
            return True

    for process in process_templates or []:
        score, _matched_tokens = _score_candidate(
            message_tokens,
            process.get("label"),
            process.get("processKey"),
            process.get("partBase"),
        )
        if score > 0:
            return True

    return False


def _message_is_sequence_chat_relevant(
    message_tokens: List[str],
    candidate_parts: List[Any],
    process_templates: List[Dict[str, Any]],
) -> bool:
    if not message_tokens:
        return False
    if _requested_part_families_from_message(message_tokens):
        return True
    if _requested_process_actions_from_message(message_tokens):
        return True
    if _extract_requested_materials(message_tokens):
        return True
    if _message_requests_fastener(message_tokens):
        return True
    if _message_has_direct_candidate_overlap(message_tokens, candidate_parts, process_templates):
        return True
    return False


def _detect_process_material(process_key: Any, label: Any) -> Optional[str]:
    haystack = f"{_normalize_chat_text(process_key)} {_normalize_chat_text(label)}".strip()
    for material, hints in _PROCESS_MATERIAL_HINTS.items():
        if any(hint in haystack for hint in hints):
            return material
    return None


def _apply_process_material_adjustment(
    score: float,
    matched_tokens: List[str],
    message_tokens: List[str],
    process_key: Any,
    label: Any,
) -> Tuple[float, List[str]]:
    adjusted_score = float(score)
    adjusted_tokens = list(matched_tokens)
    requested_materials = _extract_requested_materials(message_tokens)
    process_material = _detect_process_material(process_key, label)

    if not requested_materials or not process_material:
        return adjusted_score, adjusted_tokens

    if process_material in requested_materials:
        adjusted_score += 1.25
        return adjusted_score, adjusted_tokens

    adjusted_score -= 1.5
    adjusted_tokens.append(f"{process_material} 감점")
    return adjusted_score, adjusted_tokens


def _process_family_key(process_key: Any, label: Any) -> str:
    haystack = f"{_normalize_chat_text(process_key)} {_normalize_chat_text(label)}".strip()
    if "T/SCREW" in haystack or "SCREW" in haystack:
        return "T/SCREW"
    return str(process_key or label or "").strip().upper()


def _part_family_key(value: Any) -> str:
    normalized = _normalize_relation_key(value)
    if not normalized:
        return ""

    normalized = re.sub(r"\([^)]*\)", " ", normalized)
    normalized = re.sub(r"\[[^\]]*\]", " ", normalized)
    normalized = normalized.replace("-", " ")
    normalized = re.sub(r"[^A-Z0-9가-힣/ ]+", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()

    aliases = {
        "LED DRIVE MODULE": "LDM",
        "L D M": "LDM",
        "H S": "HEAT SINK",
    }
    return aliases.get(normalized, normalized)


def _resolve_process_display_label(process: Dict[str, Any]) -> str:
    source_sheet = str(process.get("sourceSheet") or "").strip()
    part_base = str(process.get("partBase") or "").strip()
    operation_label = str(process.get("label") or process.get("processKey") or "").strip()
    if source_sheet in {"공통 DB", "표준 동작"} and part_base:
        return part_base
    return operation_label or part_base


def _dedupe_process_family_matches(processes: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    best_by_family: Dict[str, Dict[str, Any]] = {}
    ordered_families: List[str] = []

    for item in processes:
        family = _process_family_key(item.get("processKey"), item.get("label"))
        current_best = best_by_family.get(family)
        if current_best is None:
            best_by_family[family] = item
            ordered_families.append(family)
            continue

        current_score = float(current_best.get("score") or 0)
        candidate_score = float(item.get("score") or 0)
        if candidate_score > current_score:
            best_by_family[family] = item

    return [best_by_family[family] for family in ordered_families]


def _filter_process_matches_overlapping_parts(
    processes: List[Dict[str, Any]],
    parts: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    part_keys = {
        _part_family_key(
            item.get("partBase")
            or item.get("partId")
            or item.get("nodeName")
            or item.get("partName")
            or ""
        )
        for item in (parts or [])
        if isinstance(item, dict)
    }
    part_keys = {key for key in part_keys if key}

    if not part_keys:
        return list(processes or [])

    filtered: List[Dict[str, Any]] = []
    for item in processes or []:
        process_keys = {
            _part_family_key(item.get("label")),
            _part_family_key(item.get("displayLabel")),
            _part_family_key(item.get("operationLabel")),
            _part_family_key(item.get("processKey")),
            _part_family_key(item.get("partBase")),
        }
        process_keys = {key for key in process_keys if key}
        if process_keys.intersection(part_keys):
            continue
        filtered.append(item)
    return filtered


def _build_db_part_lookup(
    effective_process_templates: List[Dict[str, Any]],
) -> Dict[str, List[Dict[str, str]]]:
    lookup: Dict[str, List[Dict[str, str]]] = {}
    seen = set()

    for item in effective_process_templates or []:
        part_base = str(item.get("partBase") or "").strip()
        source_sheet = str(item.get("sourceSheet") or "").strip()
        key = _part_family_key(part_base)
        unique_key = (key, source_sheet)
        if not key or unique_key in seen:
            continue
        seen.add(unique_key)
        lookup.setdefault(key, []).append(
            {
                "partBase": part_base,
                "sourceSheet": source_sheet,
            }
        )

    return lookup


def _resolve_db_part_entry(
    candidate_values: List[Any],
    db_part_lookup: Dict[str, List[Dict[str, str]]],
    preferred_source_sheet: Optional[Any] = None,
) -> Optional[Dict[str, str]]:
    preferred_sheet = str(preferred_source_sheet or "").strip()

    for value in candidate_values:
        normalized = _part_family_key(value)
        if not normalized:
            continue
        matches = db_part_lookup.get(normalized) or []
        if not matches:
            continue
        if preferred_sheet:
            for item in matches:
                if str(item.get("sourceSheet") or "").strip() == preferred_sheet:
                    return item
        return matches[0]

    return None


def _build_db_part_recommendation(
    entry: Dict[str, str],
    *,
    reason: str,
    score: float,
) -> Dict[str, Any]:
    part_base = str(entry.get("partBase") or "").strip()
    source_sheet = str(entry.get("sourceSheet") or "").strip()
    return {
        "nodeName": part_base,
        "partBase": part_base,
        "partName": part_base,
        "partId": part_base,
        "sourceSheet": source_sheet or None,
        "score": round(float(score), 3),
        "reason": reason,
    }


def _build_raw_part_recommendation(
    part: Any,
    *,
    reason: str,
    score: float,
) -> Dict[str, Any]:
    read = (lambda key: part.get(key) if isinstance(part, dict) else getattr(part, key, None))
    part_base = str(
        read("partBase")
        or read("partId")
        or read("nodeName")
        or read("partName")
        or ""
    ).strip()
    return {
        "nodeName": str(read("nodeName") or part_base).strip(),
        "partBase": part_base,
        "partName": str(read("partName") or part_base).strip(),
        "partId": str(read("partId") or part_base).strip(),
        "sourceSheet": str(read("sourceSheet") or "").strip() or None,
        "treePath": list(read("treePath") or []),
        "parentName": read("parentName"),
        "score": round(float(score), 3),
        "reason": reason,
    }


def _part_match_key(item: Dict[str, Any]) -> str:
    return _part_family_key(
        item.get("partBase")
        or item.get("partId")
        or item.get("nodeName")
        or item.get("partName")
        or ""
    )


def _dedupe_part_matches(parts: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    deduped: List[Dict[str, Any]] = []
    seen = set()

    for item in parts or []:
        key = _part_match_key(item)
        if not key or key in seen:
            continue
        seen.add(key)
        deduped.append(item)

    return deduped


def _normalize_relation_key(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).upper()


def _is_valid_process_label(process_label: Any, part_base: Any = None) -> bool:
    label = " ".join(str(process_label or "").split()).strip()
    if not label:
        return False

    normalized_label = _normalize_relation_key(label)
    normalized_part = _normalize_relation_key(part_base)

    if not normalized_label or set(normalized_label) <= {"-"}:
        return False
    if "OPTION" in normalized_label:
        return False
    if normalized_label in {"-", "--", "---", "---------------------"}:
        return False
    if normalized_part and normalized_label == normalized_part:
        return False
    return True


def _is_process_like_part_base(value: Any, process_label: Any = None) -> bool:
    part_base = " ".join(str(value or "").split()).strip()
    if not part_base:
        return True

    normalized_part = _normalize_relation_key(part_base)
    normalized_process = _normalize_relation_key(process_label)
    if not normalized_part or set(normalized_part) <= {"-"}:
        return True
    if normalized_part in {"AAA", "SUB ASSY 결합".upper()}:
        return True
    if normalized_process and normalized_part == normalized_process:
        return True

    process_like_keywords = (
        "AIR BLOWING",
        "BLOWING",
        "BAR CODE",
        "BARCODE",
        "BAR-CODE",
        "HAND HEAT STACKING",
        "HEAT STACKING",
        "설비 작동",
        "설비 지그에서 취출",
        "설비 스위치",
        "스위치 ON",
        "부품안착동작",
        "부품 안착 동작",
        "부품취출동작",
        "부품 취출 동작",
        "포장지제거동작",
        "포장지 제거 동작",
        "검사동작",
        "검사 동작",
        "조립동작",
        "조립 동작",
        "동작",
        "취출작업",
        "취출 작업",
        "안착작업",
        "안착 작업",
        "결합작업",
        "결합 작업",
        "체결작업",
        "체결 작업",
        "연결작업",
        "연결 작업",
        "로딩작업",
        "로딩 작업",
        "공정",
        "작업",
    )
    return any(_normalize_relation_key(keyword) in normalized_part for keyword in process_like_keywords)


def _match_process_templates_by_label(
    candidate_label: str,
    templates_by_label: Dict[str, List[Dict[str, Any]]],
) -> List[Dict[str, Any]]:
    normalized_candidate = " ".join(str(candidate_label or "").split()).strip()
    if not normalized_candidate:
        return []

    exact_matches = templates_by_label.get(normalized_candidate, [])
    if exact_matches:
        return exact_matches

    relation_candidate = _normalize_relation_key(normalized_candidate)
    loose_matches: List[Dict[str, Any]] = []

    for template_label, template_processes in templates_by_label.items():
        relation_template = _normalize_relation_key(template_label)
        if not relation_template:
            continue
        if (
            relation_candidate == relation_template
            or relation_candidate in relation_template
            or relation_template in relation_candidate
        ):
            loose_matches.extend(template_processes)

    return loose_matches


def _match_process_templates_by_db_fuzzy_label(
    candidate_label: str,
    templates_by_label: Dict[str, List[Dict[str, Any]]],
    *,
    threshold: float = 82.0,
) -> List[Dict[str, Any]]:
    normalized_candidate = " ".join(str(candidate_label or "").split()).strip()
    if not normalized_candidate:
        return []

    relation_candidate = _normalize_relation_key(normalized_candidate)
    best_label = ""
    best_score = -1.0

    for template_label in templates_by_label.keys():
        relation_template = _normalize_relation_key(template_label)
        if not relation_template:
            continue
        score = combined_score(
            rf_score(relation_candidate, relation_template),
            jw_score(relation_candidate, relation_template),
        )
        if score > best_score:
            best_score = score
            best_label = template_label

    if best_label and best_score >= threshold:
        return templates_by_label.get(best_label, [])

    return []


def _filter_process_templates_for_selected_parts(
    templates: List[Dict[str, Any]],
    selected_parts: List[Any],
) -> List[Dict[str, Any]]:
    if not templates:
        return []

    selected_part_keys = {
        _normalize_relation_key(
            getattr(part, "partBase", None)
            or getattr(part, "partId", None)
            or getattr(part, "nodeName", None)
            or getattr(part, "partName", None)
            or ""
        )
        for part in (selected_parts or [])
        if str(
            getattr(part, "partBase", None)
            or getattr(part, "partId", None)
            or getattr(part, "nodeName", None)
            or getattr(part, "partName", None)
            or ""
        ).strip()
    }
    if not selected_part_keys:
        return templates

    direct_matches = [
        template
        for template in templates
        if _normalize_relation_key(template.get("partBase")) in selected_part_keys
    ]
    if direct_matches:
        return direct_matches

    loose_matches = []
    for template in templates:
        template_key = _normalize_relation_key(template.get("partBase"))
        if not template_key:
            continue
        if any(
            template_key in selected_key or selected_key in template_key
            for selected_key in selected_part_keys
        ):
            loose_matches.append(template)
    if loose_matches:
        return loose_matches

    return templates


def _emit_sequence_recommend_log(stage: str, **payload: Any) -> None:
    return


def _filter_process_like_selected_parts(
    selected_parts: List[Any],
    effective_process_templates: List[Dict[str, Any]],
) -> List[Any]:
    process_labels = {
        _normalize_relation_key(item.get("label"))
        for item in (effective_process_templates or [])
        if str(item.get("label") or "").strip()
    }
    filtered_parts: List[Any] = []
    for part in selected_parts or []:
        candidate_values = [
            getattr(part, "partBase", None),
            getattr(part, "partId", None),
            getattr(part, "nodeName", None),
            getattr(part, "partName", None),
        ]
        normalized_candidates = {
            _normalize_relation_key(value)
            for value in candidate_values
            if str(value or "").strip()
        }
        if normalized_candidates.intersection(process_labels):
            _emit_sequence_recommend_log(
                "SKIP_PROCESS_LIKE_SELECTED_PART",
                candidate=list(normalized_candidates),
            )
            continue
        filtered_parts.append(part)
    return filtered_parts


def _normalize_selected_part_ids(selected_parts: List[Any]) -> List[str]:
    normalized_ids: List[str] = []
    for part in selected_parts or []:
        value = (
            getattr(part, "partBase", None)
            or getattr(part, "partId", None)
            or getattr(part, "nodeName", None)
            or getattr(part, "partName", None)
            or ""
        )
        normalized = " ".join(str(value).split()).strip()
        if normalized:
            normalized_ids.append(normalized)
    return normalized_ids


def _build_selected_part_matches(
    selected_parts: List[Any],
    limit: int,
    *,
    effective_process_templates: Optional[List[Dict[str, Any]]] = None,
) -> List[Dict[str, Any]]:
    recommended_parts: List[Dict[str, Any]] = []
    db_part_lookup = _build_db_part_lookup(effective_process_templates or [])

    for part in selected_parts or []:
        entry = _resolve_db_part_entry(
            [
                getattr(part, "partBase", None),
                getattr(part, "partId", None),
                getattr(part, "nodeName", None),
                getattr(part, "partName", None),
            ],
            db_part_lookup,
            getattr(part, "sourceSheet", None),
        )
        if entry is None:
            continue
        recommended_parts.append(
            _build_db_part_recommendation(
                entry,
                reason="현재 선택된 부품",
                score=1.0,
            )
        )
        if len(recommended_parts) >= limit:
            break

    return recommended_parts


def _merge_part_recommendations(
    primary_matches: List[Dict[str, Any]],
    secondary_matches: List[Dict[str, Any]],
    limit: int,
) -> List[Dict[str, Any]]:
    merged: List[Dict[str, Any]] = []
    seen = set()

    for item in [*(primary_matches or []), *(secondary_matches or [])]:
        key = _part_match_key(item)
        if not key or key in seen:
            continue
        seen.add(key)
        merged.append(item)
        if len(merged) >= limit:
            break

    return merged


def _build_candidate_part_matches(
    candidate_parts: List[Any],
    message_tokens: List[str],
    limit: int,
    effective_process_templates: Optional[List[Dict[str, Any]]] = None,
) -> List[Dict[str, Any]]:
    ranked: List[Dict[str, Any]] = []
    db_part_lookup = _build_db_part_lookup(effective_process_templates or [])
    allow_fasteners = _message_requests_fastener(message_tokens)
    process_labels = {
        _normalize_relation_key(item.get("label"))
        for item in (effective_process_templates or [])
        if str(item.get("label") or "").strip()
    }

    for part in candidate_parts or []:
        if not allow_fasteners and _is_fastener_like_part(
            getattr(part, "partBase", None),
            getattr(part, "partName", None),
            getattr(part, "partId", None),
            getattr(part, "nodeName", None),
        ):
            continue
        candidate_values = [
            getattr(part, "partBase", None),
            getattr(part, "partName", None),
            getattr(part, "partId", None),
            getattr(part, "nodeName", None),
        ]
        normalized_candidates = {
            _normalize_relation_key(value)
            for value in candidate_values
            if str(value or "").strip()
        }
        score, matched_tokens = _score_candidate(
            message_tokens,
            getattr(part, "partBase", None),
            getattr(part, "partName", None),
            getattr(part, "partId", None),
            getattr(part, "nodeName", None),
            " ".join(getattr(part, "treePath", []) or []),
            getattr(part, "parentName", None),
        )
        if score <= 0:
            continue

        db_part_entry = _resolve_db_part_entry(
            candidate_values,
            db_part_lookup,
            getattr(part, "sourceSheet", None),
        )
        if normalized_candidates.intersection(process_labels):
            continue
        if db_part_entry is None:
            continue

        ranked.append(
            _build_db_part_recommendation(
                db_part_entry,
                reason=", ".join(matched_tokens) if matched_tokens else "자연어와 연관된 부품",
                score=float(score),
            )
        )

    ranked.sort(
        key=lambda item: (
            -float(item.get("score") or 0),
            str(item.get("partBase") or item.get("partName") or item.get("nodeName") or ""),
        )
    )
    return _dedupe_part_matches(ranked)[:limit]


def _part_match_has_direct_message_overlap(item: Dict[str, Any], message_tokens: List[str]) -> bool:
    if not item or not message_tokens:
        return False
    requested_families = _requested_part_families_from_message(message_tokens)
    item_families = _detect_part_families_in_values(
        (
            item.get("displayLabel"),
            item.get("partBase"),
            item.get("partName"),
            item.get("partId"),
            item.get("nodeName"),
        )
    )
    if requested_families and item_families.intersection(requested_families):
        return True
    return any(
        _candidate_overlaps_tokens(
            item.get(key),
            message_tokens,
        )
        for key in ("displayLabel", "partBase", "partName", "partId", "nodeName")
    )


def _prioritize_direct_message_part_matches(
    part_matches: List[Dict[str, Any]],
    message_tokens: List[str],
) -> List[Dict[str, Any]]:
    deduped = _dedupe_part_matches(part_matches or [])
    if not deduped or not message_tokens:
        return deduped

    direct_matches = [
        item for item in deduped if _part_match_has_direct_message_overlap(item, message_tokens)
    ]
    if not direct_matches:
        return deduped

    direct_keys = {_part_match_key(item) for item in direct_matches}
    remainder = [item for item in deduped if _part_match_key(item) not in direct_keys]
    return [*direct_matches, *remainder]


def _build_direct_message_part_matches(
    candidate_parts: List[Any],
    message_tokens: List[str],
    *,
    effective_process_templates: List[Dict[str, Any]],
    limit: int,
) -> List[Dict[str, Any]]:
    ranked_matches = _build_candidate_part_matches(
        candidate_parts,
        message_tokens,
        limit=max(limit, 10),
        effective_process_templates=effective_process_templates,
    )
    direct_matches = [
        item for item in ranked_matches if _part_match_has_direct_message_overlap(item, message_tokens)
    ]
    return _dedupe_part_matches(direct_matches)[:limit]


def _build_fast_path_process_matches(
    selected_parts: List[Dict[str, Any]],
    *,
    effective_process_templates: List[Dict[str, Any]],
    message_tokens: List[str],
    requested_part_families: Set[str],
    limit: int,
) -> List[Dict[str, Any]]:
    if not selected_parts or not effective_process_templates:
        return []

    selected_part_sources = [
        SimpleNamespace(
            nodeName=item.get("nodeName"),
            partBase=item.get("partBase"),
            partName=item.get("partName"),
            partId=item.get("partId"),
            sourceSheet=item.get("sourceSheet"),
        )
        for item in (selected_parts or [])
    ]
    matched_templates = _filter_process_templates_for_selected_parts(
        effective_process_templates,
        selected_part_sources,
    )

    ranked: List[Dict[str, Any]] = []
    requested_actions = _requested_process_actions_from_message(message_tokens)
    for template in matched_templates:
        if requested_part_families and not _part_matches_requested_families(
            template,
            requested_part_families,
            message_tokens,
        ):
            continue

        score, reasons = _score_candidate(
            message_tokens,
            template.get("label"),
            template.get("partBase"),
            template.get("sourceSheet"),
            template.get("processKey"),
        )
        if requested_actions and _process_matches_requested_action(template, message_tokens):
            score += 2.5
            reasons.append("요청 액션 직접 일치")

        if score <= 0 and requested_actions:
            continue
        if score <= 0:
            score = 0.25
            reasons.append("선택 부품 템플릿")

        ranked.append(
            _build_process_match_from_template(
                template,
                score=score,
                reason=", ".join(dict.fromkeys(reasons)) or "선택 부품 빠른 추천",
            )
        )

    ranked = _filter_process_matches_by_requested_context(
        _dedupe_process_family_matches(ranked),
        requested_part_families,
        message_tokens,
    )
    ranked = _boost_requested_action_process_matches(
        ranked,
        effective_process_templates=effective_process_templates,
        message_tokens=message_tokens,
        limit=limit,
    )[:limit]
    ranked = _boost_connector_connection_process_matches(
        ranked,
        effective_process_templates=effective_process_templates,
        message_tokens=message_tokens,
        limit=limit,
    )[:limit]
    ranked = _prioritize_requested_action_part_base_matches(
        ranked,
        message_tokens=message_tokens,
        limit=limit,
    )[:limit]
    return ranked


def _recommend_processes_from_message_fallback(
    *,
    message_tokens: List[str],
    effective_process_templates: List[Dict[str, Any]],
    limit: int,
) -> List[Dict[str, Any]]:
    normalized_message = " ".join(message_tokens)
    requested_actions = _requested_process_actions_from_message(message_tokens)
    requested_fastener = _message_requests_fastener(message_tokens)
    requested_fastener_types = _requested_fastener_types_from_message(message_tokens)
    suppress_fastener_candidates = _message_prefers_connector_connection(message_tokens)
    ranked: List[Dict[str, Any]] = []
    seen_process_keys = set()

    for process in effective_process_templates or []:
        process_key = str(process.get("processKey") or "").strip()
        if not process_key or process_key in seen_process_keys:
            continue
        if suppress_fastener_candidates and _is_fastener_like_process_candidate(process):
            continue

        score, reasons = _score_candidate(
            message_tokens,
            process.get("label"),
            process.get("partBase"),
            process.get("sourceSheet"),
            process_key,
        )
        process_haystack = _normalize_chat_text(
            " ".join(
                [
                    str(process.get("label") or "").strip(),
                    str(process.get("partBase") or "").strip(),
                    str(process.get("sourceSheet") or "").strip(),
                    process_key,
                ]
            )
        )

        for keyword, hints in _PROCESS_MESSAGE_HINTS.items():
            if requested_actions and keyword not in requested_actions:
                continue
            if any(_normalize_chat_text(hint) in process_haystack for hint in hints):
                score += 2.0
                reasons.append(f"{keyword} 관련")

        if requested_fastener and _fastener_candidate_matches_requested_types(process, requested_fastener_types):
            score += 2.5
            reasons.append("스크류/체결 관련")

        if score <= 0:
            continue

        seen_process_keys.add(process_key)
        ranked.append(
            {
                "processKey": process_key,
                "label": process.get("label"),
                "displayLabel": _resolve_process_display_label(process),
                "operationLabel": process.get("label"),
                "partBase": process.get("partBase"),
                "contextPartBase": process.get("partBase"),
                "sourceSheet": process.get("sourceSheet"),
                "score": round(float(score), 3),
                "reason": ", ".join(dict.fromkeys(reasons)) or "자연어 기반 공정 추천",
            }
        )

    ranked.sort(
        key=lambda item: (
            -float(item.get("score") or 0),
            str(item.get("partBase") or item.get("label") or item.get("processKey") or ""),
        )
    )
    return _dedupe_process_family_matches(ranked)[:limit]


def _build_part_matches_from_process_recommendations(
    process_matches: List[Dict[str, Any]],
    limit: int,
) -> List[Dict[str, Any]]:
    recommended_parts: List[Dict[str, Any]] = []
    seen = set()

    for process in process_matches:
        part_base = str(process.get("partBase") or process.get("contextPartBase") or "").strip()
        if not part_base:
            continue
        if _is_process_like_part_base(part_base, process.get("label") or process.get("processKey")):
            continue
        part_key = _normalize_relation_key(part_base)
        if not part_key or part_key in seen:
            continue
        seen.add(part_key)
        recommended_parts.append(
            _build_db_part_recommendation(
                {
                    "partBase": part_base,
                    "sourceSheet": str(process.get("sourceSheet") or "").strip(),
                },
                reason="자연어 기준 공정 후보에서 추론한 부품",
                score=float(process.get("score") or 0),
            )
        )
        if len(recommended_parts) >= limit:
            break

    return recommended_parts


def _serialize_chat_part_candidates(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    serialized: List[Dict[str, Any]] = []
    for item in items or []:
        part_key = str(
            item.get("partBase")
            or item.get("nodeName")
            or item.get("partId")
            or item.get("partName")
            or ""
        ).strip()
        if not part_key:
            continue
        serialized.append(
            {
                "partKey": part_key,
                "nodeName": item.get("nodeName"),
                "partBase": item.get("partBase"),
                "partId": item.get("partId"),
                "partName": item.get("partName"),
                "sourceSheet": item.get("sourceSheet"),
                "reason": item.get("reason"),
                "score": item.get("score"),
            }
        )
    return serialized


def _serialize_chat_process_candidates(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    serialized: List[Dict[str, Any]] = []
    for item in items or []:
        process_key = str(item.get("processKey") or "").strip()
        if not process_key:
            continue
        serialized.append(
            {
                "processKey": process_key,
                "label": item.get("label"),
                "displayLabel": item.get("displayLabel"),
                "operationLabel": item.get("operationLabel"),
                "partBase": item.get("partBase"),
                "contextPartBase": item.get("contextPartBase"),
                "sourceSheet": item.get("sourceSheet"),
                "reason": item.get("reason"),
                "score": item.get("score"),
            }
        )
    return serialized


def _build_per_part_chat_recommendations(
    *,
    part_matches: List[Dict[str, Any]],
    effective_process_templates: List[Dict[str, Any]],
    shared_process_matches: Optional[List[Dict[str, Any]]] = None,
    message_text: str,
    message_tokens: List[str],
    requested_part_families: Set[str],
    candidate_limit: int,
    per_part_limit: int,
) -> List[Dict[str, Any]]:
    per_part_recommendations: List[Dict[str, Any]] = []
    shared_process_pool = _dedupe_process_family_matches(list(shared_process_matches or []))
    ordered_parts = _dedupe_part_matches(part_matches)
    extraction_requested = "취출" in _requested_process_actions_from_message(message_tokens)
    requested_fastener = _message_requests_fastener(message_tokens)
    requested_fastener_types = _requested_fastener_types_from_message(message_tokens)
    explicit_screw_requested = _message_requests_explicit_screw_text(message_text)
    extraction_process_match = (
        _build_requested_extraction_process_match(effective_process_templates)
        if extraction_requested
        else None
    )
    extraction_target_part_key = ""

    if extraction_requested and ordered_parts:
        ordered_parts = sorted(
            ordered_parts,
            key=lambda part: _find_part_message_order_index(part, message_text),
        )
        target_part = next(
            (
                part
                for part in reversed(ordered_parts)
                if _find_part_message_order_index(part, message_text) != float("inf")
            ),
            ordered_parts[-1],
        )
        extraction_target_part_key = _part_match_key(target_part)

    for part in ordered_parts[:per_part_limit]:
        part_family = _part_family_key(
            part.get("partBase") or part.get("partId") or part.get("nodeName") or part.get("partName") or ""
        )
        process_matches = [
            process
            for process in shared_process_pool
            if (
                _part_family_key(process.get("partBase") or process.get("contextPartBase") or "") == part_family
                or (
                    requested_fastener
                    and _fastener_candidate_matches_requested_types(process, requested_fastener_types)
                )
            )
        ]

        if not process_matches:
            part_source = SimpleNamespace(
                nodeName=part.get("nodeName"),
                partBase=part.get("partBase"),
                partName=part.get("partName"),
                partId=part.get("partId"),
                sourceSheet=part.get("sourceSheet"),
            )
            process_matches = _recommend_next_processes_from_graph(
                selected_parts=[part_source],
                effective_process_templates=effective_process_templates,
                limit=candidate_limit,
                message_tokens=message_tokens,
                apply_message_adjustment=True,
            )
        if not process_matches:
            process_matches = _recommend_processes_from_message_fallback(
                message_tokens=message_tokens,
                effective_process_templates=effective_process_templates,
                limit=candidate_limit,
            )
        process_matches = _filter_manual_barcode_reading_processes(
            _filter_process_matches_by_requested_context(
                _dedupe_process_family_matches(process_matches),
                requested_part_families,
                message_tokens,
            )
        )[:per_part_limit]
        process_matches = _boost_fastener_process_matches(
            process_matches,
            effective_process_templates=effective_process_templates,
            requested_part_families=requested_part_families,
            message_tokens=message_tokens,
            limit=per_part_limit,
        )[:per_part_limit]
        process_matches = _boost_requested_action_process_matches(
            process_matches,
            effective_process_templates=effective_process_templates,
            message_tokens=message_tokens,
            limit=per_part_limit,
        )[:per_part_limit]
        process_matches = _prioritize_requested_action_part_base_matches(
            process_matches,
            message_tokens=message_tokens,
            limit=per_part_limit,
        )[:per_part_limit]

        if explicit_screw_requested and not any(
            _fastener_candidate_matches_requested_types(item, {"SCREW"})
            for item in process_matches
        ):
            screw_candidates = [
                _build_process_match_from_template(
                    template,
                    score=999.0,
                    reason="스크류 체결 요청 보정",
                )
                for template in (effective_process_templates or [])
                if _fastener_candidate_matches_requested_types(template, {"SCREW"})
            ]
            if screw_candidates:
                process_matches = _dedupe_process_family_matches(
                    [*screw_candidates, *process_matches]
                )[:per_part_limit]

        if (
            extraction_requested
            and extraction_process_match is not None
            and _part_match_key(part) == extraction_target_part_key
        ):
            process_matches = _dedupe_process_family_matches(
                [extraction_process_match, *process_matches]
            )
        process_matches = _filter_process_matches_overlapping_parts(process_matches, [part])[
            :per_part_limit
        ]

        recommended_options: List[Dict[str, Any]] = []
        part_base = str(part.get("partBase") or "").strip()
        if part_base:
            part_options = _load_options_from_excel(part_base, part.get("sourceSheet"))
            if part_options:
                recommended_options.append(
                    {
                        "targetType": "PART",
                        "targetKey": part_base,
                        "sourceSheet": part.get("sourceSheet"),
                        "options": part_options,
                    }
                )

        for process in process_matches:
            process_part_base = str(process.get("partBase") or part.get("partBase") or "").strip()
            if not process_part_base:
                continue
            process_label = process.get("label") or process.get("operationLabel")
            process_options = _load_process_options_from_excel(
                process_part_base,
                process_label,
                process.get("sourceSheet") or part.get("sourceSheet"),
            )
            if not process_options:
                continue
            recommended_options.append(
                {
                    "targetType": "PROCESS",
                    "targetKey": process_label or process.get("processKey") or process_part_base,
                    "sourceSheet": process.get("sourceSheet") or part.get("sourceSheet"),
                    "options": process_options,
                }
            )

        deduped_options: List[Dict[str, Any]] = []
        seen_option_keys = set()
        for item in recommended_options:
            option_key = (item.get("targetType"), item.get("targetKey"))
            if option_key in seen_option_keys:
                continue
            seen_option_keys.add(option_key)
            deduped_options.append(item)

        per_part_recommendations.append(
            {
                "part": part,
                "recommendedProcesses": process_matches,
                "recommendedOptions": deduped_options,
                "reply": "",
            }
        )

    return per_part_recommendations


def _build_chat_part_matches_from_inputs(selected_parts: List[Any]) -> List[Dict[str, Any]]:
    part_matches: List[Dict[str, Any]] = []
    for index, part in enumerate(selected_parts or []):
        part_base = str(getattr(part, "partBase", None) or "").strip()
        source_sheet = str(getattr(part, "sourceSheet", None) or "").strip()
        if not part_base or not source_sheet:
            continue
        part_matches.append(
            _build_db_part_recommendation(
                {
                    "nodeName": getattr(part, "nodeName", None) or part_base,
                    "partId": getattr(part, "partId", None) or part_base,
                    "partName": getattr(part, "partName", None) or part_base,
                    "partBase": part_base,
                    "sourceSheet": source_sheet,
                },
                reason="선택 부품",
                score=1000.0 - index,
            )
        )
    return _dedupe_part_matches(part_matches)


def _rerank_chat_part_matches_with_openai(
    message: str,
    *,
    selected_parts: List[Any],
    part_candidates: List[Dict[str, Any]],
    process_candidates: List[Dict[str, Any]],
    limit: int,
) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]], str]:
    if not part_candidates and not process_candidates:
        return part_candidates, process_candidates, ""

    selected_payload = [
        {
            "nodeName": getattr(item, "nodeName", None),
            "partBase": getattr(item, "partBase", None),
            "partId": getattr(item, "partId", None),
            "partName": getattr(item, "partName", None),
            "sourceSheet": getattr(item, "sourceSheet", None),
        }
        for item in (selected_parts or [])
    ]

    ai_result = generate_sequence_chat_recommendations(
        message=message,
        selected_parts=selected_payload,
        candidate_parts=_serialize_chat_part_candidates(part_candidates),
        candidate_processes=_serialize_chat_process_candidates(process_candidates),
        limit=limit,
    )
    result = dict(ai_result.get("result") or {})

    requested_part_keys = [
        _normalize_relation_key(value)
        for value in (result.get("recommendedPartKeys") or [])
        if str(value or "").strip()
    ]
    requested_process_keys = [
        _normalize_relation_key(value)
        for value in (result.get("recommendedProcessKeys") or [])
        if str(value or "").strip()
    ]

    part_candidates_by_key = {
        _normalize_relation_key(
            item.get("partBase")
            or item.get("nodeName")
            or item.get("partId")
            or item.get("partName")
            or ""
        ): item
        for item in (part_candidates or [])
    }
    process_candidates_by_key = {
        _normalize_relation_key(item.get("processKey") or ""): item
        for item in (process_candidates or [])
    }

    reranked_parts = [
        part_candidates_by_key[key]
        for key in requested_part_keys
        if key in part_candidates_by_key
    ]
    reranked_processes = [
        process_candidates_by_key[key]
        for key in requested_process_keys
        if key in process_candidates_by_key
    ]

    if not reranked_parts:
        reranked_parts = part_candidates[:limit]
    else:
        reranked_parts = _dedupe_part_matches(reranked_parts)[:limit]

    if not reranked_processes:
        reranked_processes = process_candidates[:limit]
    else:
        reranked_processes = _dedupe_process_family_matches(reranked_processes)[:limit]

    reply = str(result.get("reply") or "").strip()
    return reranked_parts, reranked_processes, reply


def _build_chat_part_matches_without_ai(
    message: str,
    *,
    selected_parts: List[Any],
    candidate_parts: List[Any],
    process_candidates: List[Dict[str, Any]],
    limit: int,
) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]], str]:
    selected_matches: List[Dict[str, Any]] = []
    for index, part in enumerate(selected_parts or []):
        part_base = str(getattr(part, "partBase", None) or getattr(part, "partId", None) or getattr(part, "nodeName", None) or getattr(part, "partName", None) or "").strip()
        if not part_base:
            continue
        selected_matches.append(
            _build_db_part_recommendation(
                {
                    "nodeName": getattr(part, "nodeName", None) or part_base,
                    "partId": getattr(part, "partId", None) or part_base,
                    "partName": getattr(part, "partName", None) or part_base,
                    "partBase": getattr(part, "partBase", None) or part_base,
                    "sourceSheet": str(getattr(part, "sourceSheet", None) or "").strip(),
                },
                reason="현재 선택된 부품",
                score=2000.0 - index,
            )
        )

    candidate_part_matches: List[Dict[str, Any]] = []
    for index, part in enumerate(candidate_parts or []):
        part_base = str(getattr(part, "partBase", None) or "").strip()
        source_sheet = str(getattr(part, "sourceSheet", None) or "").strip()
        part_id = str(getattr(part, "partId", None) or "").strip()
        node_name = str(getattr(part, "nodeName", None) or "").strip()
        part_name = str(getattr(part, "partName", None) or "").strip()
        if not part_base or not source_sheet:
            continue
        candidate_part_matches.append(
            _build_db_part_recommendation(
                {
                    "nodeName": node_name or part_base or part_id or part_name,
                    "partId": part_id or part_base or node_name or part_name,
                    "partName": part_name or part_base or part_id or node_name,
                    "partBase": part_base,
                    "sourceSheet": source_sheet,
                    "treePath": list(getattr(part, "treePath", None) or []),
                    "parentName": getattr(part, "parentName", None),
                },
                reason="규칙 기반 채팅 후보",
                score=max(1.0, 1000.0 - index),
            )
        )

    recommended_parts = _dedupe_part_matches([*selected_matches, *candidate_part_matches])[:limit]
    recommended_processes = _dedupe_process_family_matches(process_candidates)[:limit]
    return recommended_parts, recommended_processes, ""


def _should_use_ai_chat_rerank(
    *,
    message_tokens: List[str],
    requested_part_families: Set[str],
    selected_parts: List[Any],
    part_candidates: List[Any],
    process_candidates: List[Dict[str, Any]],
    limit: int,
) -> bool:
    mode = os.getenv("SEQUENCE_CHAT_AI_RERANK", "auto").strip().lower()
    if mode in {"0", "false", "off", "no"}:
        return False
    if mode in {"1", "true", "on", "yes"}:
        return True

    requested_actions = _requested_process_actions_from_message(message_tokens)
    if requested_actions and len(process_candidates) >= limit:
        return False
    if requested_part_families and len(part_candidates) >= min(limit, 2):
        return False
    if selected_parts and process_candidates:
        return False
    return True


def _build_chat_part_matches_with_ai(
    message: str,
    *,
    selected_parts: List[Any],
    candidate_parts: List[Any],
    process_candidates: List[Dict[str, Any]],
    limit: int,
) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]], str]:
    candidate_part_matches: List[Dict[str, Any]] = []
    for index, part in enumerate(candidate_parts or []):
        part_base = str(getattr(part, "partBase", None) or "").strip()
        part_id = str(getattr(part, "partId", None) or "").strip()
        node_name = str(getattr(part, "nodeName", None) or "").strip()
        part_name = str(getattr(part, "partName", None) or "").strip()
        if not any((part_base, part_id, node_name, part_name)):
            continue
        candidate_part_matches.append(
            _build_db_part_recommendation(
                {
                    "nodeName": node_name or part_base or part_id or part_name,
                    "partId": part_id or part_base or node_name or part_name,
                    "partName": part_name or part_base or part_id or node_name,
                    "partBase": part_base or part_id or node_name or part_name,
                    "sourceSheet": str(getattr(part, "sourceSheet", None) or "").strip(),
                    "treePath": list(getattr(part, "treePath", None) or []),
                    "parentName": getattr(part, "parentName", None),
                },
                reason="AI 채팅 후보",
                score=max(1.0, 1000.0 - index),
            )
        )

    return _rerank_chat_part_matches_with_openai(
        message,
        selected_parts=selected_parts,
        part_candidates=_dedupe_part_matches(candidate_part_matches),
        process_candidates=_dedupe_process_family_matches(process_candidates),
        limit=limit,
    )


def _part_match_from_embedding_document(document: Dict[str, Any]) -> Dict[str, Any]:
    part = document.get("item")
    return _build_db_part_recommendation(
        {
            "nodeName": getattr(part, "nodeName", None) if not isinstance(part, dict) else part.get("nodeName"),
            "partId": getattr(part, "partId", None) if not isinstance(part, dict) else part.get("partId"),
            "partName": getattr(part, "partName", None) if not isinstance(part, dict) else part.get("partName"),
            "partBase": getattr(part, "partBase", None) if not isinstance(part, dict) else part.get("partBase"),
            "sourceSheet": getattr(part, "sourceSheet", None) if not isinstance(part, dict) else part.get("sourceSheet"),
            "treePath": getattr(part, "treePath", None) if not isinstance(part, dict) else part.get("treePath"),
            "parentName": getattr(part, "parentName", None) if not isinstance(part, dict) else part.get("parentName"),
        },
        reason=f"bge-m3 유사도 {float(document.get('embeddingScore') or 0):.3f}",
        score=float(document.get("embeddingScore") or 0),
    )


def _process_match_from_embedding_document(document: Dict[str, Any]) -> Dict[str, Any]:
    process = document.get("item")
    read = (lambda key: process.get(key) if isinstance(process, dict) else getattr(process, key, None))
    return {
        "processKey": str(read("processKey") or "").strip(),
        "label": read("label"),
        "displayLabel": _resolve_process_display_label(
            {
                "processKey": read("processKey"),
                "label": read("label"),
                "partBase": read("partBase"),
            }
        ),
        "operationLabel": read("label"),
        "partBase": read("partBase"),
        "contextPartBase": read("partBase"),
        "sourceSheet": read("sourceSheet"),
        "score": float(document.get("embeddingScore") or 0),
        "reason": f"bge-m3 유사도 {float(document.get('embeddingScore') or 0):.3f}",
    }


def _read_match_key(item: Dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = str(item.get(key) or "").strip()
        if value:
            return value
    return ""


def _build_process_match_from_template(
    template: Dict[str, Any],
    *,
    score: float,
    reason: str,
) -> Dict[str, Any]:
    return {
        "processKey": str(template.get("processKey") or "").strip(),
        "label": template.get("label"),
        "displayLabel": _resolve_process_display_label(template),
        "operationLabel": template.get("label"),
        "partBase": template.get("partBase"),
        "contextPartBase": template.get("partBase"),
        "sourceSheet": template.get("sourceSheet"),
        "score": round(float(score), 3),
        "reason": reason,
    }


def _match_process_template_from_graph_candidate(
    candidate: Dict[str, Any],
    process_templates: List[Dict[str, Any]],
) -> Optional[Dict[str, Any]]:
    candidate_values = [
        candidate.get("processKey"),
        candidate.get("processLabel"),
        candidate.get("label"),
    ]
    normalized_values = {
        _normalize_relation_key(value)
        for value in candidate_values
        if str(value or "").strip()
    }
    if not normalized_values:
        return None

    for template in process_templates or []:
        template_values = {
            _normalize_relation_key(template.get("processKey")),
            _normalize_relation_key(template.get("label")),
            _normalize_relation_key(template.get("partBase")),
        }
        if normalized_values.intersection({value for value in template_values if value}):
            return template
    return None


def _build_candidate_part_lookup(candidate_parts: List[Any]) -> Dict[str, Any]:
    lookup: Dict[str, Any] = {}
    for part in candidate_parts or []:
        for value in (
            getattr(part, "partBase", None),
            getattr(part, "partId", None),
            getattr(part, "partName", None),
            getattr(part, "nodeName", None),
        ):
            key = _normalize_relation_key(value)
            if key and key not in lookup:
                lookup[key] = part
    return lookup


def _part_match_from_candidate_part(
    part: Any,
    *,
    reason: str,
    score: float,
) -> Dict[str, Any]:
    return _build_db_part_recommendation(
        {
            "nodeName": str(getattr(part, "nodeName", None) or "").strip(),
            "partId": str(getattr(part, "partId", None) or "").strip(),
            "partName": str(getattr(part, "partName", None) or "").strip(),
            "partBase": str(
                getattr(part, "partBase", None)
                or getattr(part, "partId", None)
                or getattr(part, "nodeName", None)
                or getattr(part, "partName", None)
                or ""
            ).strip(),
            "sourceSheet": str(getattr(part, "sourceSheet", None) or "").strip(),
        },
        reason=reason,
        score=score,
    )


def _match_candidate_part_from_graph_candidate(
    candidate: Dict[str, Any],
    candidate_part_lookup: Dict[str, Any],
) -> Optional[Any]:
    for value in (
        candidate.get("partBase"),
        candidate.get("partId"),
        candidate.get("label"),
        candidate.get("nodeName"),
        candidate.get("partName"),
    ):
        key = _normalize_relation_key(value)
        if key and key in candidate_part_lookup:
            return candidate_part_lookup[key]
    return None


def _expand_chat_candidates_by_graph(
    *,
    seed_part_matches: List[Dict[str, Any]],
    seed_process_matches: List[Dict[str, Any]],
    candidate_parts: List[Any],
    process_templates: List[Dict[str, Any]],
    limit: int,
) -> Dict[str, List[Dict[str, Any]]]:
    def append_seed(target: List[str], value: Any) -> None:
        normalized = " ".join(str(value or "").split()).strip()
        if normalized and normalized not in target:
            target.append(normalized)

    seed_part_ids: List[str] = []
    for item in seed_part_matches or []:
        for key in ("partBase", "partId", "nodeName", "partName"):
            append_seed(seed_part_ids, item.get(key))

    seed_process_keys: List[str] = []
    for item in seed_process_matches or []:
        for key in ("processKey", "label", "partBase"):
            append_seed(seed_process_keys, item.get(key))

    allowed_process_labels = []
    for template in process_templates or []:
        for value in (template.get("label"), template.get("processKey"), template.get("partBase")):
            normalized = " ".join(str(value or "").split()).strip()
            if normalized:
                allowed_process_labels.append(normalized)

    candidate_part_lookup = _build_candidate_part_lookup(candidate_parts)

    expanded = retrieve_expanded_nodes_from_neo4j(
        seed_part_ids=seed_part_ids[:10],
        seed_process_keys=seed_process_keys[:10],
        allowed_process_labels=allowed_process_labels,
        limit=limit,
        max_depth=3,
    )

    process_matches: List[Dict[str, Any]] = []
    for candidate in expanded.get("processCandidates") or []:
        template = _match_process_template_from_graph_candidate(candidate, process_templates)
        if not template:
            continue
        depth = int(candidate.get("depth") or 1)
        matched_paths = float(candidate.get("matchedWindows") or candidate.get("matchedPaths") or 1)
        process_matches.append(
            _build_process_match_from_template(
                template,
                score=max(0.1, matched_paths / max(depth, 1)),
                reason=f"graph traversal depth={depth}",
            )
        )

    part_matches: List[Dict[str, Any]] = []
    for candidate in expanded.get("partCandidates") or []:
        part = _match_candidate_part_from_graph_candidate(candidate, candidate_part_lookup)
        if not part:
            continue
        depth = int(candidate.get("depth") or 1)
        matched_paths = float(candidate.get("matchedPaths") or candidate.get("matchedWindows") or 1)
        part_matches.append(
            _part_match_from_candidate_part(
                part,
                score=max(0.1, matched_paths / max(depth, 1)),
                reason=f"graph traversal depth={depth}",
            )
        )

    if process_matches or part_matches:
        return {
            "parts": _dedupe_part_matches(part_matches)[:limit],
            "processes": _dedupe_process_family_matches(process_matches)[:limit],
        }

    try:
        index = get_or_build_index()
    except Exception:
        return {
            "parts": [],
            "processes": [],
        }

    seed_keys = {_normalize_relation_key(value) for value in [*seed_part_ids, *seed_process_keys]}
    seed_keys = {value for value in seed_keys if value}
    if not seed_keys:
        return {
            "parts": [],
            "processes": [],
        }

    part_score_by_key: Dict[str, float] = {}
    process_score_by_label: Dict[str, float] = {}
    for document in index.documents:
        snippet = list(document.snippet or [])
        for index_in_snippet, step in enumerate(snippet):
            step_key = _normalize_relation_key(step.get("key")) or _normalize_relation_key(step.get("label"))
            if step_key not in seed_keys:
                continue
            for distance, next_step in enumerate(snippet[index_in_snippet + 1 : index_in_snippet + 4], start=1):
                next_type = str(next_step.get("type") or "").upper()
                if next_type == "PART":
                    part_key = _normalize_relation_key(next_step.get("key")) or _normalize_relation_key(next_step.get("label"))
                    if part_key:
                        part_score_by_key[part_key] = part_score_by_key.get(part_key, 0.0) + (1.0 / distance)
                    continue
                if next_type != "PROCESS":
                    continue
                process_label = " ".join(
                    str(next_step.get("label") or next_step.get("key") or "").split()
                ).strip()
                if not process_label:
                    continue
                process_score_by_label[process_label] = process_score_by_label.get(process_label, 0.0) + (1.0 / distance)

    for part_key, score in sorted(part_score_by_key.items(), key=lambda item: (-item[1], item[0]))[:limit]:
        part = candidate_part_lookup.get(part_key)
        if not part:
            continue
        part_matches.append(
            _part_match_from_candidate_part(
                part,
                score=score,
                reason="local graph traversal",
            )
        )

    ranked_process_labels = [
        label
        for label, _ in sorted(process_score_by_label.items(), key=lambda item: (-item[1], item[0]))
    ][:limit]
    for label in ranked_process_labels:
        candidate = {"processLabel": label}
        template = _match_process_template_from_graph_candidate(candidate, process_templates)
        if not template:
            continue
        process_matches.append(
            _build_process_match_from_template(
                template,
                score=process_score_by_label.get(label, 0.0),
                reason="local graph traversal",
            )
        )

    return {
        "parts": _dedupe_part_matches(part_matches)[:limit],
        "processes": _dedupe_process_family_matches(process_matches)[:limit],
    }

@router.get("/process-templates")
def get_process_templates():
    """
    Sequence 구성용 PROCESS 템플릿
    - 작업시간 분석표 DB 전 시트 대상
    - '부품 기준' = 관계 대상 부품
    - '요소작업' = 실제 공정 라벨
    """

    if not EXCEL_DB_PATH.exists():
        raise HTTPException(
            status_code=404,
            detail="작업시간 분석표 DB 엑셀 파일 없음"
        )

    return _get_cached_process_templates(
        str(EXCEL_DB_PATH.resolve()),
        EXCEL_DB_PATH.stat().st_mtime,
    )


def _recommend_next_processes_from_graph(
    *,
    selected_parts: List[Any],
    effective_process_templates: List[Dict[str, Any]],
    limit: int,
    message_tokens: Optional[List[str]] = None,
    apply_message_adjustment: bool = False,
) -> List[Dict[str, Any]]:
    selected_parts = _filter_process_like_selected_parts(selected_parts or [], effective_process_templates)
    selected_part_base = str(
        getattr((selected_parts or [None])[0], "partBase", None)
        or getattr((selected_parts or [None])[0], "partId", None)
        or getattr((selected_parts or [None])[0], "nodeName", None)
        or ""
    ).strip()
    _emit_sequence_recommend_log(
        "START",
        selectedParts=len(selected_parts or []),
        templates=len(effective_process_templates or []),
        limit=limit,
        applyMessageAdjustment=apply_message_adjustment,
    )
    normalized_selected_part_ids = [
        " ".join(
            str(
                getattr(part, "partBase", None)
                or getattr(part, "partId", None)
                or getattr(part, "nodeName", None)
                or ""
            ).split()
        ).strip()
        for part in (selected_parts or [])
        if str(
            getattr(part, "partBase", None)
            or getattr(part, "partId", None)
            or getattr(part, "nodeName", None)
            or ""
        ).strip()
    ]
    if not normalized_selected_part_ids:
        _emit_sequence_recommend_log("NO_SELECTED_PART_IDS")
        return []

    _emit_sequence_recommend_log(
        "SELECTED_PART_IDS",
        partIds=normalized_selected_part_ids,
    )

    allowed_process_labels = [
        " ".join(str(item.get("label") or "").split()).strip()
        for item in effective_process_templates
        if str(item.get("label") or "").strip()
    ]

    neo4j_result = retrieve_references_from_neo4j(
        selected_part_ids=normalized_selected_part_ids,
        allowed_process_labels=allowed_process_labels,
        limit=max(limit, 5),
    )
    graph_candidates = list(neo4j_result.get("processCandidates") or [])
    graph_windows = list(neo4j_result.get("referenceWindows") or [])
    _emit_sequence_recommend_log(
        "GRAPH_RESULT",
        referenceWindows=len(graph_windows),
        processCandidates=len(graph_candidates),
    )

    if not graph_windows and not graph_candidates:
        try:
            index = get_or_build_index()
            json_next_counts: Dict[str, int] = {}
            for part_id in normalized_selected_part_ids:
                for item in summarize_next_process_candidates(index, part_id, limit=max(limit * 3, 10)):
                    process_label = " ".join(str(item.get("processLabel") or "").split()).strip()
                    count = int(item.get("count") or item.get("matchedWindows") or 0)
                    if not process_label or count <= 0:
                        continue
                    json_next_counts[process_label] = json_next_counts.get(process_label, 0) + count

            if json_next_counts:
                graph_candidates = [
                    {
                        "processLabel": process_label,
                        "matchedWindows": matched_windows,
                        "retrievalBackend": "graph-index-next-process-fallback",
                    }
                    for process_label, matched_windows in sorted(
                        json_next_counts.items(),
                        key=lambda item: (-item[1], item[0]),
                    )
                ]
                _emit_sequence_recommend_log(
                    "GRAPH_INDEX_NEXT_PROCESS_FALLBACK",
                    candidates=graph_candidates[:10],
                )
        except Exception as exc:
            _emit_sequence_recommend_log(
                "GRAPH_INDEX_FALLBACK_FAILED",
                error=str(exc),
            )

    templates_by_label: Dict[str, List[Dict[str, Any]]] = {}
    for process in effective_process_templates:
        normalized_label = " ".join(str(process.get("label") or "").split()).strip()
        if not normalized_label:
            continue
        templates_by_label.setdefault(normalized_label, []).append(process)
    _emit_sequence_recommend_log(
        "TEMPLATES_BY_LABEL",
        uniqueLabels=len(templates_by_label),
    )

    next_process_counts: Dict[str, int] = {}
    selected_part_key_set = {_normalize_relation_key(value) for value in normalized_selected_part_ids}
    for window in graph_windows:
        snippet = list(window.get("snippet") or [])
        for index, item in enumerate(snippet):
            item_type = str(item.get("type") or "").strip().upper()
            if item_type != "PART":
                continue
            item_keys = {
                _normalize_relation_key(item.get("key")),
                _normalize_relation_key(item.get("label")),
            }
            if not item_keys.intersection(selected_part_key_set):
                continue
            for next_item in snippet[index + 1 :]:
                next_type = str(next_item.get("type") or "").strip().upper()
                if next_type == "PROCESS":
                    next_label = " ".join(
                        str(next_item.get("label") or next_item.get("key") or "").split()
                    ).strip()
                    if next_label:
                        next_process_counts[next_label] = next_process_counts.get(next_label, 0) + 1
                    break
                if next_type == "PART":
                    break

    _emit_sequence_recommend_log(
        "NEXT_PROCESS_COUNTS",
        counts=list(sorted(next_process_counts.items(), key=lambda item: (-item[1], item[0])))[:10],
    )

    if next_process_counts:
        graph_candidates = [
            {
                "processLabel": process_label,
                "matchedWindows": matched_windows,
                "retrievalBackend": "neo4j-next-process-fallback",
            }
            for process_label, matched_windows in sorted(
                next_process_counts.items(),
                key=lambda item: (-item[1], item[0]),
            )
        ]
        _emit_sequence_recommend_log(
            "USING_NEXT_PROCESS_FALLBACK",
            candidates=graph_candidates[:10],
        )
    elif not graph_candidates and graph_windows:
        window_process_counts: Dict[str, int] = {}
        for window in graph_windows:
            for raw_label in window.get("processLabels") or []:
                normalized_label = " ".join(str(raw_label or "").split()).strip()
                if not normalized_label:
                    continue
                window_process_counts[normalized_label] = window_process_counts.get(normalized_label, 0) + 1

        graph_candidates = [
            {
                "processLabel": process_label,
                "matchedWindows": matched_windows,
                "retrievalBackend": "neo4j-window-fallback",
            }
            for process_label, matched_windows in sorted(
                window_process_counts.items(),
                key=lambda item: (-item[1], item[0]),
            )
        ]
        _emit_sequence_recommend_log(
            "USING_WINDOW_PROCESS_FALLBACK",
            candidates=graph_candidates[:10],
        )
    else:
        _emit_sequence_recommend_log(
            "USING_GRAPH_CANDIDATES",
            candidates=graph_candidates[:10],
        )

    recommended_processes: List[Dict[str, Any]] = []
    seen_process_keys = set()

    for candidate in graph_candidates:
        candidate_label = " ".join(str(candidate.get("processLabel") or "").split()).strip()
        if not candidate_label:
            continue

        matched_templates = _match_process_templates_by_label(candidate_label, templates_by_label)
        fuzzy_matched_templates: List[Dict[str, Any]] = []
        if not matched_templates:
            fuzzy_matched_templates = _match_process_templates_by_db_fuzzy_label(
                candidate_label,
                templates_by_label,
            )
            matched_templates = fuzzy_matched_templates
        matched_templates = _filter_process_templates_for_selected_parts(
            matched_templates,
            selected_parts,
        )
        _emit_sequence_recommend_log(
            "MATCH_TEMPLATE",
            candidateLabel=candidate_label,
            matchedTemplateLabels=[
                str(item.get("label") or item.get("processKey") or "").strip()
                for item in matched_templates[:10]
            ],
            matchedTemplatePartBases=[
                str(item.get("partBase") or "").strip()
                for item in matched_templates[:10]
            ],
            matchedTemplateCount=len(matched_templates),
            matchMode="fuzzy" if fuzzy_matched_templates else ("exact_or_loose" if matched_templates else "none"),
        )
        if not matched_templates:
            _emit_sequence_recommend_log(
                "SKIP_NON_DB_PROCESS_CANDIDATE",
                candidateLabel=candidate_label,
            )
            continue

        for process in matched_templates:
            process_key = str(process.get("processKey") or "").strip()
            if not process_key or process_key in seen_process_keys:
                if process_key:
                    _emit_sequence_recommend_log(
                        "SKIP_DUPLICATED_PROCESS",
                        processKey=process_key,
                    )
                continue
            score = float(candidate.get("matchedWindows") or candidate.get("count") or 0)
            reasons = []
            if apply_message_adjustment:
                score, reasons = _apply_process_material_adjustment(
                    score,
                    reasons,
                    message_tokens or [],
                    process.get("processKey"),
                    process.get("label"),
                )
            if score <= 0:
                _emit_sequence_recommend_log(
                    "SKIP_NON_POSITIVE_SCORE",
                    processKey=process_key,
                    label=process.get("label"),
                    score=score,
                    reasons=reasons,
                )
                continue

            display_label = _resolve_process_display_label(process)
            seen_process_keys.add(process_key)
            recommended_processes.append(
                {
                    "processKey": process_key,
                    "label": process.get("label"),
                    "displayLabel": display_label,
                    "operationLabel": process.get("label"),
                    "partBase": process.get("partBase"),
                    "contextPartBase": selected_part_base,
                    "sourceSheet": process.get("sourceSheet"),
                    "score": round(score, 3),
                    "reason": ", ".join(dict.fromkeys(reasons)),
                }
            )
            _emit_sequence_recommend_log(
                "ADD_RECOMMENDED_PROCESS",
                processKey=process_key,
                label=process.get("label"),
                score=round(score, 3),
                reasons=reasons,
            )
            if len(recommended_processes) >= limit:
                break
        if len(recommended_processes) >= limit:
            break

    _emit_sequence_recommend_log(
        "DONE",
        recommendedCount=len(recommended_processes),
        recommendedLabels=[
            str(item.get("label") or item.get("processKey") or "").strip()
            for item in recommended_processes
        ],
    )
    return recommended_processes


def _recommend_common_process_path_from_graph(
    *,
    selected_parts: List[Any],
    effective_process_templates: List[Dict[str, Any]],
    limit: int,
) -> List[Dict[str, Any]]:
    selected_parts = _filter_process_like_selected_parts(selected_parts or [], effective_process_templates)
    normalized_selected_part_ids = _normalize_selected_part_ids(selected_parts)
    if not normalized_selected_part_ids:
        _emit_sequence_recommend_log("COMMON_PATH_NO_SELECTED_PART_IDS")
        return []

    allowed_process_labels: List[str] = []
    for item in effective_process_templates or []:
        for value in (item.get("label"), item.get("processKey"), item.get("partBase")):
            normalized = " ".join(str(value or "").split()).strip()
            if normalized:
                allowed_process_labels.append(normalized)

    try:
        index = get_or_build_index()
    except Exception as exc:
        _emit_sequence_recommend_log("COMMON_PATH_INDEX_FAILED", error=str(exc))
        return []

    ranked_windows = recommend_windows(
        index,
        selected_part_ids=normalized_selected_part_ids,
        allowed_process_labels=allowed_process_labels,
        limit=max(limit * 4, 12),
    )
    _emit_sequence_recommend_log(
        "COMMON_PATH_WINDOWS",
        selectedParts=normalized_selected_part_ids,
        windowCount=len(ranked_windows),
        topWindowIds=[str(item.get("docId") or "").strip() for item in ranked_windows[:5]],
    )
    if not ranked_windows:
        return []

    best_anchor_hits = max(
        int((window.get("score") or {}).get("anchorHits") or 0)
        for window in ranked_windows
    )
    fully_matched_windows = [
        window
        for window in ranked_windows
        if int((window.get("score") or {}).get("anchorHits") or 0) >= len(normalized_selected_part_ids)
    ]
    candidate_windows = fully_matched_windows or [
        window
        for window in ranked_windows
        if int((window.get("score") or {}).get("anchorHits") or 0) == best_anchor_hits
    ]
    _emit_sequence_recommend_log(
        "COMMON_PATH_FILTERED_WINDOWS",
        bestAnchorHits=best_anchor_hits,
        filteredWindowCount=len(candidate_windows),
        filteredWindowIds=[str(item.get("docId") or "").strip() for item in candidate_windows[:5]],
    )
    if not candidate_windows:
        return []

    templates_by_label: Dict[str, List[Dict[str, Any]]] = {}
    for process in effective_process_templates or []:
        for candidate in (
            " ".join(str(process.get("label") or "").split()).strip(),
            " ".join(str(process.get("processKey") or "").split()).strip(),
            " ".join(str(process.get("partBase") or "").split()).strip(),
        ):
            if not candidate:
                continue
            templates_by_label.setdefault(candidate, []).append(process)

    selected_part_context = ", ".join(
        [
            str(
                getattr(part, "partBase", None)
                or getattr(part, "partId", None)
                or getattr(part, "nodeName", None)
                or ""
            ).strip()
            for part in selected_parts
        ][:3]
    )
    aggregate_by_process_key: Dict[str, Dict[str, Any]] = {}

    for window_rank, window in enumerate(candidate_windows):
        score = window.get("score") or {}
        weight = (
            int(score.get("exactOrderHit") or 0) * 100
            + int(score.get("anchorHits") or 0) * 10
            + int(score.get("uniqueAllowedProcessHits") or 0) * 4
            + int(score.get("repeatedAllowedProcessHits") or 0)
        )
        if weight <= 0:
            weight = max(1, len(ranked_windows) - window_rank)

        snippet = list(window.get("snippet") or [])
        process_position = 0
        seen_process_keys_in_window = set()

        for snippet_item in snippet:
            snippet_type = str(snippet_item.get("type") or "").strip().upper()
            if snippet_type != "PROCESS":
                continue

            process_position += 1
            matched_templates: List[Dict[str, Any]] = []
            for candidate_value in (snippet_item.get("label"), snippet_item.get("key")):
                candidate_label = " ".join(str(candidate_value or "").split()).strip()
                if not candidate_label:
                    continue
                matched_templates = _match_process_templates_by_label(candidate_label, templates_by_label)
                if not matched_templates:
                    matched_templates = _match_process_templates_by_db_fuzzy_label(
                        candidate_label,
                        templates_by_label,
                    )
                if matched_templates:
                    break

            matched_templates = _filter_process_templates_for_selected_parts(
                matched_templates,
                selected_parts,
            )
            if not matched_templates:
                continue

            template = matched_templates[0]
            process_key = str(template.get("processKey") or "").strip()
            if not process_key or process_key in seen_process_keys_in_window:
                continue
            if not _is_process_semantically_allowed_for_selected_parts(template, selected_parts):
                continue
            seen_process_keys_in_window.add(process_key)

            bucket = aggregate_by_process_key.setdefault(
                process_key,
                {
                    "template": template,
                    "weight": 0,
                    "positionWeight": 0,
                    "positionWeightTotal": 0,
                    "matchedWindows": 0,
                    "firstWindowRank": window_rank,
                    "firstProcessPosition": process_position,
                },
            )
            bucket["weight"] += weight
            bucket["positionWeight"] += process_position * weight
            bucket["positionWeightTotal"] += weight
            bucket["matchedWindows"] += 1
            bucket["firstWindowRank"] = min(int(bucket["firstWindowRank"]), int(window_rank))
            bucket["firstProcessPosition"] = min(
                int(bucket["firstProcessPosition"]),
                int(process_position),
            )

    ranked_common_processes = sorted(
        aggregate_by_process_key.values(),
        key=lambda item: (
            int(item["firstWindowRank"]),
            (item["positionWeight"] / item["positionWeightTotal"]) if item["positionWeightTotal"] else 9999,
            int(item["firstProcessPosition"]),
            -item["weight"],
            str(item["template"].get("label") or item["template"].get("processKey") or ""),
        ),
    )

    recommended_processes: List[Dict[str, Any]] = []
    seen_families = set()
    for item in ranked_common_processes:
        template = item["template"]
        family_key = _process_family_key(template.get("processKey"), template.get("label"))
        if family_key in seen_families:
            continue
        if not _is_process_semantically_allowed_for_selected_parts(template, selected_parts):
            continue
        seen_families.add(family_key)

        recommended_processes.append(
            {
                "processKey": str(template.get("processKey") or "").strip(),
                "label": template.get("label"),
                "displayLabel": _resolve_process_display_label(template),
                "operationLabel": template.get("label"),
                "partBase": template.get("partBase"),
                "contextPartBase": selected_part_context,
                "sourceSheet": template.get("sourceSheet"),
                "score": round(float(item["weight"]), 3),
                "reason": "선택 부품 조합의 graph DB 공통 경로",
            }
        )
        if len(recommended_processes) >= limit:
            break

    _emit_sequence_recommend_log(
        "COMMON_PATH_RESULT",
        recommendedCount=len(recommended_processes),
        recommendedLabels=[
            str(item.get("label") or item.get("processKey") or "").strip()
            for item in recommended_processes
        ],
    )
    return recommended_processes


def _build_process_templates_by_label(
    effective_process_templates: List[Dict[str, Any]],
) -> Dict[str, List[Dict[str, Any]]]:
    templates_by_label: Dict[str, List[Dict[str, Any]]] = {}
    for process in effective_process_templates or []:
        for candidate in (
            " ".join(str(process.get("label") or "").split()).strip(),
            " ".join(str(process.get("processKey") or "").split()).strip(),
            " ".join(str(process.get("partBase") or "").split()).strip(),
        ):
            if candidate:
                templates_by_label.setdefault(candidate, []).append(process)
    return templates_by_label


def _match_graph_process_template(
    process_label: Any,
    templates_by_label: Dict[str, List[Dict[str, Any]]],
) -> Optional[Dict[str, Any]]:
    label = " ".join(str(process_label or "").split()).strip()
    if not label:
        return None

    matched_templates = _match_process_templates_by_label(label, templates_by_label)
    if not matched_templates:
        matched_templates = _match_process_templates_by_db_fuzzy_label(
            label,
            templates_by_label,
        )
    return matched_templates[0] if matched_templates else None


def _selected_part_graph_aliases(
    record: Dict[str, Any],
    selected_part_records: List[Dict[str, Any]],
) -> List[str]:
    aliases = [
        str(record.get("partBase") or "").strip(),
        str(record.get("partId") or "").strip(),
        str(record.get("partName") or "").strip(),
        str(record.get("nodeName") or "").strip(),
        str(record.get("label") or "").strip(),
    ]

    selected_blob = " ".join(
        str(item.get("partBase") or item.get("partId") or item.get("label") or "").upper()
        for item in selected_part_records
    )
    record_blob = " ".join(value.upper() for value in aliases if value)
    if "HEAT SINK" in selected_blob and "LAM" in selected_blob:
        if "HEAT SINK" in record_blob or "LAM" in record_blob:
            aliases.extend([
                "HEAT SINK+LAM 모듈",
                "HEAT SINK + LAM",
                "H/S+LAM 모듈",
            ])

    deduped: List[str] = []
    seen = set()
    for alias in aliases:
        normalized = " ".join(str(alias or "").split()).strip()
        if not normalized:
            continue
        relation_key = _normalize_relation_key(normalized)
        if relation_key in seen:
            continue
        seen.add(relation_key)
        deduped.append(normalized)
    return deduped


def _score_process_for_part_pair(process_label: str, current_part: Dict[str, Any], next_part: Dict[str, Any]) -> int:
    process_key = _normalize_relation_key(process_label)
    current_key = _normalize_relation_key(
        current_part.get("partBase") or current_part.get("partId") or current_part.get("label")
    )
    next_key = _normalize_relation_key(
        next_part.get("partBase") or next_part.get("partId") or next_part.get("label")
    )
    score = 0
    for token_source, weight in ((next_key, 12), (current_key, 5)):
        for token in re.split(r"[^A-Z0-9가-힣/]+", token_source):
            token = token.strip()
            if len(token) < 2:
                continue
            if token in process_key:
                score += weight
    return score


def _is_terminal_equipment_process(process_label: Any, process_key: Any = "") -> bool:
    haystack = _normalize_relation_key(f"{process_label or ''} {process_key or ''}")
    terminal_keywords = (
        "설비 스위치 ON",
        "설비 작동",
        "설비 동작",
        "START",
        "SWITCH ON",
        "스위치 ON",
    )
    return any(_normalize_relation_key(keyword) in haystack for keyword in terminal_keywords)


def _is_post_assembly_barcode_process(item: Any, process_key: Any = "", label: Any = "") -> bool:
    if isinstance(item, dict):
        values = [
            item.get("processKey"),
            item.get("label"),
            item.get("displayLabel"),
            item.get("operationLabel"),
            item.get("partBase"),
            item.get("contextPartBase"),
            item.get("reason"),
        ]
    else:
        values = [
            process_key,
            label,
            getattr(item, "processKey", None),
            getattr(item, "label", None),
            getattr(item, "displayLabel", None),
            getattr(item, "operationLabel", None),
            getattr(item, "partBase", None),
            getattr(item, "contextPartBase", None),
            getattr(item, "reason", None),
        ]

    haystack = _normalize_relation_key(" ".join(str(value or "") for value in values))
    has_barcode = any(keyword in haystack for keyword in ("바코드", "BAR CODE", "BAR-CODE", "BARCODE"))
    return has_barcode and not _is_manual_barcode_reading_process(item, process_key, label)


def _sort_post_assembly_processes_last(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return sorted(
        items or [],
        key=lambda item: (
            1 if _is_post_assembly_barcode_process(item) else 0,
            -float(item.get("score") or 0),
            str(item.get("label") or item.get("processKey") or ""),
        ),
    )


def _is_unload_process(process_label: Any, process_key: Any = "") -> bool:
    haystack = _normalize_relation_key(f"{process_label or ''} {process_key or ''}")
    unload_keywords = (
        "취출",
        "UNLOAD",
    )
    return any(_normalize_relation_key(keyword) in haystack for keyword in unload_keywords)


def _is_assy_part(part: Any) -> bool:
    if not part:
        return False
    if isinstance(part, dict):
        values = [
            part.get("partBase"),
            part.get("partId"),
            part.get("partName"),
            part.get("nodeName"),
            part.get("label"),
            part.get("displayLabel"),
        ]
    else:
        values = [
            getattr(part, "partBase", None),
            getattr(part, "partId", None),
            getattr(part, "partName", None),
            getattr(part, "nodeName", None),
            getattr(part, "label", None),
            getattr(part, "displayLabel", None),
        ]
    haystack = " ".join(str(value or "").upper() for value in values)
    compact = re.sub(r"[^A-Z0-9가-힣]+", "", haystack)
    return "ASSY" in compact or "ASS'Y" in haystack or "SUBASSY" in compact


def _is_simple_assy_seating_process(process_label: Any, process_key: Any = "") -> bool:
    haystack = _normalize_relation_key(f"{process_label or ''} {process_key or ''}")
    simple_assy_keywords = (
        "부품(단품)류 단순 안착작업",
        "부품 단품류 단순 안착작업",
        "부품/단품 단순 안착",
    )
    return any(_normalize_relation_key(keyword) in haystack for keyword in simple_assy_keywords)


def _semantic_part_text(part: Any) -> str:
    if not part:
        return ""
    if isinstance(part, dict):
        values = [
            part.get("partBase"),
            part.get("partId"),
            part.get("partName"),
            part.get("nodeName"),
            part.get("label"),
            part.get("displayLabel"),
        ]
    else:
        values = [
            getattr(part, "partBase", None),
            getattr(part, "partId", None),
            getattr(part, "partName", None),
            getattr(part, "nodeName", None),
            getattr(part, "label", None),
            getattr(part, "displayLabel", None),
        ]
    return _normalize_relation_key(" ".join(str(value or "") for value in values))


def _semantic_process_text(template: Dict[str, Any]) -> str:
    return _normalize_relation_key(
        " ".join(
            str(value or "")
            for value in (
                template.get("label"),
                template.get("processKey"),
                template.get("partBase"),
                template.get("operationLabel"),
            )
        )
    )


def _semantic_pair_text(current_part: Optional[Dict[str, Any]], next_part: Optional[Dict[str, Any]]) -> str:
    return f"{_semantic_part_text(current_part)} {_semantic_part_text(next_part)}".strip()


def _semantic_text_has_any(haystack: str, keywords: Tuple[str, ...]) -> bool:
    return any(_normalize_relation_key(keyword) in haystack for keyword in keywords)


def _semantic_process_part_hints(process_text: str) -> List[Tuple[str, ...]]:
    hint_groups: List[Tuple[str, ...]] = []
    rules: List[Tuple[Tuple[str, ...], Tuple[str, ...]]] = [
        (("LDM", "LED DRIVE MODULE"), ("LDM", "LED DRIVE MODULE")),
        (("HEAT SINK", "히트싱크", "H/S"), ("HEAT SINK", "히트싱크", "H/S")),
        (("LAM",), ("LAM",)),
        (("DUST CAP", "더스트 캡"), ("DUST CAP", "DUST", "CAP", "더스트", "캡")),
        (("BEZEL", "베젤"), ("BEZEL", "베젤")),
        (("LENS", "렌즈"), ("LENS", "렌즈")),
        (("HOUSING", "하우징"), ("HOUSING", "하우징")),
        (("WIRE", "와이어"), ("WIRE", "와이어")),
        (("CONNECTOR", "커넥터"), ("CONNECTOR", "커넥터")),
        (("BRACKET", "BRKT", "브라켓"), ("BRACKET", "BRKT", "브라켓")),
        (("CAP", "캡"), ("CAP", "캡")),
        (("PCB", "기판"), ("PCB", "기판")),
    ]
    for process_keywords, part_keywords in rules:
        if _semantic_text_has_any(process_text, process_keywords):
            hint_groups.append(part_keywords)
    return hint_groups


def _is_generic_process_for_auto_sequence(process_text: str) -> bool:
    generic_keywords = (
        "BAR-CODE",
        "BAR CODE",
        "바코드",
        "에어 블로윙",
        "AIR BLOWING",
        "유성펜",
        "마킹",
        "커넥터류 연결",
    )
    return _semantic_text_has_any(process_text, generic_keywords)


def _is_process_semantically_allowed_between_parts(
    template: Dict[str, Any],
    current_part: Optional[Dict[str, Any]],
    next_part: Optional[Dict[str, Any]],
) -> bool:
    process_label = template.get("label")
    process_key = template.get("processKey")
    process_text = _semantic_process_text(template)
    pair_text = _semantic_pair_text(current_part, next_part)

    if _is_unload_process(process_label, process_key):
        return False
    if _is_terminal_equipment_process(process_label, process_key):
        return True
    if _is_simple_assy_seating_process(process_label, process_key):
        return _is_assy_part(current_part) or _is_assy_part(next_part)

    hint_groups = _semantic_process_part_hints(process_text)
    if hint_groups:
        return all(_semantic_text_has_any(pair_text, hints) for hints in hint_groups)

    if _is_generic_process_for_auto_sequence(process_text):
        return True

    return True


def _is_process_semantically_allowed_for_selected_parts(
    template: Dict[str, Any],
    selected_parts: List[Any],
) -> bool:
    process_label = template.get("label")
    process_key = template.get("processKey")
    process_text = _semantic_process_text(template)
    selected_text = " ".join(_semantic_part_text(part) for part in selected_parts or [])

    if _is_unload_process(process_label, process_key):
        return False
    if _is_terminal_equipment_process(process_label, process_key):
        return True
    if _is_simple_assy_seating_process(process_label, process_key):
        return any(_is_assy_part(part) for part in selected_parts or [])

    hint_groups = _semantic_process_part_hints(process_text)
    if hint_groups:
        return all(_semantic_text_has_any(selected_text, hints) for hints in hint_groups)

    if _is_generic_process_for_auto_sequence(process_text):
        return True

    return True


def _is_manual_barcode_reading_process(item: Any, process_key: Any = "", label: Any = "") -> bool:
    if isinstance(item, dict):
        values = [
            item.get("processKey"),
            item.get("label"),
            item.get("displayLabel"),
            item.get("operationLabel"),
            item.get("partBase"),
            item.get("contextPartBase"),
            item.get("reason"),
        ]
    else:
        values = [
            process_key,
            label,
            getattr(item, "processKey", None),
            getattr(item, "label", None),
            getattr(item, "displayLabel", None),
            getattr(item, "operationLabel", None),
            getattr(item, "partBase", None),
            getattr(item, "contextPartBase", None),
            getattr(item, "reason", None),
        ]

    haystack = _normalize_relation_key(" ".join(str(value or "") for value in values))
    compact = re.sub(r"[^A-Z0-9가-힣]+", "", haystack)
    has_barcode = any(keyword in haystack for keyword in ("바코드", "BAR CODE", "BAR-CODE", "BARCODE"))
    has_reading = any(keyword in haystack for keyword in ("리딩", "READ", "READING", "SCAN", "스캔"))
    return (has_barcode and has_reading) or "단순 리딩 작업" in haystack or "단순리딩작업" in compact


def _filter_manual_barcode_reading_processes(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return list(items or [])


def _filter_manual_barcode_reading_sequence_steps(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return list(items or [])


def _is_fastening_after_parts_process(process_label: Any, process_key: Any = "") -> bool:
    haystack = _normalize_relation_key(f"{process_label or ''} {process_key or ''}")
    fastening_keywords = (
        "T/SCREW",
        "SCREW",
        "스크류",
        "볼트",
        "체결",
    )
    return any(_normalize_relation_key(keyword) in haystack for keyword in fastening_keywords)


def _canonical_selected_part_base_for_process(
    process_part_base: Any,
    selected_part_records: List[Dict[str, Any]],
) -> str:
    process_family = _part_family_key(process_part_base)
    if not process_family:
        return str(process_part_base or "").strip()

    for part in selected_part_records or []:
        selected_value = (
            part.get("partBase")
            or part.get("partId")
            or part.get("partName")
            or part.get("nodeName")
            or part.get("label")
            or ""
        )
        if _part_family_key(selected_value) == process_family:
            return str(selected_value or "").strip()

    return str(process_part_base or "").strip()


def _build_transition_flow_sequence(
    *,
    selected_part_records: List[Dict[str, Any]],
    effective_process_templates: List[Dict[str, Any]],
    index: Any,
    limit: int,
) -> List[Dict[str, Any]]:
    if len(selected_part_records) < 2:
        return selected_part_records

    templates_by_label = _build_process_templates_by_label(effective_process_templates)
    part_to_process_counts = getattr(index, "part_to_process_counts", {}) or {}
    process_to_part_counts = getattr(index, "process_to_part_counts", {}) or {}

    output_steps: List[Dict[str, Any]] = []
    seen_process_families = set()
    deferred_terminal_processes: List[Dict[str, Any]] = []
    deferred_terminal_families = set()
    context_part_base = ", ".join(
        [
            str(item.get("partBase") or item.get("label") or "").strip()
            for item in selected_part_records
        ]
    )

    def add_process_candidates(
        candidates: Dict[str, int],
        process_label: Any,
        weight: int,
    ) -> None:
        label = " ".join(str(process_label or "").split()).strip()
        if not label:
            return
        candidates[label] = max(int(candidates.get(label, 0)), int(weight))

    for index_in_sequence, current_part in enumerate(selected_part_records):
        output_steps.append(dict(current_part))
        if index_in_sequence >= len(selected_part_records) - 1:
            continue

        next_part = selected_part_records[index_in_sequence + 1]
        current_aliases = _selected_part_graph_aliases(current_part, selected_part_records)
        next_aliases = _selected_part_graph_aliases(next_part, selected_part_records)
        current_alias_keys = {_normalize_relation_key(alias) for alias in current_aliases}
        next_alias_keys = {_normalize_relation_key(alias) for alias in next_aliases}

        candidates: Dict[str, int] = {}
        for part_key, process_counts in part_to_process_counts.items():
            if _normalize_relation_key(part_key) not in current_alias_keys:
                continue
            for process_label, count in (process_counts or {}).items():
                add_process_candidates(
                    candidates,
                    process_label,
                    int(count or 0) * 10 + _score_process_for_part_pair(process_label, current_part, next_part),
                )

        for process_label, part_counts in process_to_part_counts.items():
            matched_count = 0
            for part_key, count in (part_counts or {}).items():
                if _normalize_relation_key(part_key) in next_alias_keys:
                    matched_count += int(count or 0)
            if matched_count:
                add_process_candidates(
                    candidates,
                    process_label,
                    matched_count * 10 + 6 + _score_process_for_part_pair(process_label, current_part, next_part),
                )

        ranked_candidates = sorted(candidates.items(), key=lambda item: (-item[1], item[0]))
        inserted_for_pair = 0
        for process_label, score in ranked_candidates:
            template = _match_graph_process_template(process_label, templates_by_label)
            if not template:
                continue
            process_part_family = _part_family_key(template.get("partBase"))
            if process_part_family and any(
                _part_family_key(
                    part.get("partBase")
                    or part.get("partId")
                    or part.get("partName")
                    or part.get("nodeName")
                    or part.get("label")
                )
                == process_part_family
                for part in selected_part_records
            ):
                continue
            process_key = str(template.get("processKey") or "").strip()
            family_key = _process_family_key(process_key, template.get("label"))
            if not process_key or family_key in seen_process_families:
                continue
            if _is_unload_process(template.get("label"), process_key):
                continue
            if not _is_process_semantically_allowed_for_selected_parts(template, selected_part_records):
                continue
            should_defer_process = _is_terminal_equipment_process(
                template.get("label"),
                process_key,
            ) or _is_post_assembly_barcode_process(template)
            if should_defer_process:
                if family_key not in deferred_terminal_families:
                    deferred_terminal_families.add(family_key)
                    deferred_terminal_processes.append(
                        {
                            "type": "PROCESS",
                            "processKey": process_key,
                            "label": template.get("label"),
                            "displayLabel": _resolve_process_display_label(template),
                            "operationLabel": template.get("label"),
                            "partBase": template.get("partBase"),
                            "contextPartBase": context_part_base,
                            "sourceSheet": template.get("sourceSheet"),
                            "reason": "선택 부품 조립 조건 충족 후 후행 공정",
                            "score": float(score),
                        }
                    )
                continue
            seen_process_families.add(family_key)
            canonical_part_base = _canonical_selected_part_base_for_process(
                template.get("partBase"),
                selected_part_records,
            )
            output_steps.append(
                {
                    "type": "PROCESS",
                    "processKey": process_key,
                    "label": template.get("label"),
                    "displayLabel": _resolve_process_display_label(template),
                    "operationLabel": template.get("label"),
                    "partBase": canonical_part_base or template.get("partBase"),
                    "contextPartBase": context_part_base,
                    "sourceSheet": template.get("sourceSheet"),
                    "reason": "선택 부품 순서 기반 graph transition 흐름",
                    "score": float(score),
                }
            )
            inserted_for_pair += 1
            if inserted_for_pair >= 1:
                break
            if sum(1 for step in output_steps if step.get("type") == "PROCESS") >= limit:
                break

    remaining_slots = max(0, limit - sum(1 for step in output_steps if step.get("type") == "PROCESS"))
    for terminal_process in deferred_terminal_processes[:remaining_slots]:
        family_key = _process_family_key(terminal_process.get("processKey"), terminal_process.get("label"))
        if family_key in seen_process_families:
            continue
        seen_process_families.add(family_key)
        output_steps.append(terminal_process)

    process_count = sum(1 for step in output_steps if step.get("type") == "PROCESS")
    output_steps = output_steps if process_count > 0 else []
    if output_steps:
        process_steps = _sort_post_assembly_processes_last(
            [step for step in output_steps if step.get("type") == "PROCESS"]
        )
        part_steps = [step for step in output_steps if step.get("type") != "PROCESS"]
        output_steps = part_steps + process_steps
    return output_steps


def _recommend_selected_part_combination_sequence(
    *,
    selected_parts: List[Any],
    effective_process_templates: List[Dict[str, Any]],
    limit: int,
) -> List[Dict[str, Any]]:
    selected_parts = _filter_process_like_selected_parts(selected_parts or [], effective_process_templates)
    normalized_selected_part_ids = _normalize_selected_part_ids(selected_parts)
    if len(normalized_selected_part_ids) < 2:
        return []

    selected_part_records: List[Dict[str, Any]] = []
    selected_part_by_key: Dict[str, Dict[str, Any]] = {}
    for part in selected_parts or []:
        part_values = [
            getattr(part, "partBase", None),
            getattr(part, "partId", None),
            getattr(part, "nodeName", None),
            getattr(part, "partName", None),
        ]
        display_value = next(
            (str(value or "").strip() for value in part_values if str(value or "").strip()),
            "",
        )
        if not display_value:
            continue

        record = {
            "type": "PART",
            "nodeName": str(getattr(part, "nodeName", None) or display_value).strip(),
            "partId": str(getattr(part, "partId", None) or display_value).strip(),
            "partName": str(getattr(part, "partName", None) or display_value).strip(),
            "partBase": str(getattr(part, "partBase", None) or display_value).strip(),
            "sourceSheet": str(getattr(part, "sourceSheet", None) or "").strip(),
            "label": display_value,
            "reason": "선택 부품",
        }

        record_keys = {
            key
            for value in part_values
            for key in (_normalize_relation_key(value), _part_family_key(value))
        }
        record_keys = {key for key in record_keys if key}
        if any(key in selected_part_by_key for key in record_keys):
            continue

        selected_part_records.append(record)
        for key in record_keys:
            selected_part_by_key[key] = record

    if len(selected_part_records) < 2:
        return selected_part_records

    allowed_process_labels: List[str] = []
    for item in effective_process_templates or []:
        for value in (item.get("label"), item.get("processKey"), item.get("partBase")):
            normalized = " ".join(str(value or "").split()).strip()
            if normalized:
                allowed_process_labels.append(normalized)

    try:
        index = get_or_build_index()
    except Exception as exc:
        _emit_sequence_recommend_log("COMBINATION_SEQUENCE_INDEX_FAILED", error=str(exc))
        return selected_part_records

    transition_flow_steps = _build_transition_flow_sequence(
        selected_part_records=selected_part_records,
        effective_process_templates=effective_process_templates,
        index=index,
        limit=limit,
    )
    if transition_flow_steps:
        _emit_sequence_recommend_log(
            "COMBINATION_SEQUENCE_TRANSITION_FLOW_RESULT",
            stepCount=len(transition_flow_steps),
            processCount=sum(1 for step in transition_flow_steps if step.get("type") == "PROCESS"),
        )
        return transition_flow_steps

    ranked_windows = recommend_windows(
        index,
        selected_part_ids=normalized_selected_part_ids,
        allowed_process_labels=allowed_process_labels,
        limit=max(limit * 4, 12),
    )
    if not ranked_windows:
        return selected_part_records

    fully_matched_windows = [
        window
        for window in ranked_windows
        if int((window.get("score") or {}).get("anchorHits") or 0) >= len(selected_part_records)
    ]
    candidate_windows = fully_matched_windows or ranked_windows

    templates_by_label = _build_process_templates_by_label(effective_process_templates)

    best_steps: List[Dict[str, Any]] = []
    best_score: Tuple[int, int, int] = (-1, -1, -1)

    for window_rank, window in enumerate(candidate_windows):
        output_steps: List[Dict[str, Any]] = []
        seen_part_keys = set()
        seen_process_families = set()
        deferred_terminal_processes: List[Dict[str, Any]] = []
        deferred_terminal_families = set()
        selected_parts_seen = 0
        pending_processes: List[Dict[str, Any]] = []

        for snippet_item in list(window.get("snippet") or []):
            snippet_type = str(snippet_item.get("type") or "").strip().upper()

            if snippet_type == "PART":
                item_keys = {
                    _normalize_relation_key(snippet_item.get("key")),
                    _normalize_relation_key(snippet_item.get("label")),
                }
                matched_part = next(
                    (
                        selected_part_by_key[key]
                        for key in item_keys
                        if key and key in selected_part_by_key
                    ),
                    None,
                )
                if not matched_part:
                    continue

                part_key = _normalize_relation_key(matched_part.get("partBase")) or _normalize_relation_key(
                    matched_part.get("nodeName")
                )
                if not part_key or part_key in seen_part_keys:
                    pending_processes = []
                    continue

                if selected_parts_seen > 0:
                    output_steps.extend(pending_processes)
                output_steps.append(dict(matched_part))
                seen_part_keys.add(part_key)
                selected_parts_seen += 1
                pending_processes = []
                continue

            if snippet_type != "PROCESS" or selected_parts_seen <= 0:
                continue

            matched_templates: List[Dict[str, Any]] = []
            for candidate_value in (snippet_item.get("label"), snippet_item.get("key")):
                candidate_label = " ".join(str(candidate_value or "").split()).strip()
                if not candidate_label:
                    continue
                matched_templates = _match_process_templates_by_label(candidate_label, templates_by_label)
                if not matched_templates:
                    matched_templates = _match_process_templates_by_db_fuzzy_label(
                        candidate_label,
                        templates_by_label,
                    )
                if matched_templates:
                    break

            if not matched_templates:
                continue

            template = matched_templates[0]
            process_key = str(template.get("processKey") or "").strip()
            process_part_key = _normalize_relation_key(template.get("partBase"))
            process_part_family_key = _part_family_key(template.get("partBase"))
            if process_part_key and process_part_key in selected_part_by_key:
                continue
            if process_part_family_key and process_part_family_key in selected_part_by_key:
                continue
            family_key = _process_family_key(process_key, template.get("label"))
            if not process_key or family_key in seen_process_families:
                continue
            if _is_unload_process(template.get("label"), process_key):
                continue
            if not _is_process_semantically_allowed_for_selected_parts(template, selected_part_records):
                continue
            canonical_part_base = _canonical_selected_part_base_for_process(
                template.get("partBase"),
                selected_part_records,
            )
            if _is_terminal_equipment_process(template.get("label"), process_key) or _is_post_assembly_barcode_process(template):
                if family_key not in deferred_terminal_families:
                    deferred_terminal_families.add(family_key)
                    deferred_terminal_processes.append(
                        {
                            "type": "PROCESS",
                            "processKey": process_key,
                            "label": template.get("label"),
                            "displayLabel": _resolve_process_display_label(template),
                            "operationLabel": template.get("label"),
                            "partBase": canonical_part_base or template.get("partBase"),
                            "contextPartBase": ", ".join(
                                [str(item.get("partBase") or item.get("label") or "").strip() for item in selected_part_records]
                            ),
                            "sourceSheet": template.get("sourceSheet"),
                            "reason": "선택 부품 조립 조건 충족 후 후행 공정",
                            "score": float((window.get("score") or {}).get("anchorHits") or 0),
                        }
                    )
                continue
            seen_process_families.add(family_key)

            pending_processes.append(
                {
                    "type": "PROCESS",
                    "processKey": process_key,
                    "label": template.get("label"),
                    "displayLabel": _resolve_process_display_label(template),
                    "operationLabel": template.get("label"),
                    "partBase": canonical_part_base or template.get("partBase"),
                    "contextPartBase": ", ".join(
                        [str(item.get("partBase") or item.get("label") or "").strip() for item in selected_part_records]
                    ),
                    "sourceSheet": template.get("sourceSheet"),
                    "reason": "선택 부품 사이 graph DB 경로",
                    "score": float((window.get("score") or {}).get("anchorHits") or 0),
                }
            )

        for terminal_process in deferred_terminal_processes:
            family_key = _process_family_key(terminal_process.get("processKey"), terminal_process.get("label"))
            if family_key in seen_process_families:
                continue
            seen_process_families.add(family_key)
            output_steps.append(terminal_process)

        part_count = sum(1 for step in output_steps if step.get("type") == "PART")
        process_count = sum(1 for step in output_steps if step.get("type") == "PROCESS")
        score = (part_count, process_count, -window_rank)
        if part_count >= 2 and score > best_score:
            best_score = score
            best_steps = output_steps

    if not best_steps:
        return selected_part_records

    missing_parts = []
    output_part_keys = {
        _normalize_relation_key(step.get("partBase")) or _normalize_relation_key(step.get("nodeName"))
        for step in best_steps
        if step.get("type") == "PART"
    }
    for part in selected_part_records:
        part_key = _normalize_relation_key(part.get("partBase")) or _normalize_relation_key(part.get("nodeName"))
        if part_key and part_key not in output_part_keys:
            missing_parts.append(part)

    process_steps = _sort_post_assembly_processes_last(
        [step for step in best_steps if step.get("type") == "PROCESS"]
    )
    part_steps = [step for step in best_steps if step.get("type") != "PROCESS"]
    return part_steps + process_steps + missing_parts


@router.get("/process/options")
def get_options(
    part_base: str = Query(..., alias="partBase"),
    process_label: Optional[str] = Query(None, alias="processLabel"),
    source_sheet: Optional[str] = Query(None, alias="sourceSheet"),
):
    requested_sheet = (source_sheet or "").strip()
    options = (
        _load_process_options_from_excel(part_base, process_label or "", requested_sheet)
        if str(process_label or "").strip()
        else _load_options_from_excel(part_base, requested_sheet)
    )

    return {
        "partBase": part_base,
        "processLabel": process_label,
        "sourceSheet": requested_sheet,
        "options": options,
    }

@router.get("/part/options")
def get_part_options(
    part_base: str = Query(..., alias="partBase"),
    source_sheet: Optional[str] = Query(None, alias="sourceSheet"),
):
    """
    부품 기준 OPTION 조회
    source: 작업시간 분석표 DB 엑셀
    """

    if not EXCEL_DB_PATH.exists():
        raise HTTPException(404, "작업시간 분석표 DB 엑셀 파일 없음")

    requested_sheet = (source_sheet or "").strip()
    option_index = _get_sequence_option_index()
    rows_by_sheet = option_index.get("rowsBySheet") or {}
    sheet_names = _resolve_option_sheet_names(option_index, requested_sheet)
    matched_sheets = [
        sheet_name
        for sheet_name in sheet_names
        if any(
            item.get("partBase") == part_base and item.get("option")
            for item in rows_by_sheet.get(sheet_name, [])
        )
    ]
    options = _load_options_from_excel(part_base, requested_sheet)

    return {
        "partBase": part_base,
        "sourceSheet": requested_sheet,
        "matchedSheets": matched_sheets,
        "options": options,
        "count": len(options),
    }


@router.get("/node/action-elements")
def get_node_action_elements(
    node_type: str = Query(..., alias="type"),
    part_base: str = Query(..., alias="partBase"),
    option_value: str = Query(..., alias="option"),
    process_label: Optional[str] = Query(None, alias="processLabel"),
    source_sheet: Optional[str] = Query(None, alias="sourceSheet"),
):
    if not EXCEL_DB_PATH.exists():
        raise HTTPException(404, "작업시간 분석표 DB 엑셀 파일 없음")

    normalized_type = str(node_type or "").strip().upper()
    requested_process_label = (process_label or "").strip() if normalized_type == "PROCESS" else ""
    requested_sheet = (source_sheet or "").strip()
    result = _load_action_elements_from_excel(
        part_base,
        option_value,
        process_label=requested_process_label,
        source_sheet=requested_sheet,
    )

    return {
        "type": normalized_type,
        "partBase": part_base,
        "processLabel": requested_process_label,
        "option": option_value,
        "resolvedOption": result.get("resolvedOption", ""),
        "matchStrategy": result.get("matchStrategy", "none"),
        "sourceSheet": requested_sheet,
        "matchedSheets": result["matchedSheets"],
        "rows": result["rows"],
        "count": len(result["rows"]),
    }


@router.post("/chat", response_model=SequenceChatResponse)
def chat_sequence_recommendations(req: SequenceChatRequest):
    message = str(req.message or "").strip()
    if not message:
        raise HTTPException(status_code=400, detail="message가 비어 있습니다.")

    _emit_sequence_recommend_log(
        "CHAT_REQUEST",
        bomId=req.bomId,
        spec=req.spec,
        message=message,
        selectedParts=[
            str(part.partBase or part.partId or part.nodeName or "").strip()
            for part in (req.selectedParts or [])
        ],
    )

    message_tokens = _tokenize_chat_text(message)
    requested_part_families = _requested_part_families_from_message(message_tokens)
    limit = max(1, min(int(req.limit or 3), 5))
    candidate_limit = max(limit * 3, 10)
    effective_process_templates = _resolve_effective_process_templates(req.processTemplates)
    ai_reply = ""
    recommended_parts: List[Dict[str, Any]] = []
    recommended_processes: List[Dict[str, Any]] = []
    embedding_part_matches: List[Dict[str, Any]] = []
    embedding_process_matches: List[Dict[str, Any]] = []
    graph_part_matches: List[Dict[str, Any]] = []
    graph_process_matches: List[Dict[str, Any]] = []
    per_part_recommendations: List[Dict[str, Any]] = []
    used_fast_path = False

    if not list(req.selectedParts or []) and not _message_is_sequence_chat_relevant(
        message_tokens,
        list(req.candidateParts or []),
        effective_process_templates,
    ):
        return SequenceChatResponse(
            reply="제조 공정이나 부품과 관련된 표현을 찾지 못했습니다. 부품명, 공정명, 동작(안착/체결/블로잉/작동 등)을 포함해서 다시 입력해 주세요.",
            recommendedParts=[],
            recommendedProcesses=[],
            recommendedOptions=[],
            perPartRecommendations=[],
        )

    direct_message_part_matches = _build_direct_message_part_matches(
        list(req.candidateParts or []),
        message_tokens,
        effective_process_templates=effective_process_templates,
        limit=limit,
    )
    direct_message_part_matches = _apply_chat_part_display_labels(
        direct_message_part_matches,
        message_tokens,
    )

    if direct_message_part_matches and not req.selectedParts:
        fast_path_processes = _build_fast_path_process_matches(
            direct_message_part_matches,
            effective_process_templates=effective_process_templates,
            message_tokens=message_tokens,
            requested_part_families=requested_part_families,
            limit=limit,
        )
        if fast_path_processes:
            recommended_parts = direct_message_part_matches
            recommended_processes = fast_path_processes
            used_fast_path = True

    if not used_fast_path:
        try:
            embedding_result = search_chat_candidates_with_bge_m3(
                message,
                candidate_parts=list(req.candidateParts or []),
                process_templates=effective_process_templates,
                part_limit=max(candidate_limit, 30),
                process_limit=max(candidate_limit * 2, 60),
            )
            embedding_part_matches = [
                _part_match_from_embedding_document(item)
                for item in (embedding_result.get("parts") or [])
            ]
            embedding_part_matches = _filter_part_matches_by_requested_families(
                embedding_part_matches,
                requested_part_families,
                message_tokens,
            )
            embedding_process_matches = [
                _process_match_from_embedding_document(item)
                for item in (embedding_result.get("processes") or [])
                if str((item.get("item") or {}).get("processKey") if isinstance(item.get("item"), dict) else getattr(item.get("item"), "processKey", "") or "").strip()
            ]
            embedding_process_matches = _filter_process_matches_by_requested_context(
                embedding_process_matches,
                requested_part_families,
                message_tokens,
            )
            _emit_sequence_recommend_log(
                "CHAT_EMBEDDING_RESULT",
                model="BAAI/bge-m3",
                partCount=len(embedding_part_matches),
                processCount=len(embedding_process_matches),
                topParts=[
                    str(item.get("partBase") or item.get("partId") or item.get("nodeName") or "").strip()
                    for item in embedding_part_matches[:5]
                ],
                topProcesses=[
                    str(item.get("processKey") or "").strip()
                    for item in embedding_process_matches[:5]
                ],
            )
        except Exception as exc:
            logger.warning("bge-m3 chat embedding search failed; falling back to heuristic candidates: %s", exc)
            _emit_sequence_recommend_log(
                "CHAT_EMBEDDING_FAILED",
                model="BAAI/bge-m3",
                error=str(exc),
            )

    if not used_fast_path:
        try:
            selected_seed_matches = [
                _part_match_from_candidate_part(
                    part,
                    reason="선택 부품 graph traversal seed",
                    score=1.0,
                )
                for part in (req.selectedParts or [])
            ]
            graph_expanded = _expand_chat_candidates_by_graph(
                seed_part_matches=[*selected_seed_matches, *embedding_part_matches],
                seed_process_matches=embedding_process_matches,
                candidate_parts=list(req.selectedParts or []) + list(req.candidateParts or []),
                process_templates=effective_process_templates,
                limit=max(candidate_limit, 20),
            )
            graph_part_matches = list(graph_expanded.get("parts") or [])
            graph_part_matches = _filter_part_matches_by_requested_families(
                graph_part_matches,
                requested_part_families,
                message_tokens,
            )
            graph_process_matches = list(graph_expanded.get("processes") or [])
            graph_process_matches = _filter_process_matches_by_requested_context(
                graph_process_matches,
                requested_part_families,
                message_tokens,
            )
            _emit_sequence_recommend_log(
                "CHAT_GRAPH_TRAVERSAL_RESULT",
                seedPartCount=len(selected_seed_matches) + len(embedding_part_matches),
                seedProcessCount=len(embedding_process_matches),
                partCount=len(graph_part_matches),
                processCount=len(graph_process_matches),
                topParts=[
                    str(item.get("partBase") or item.get("partId") or item.get("nodeName") or "").strip()
                    for item in graph_part_matches[:5]
                ],
                topProcesses=[
                    str(item.get("processKey") or item.get("label") or "").strip()
                    for item in graph_process_matches[:5]
                ],
            )
        except Exception as exc:
            logger.warning("Graph traversal expansion failed; falling back to embedding candidates: %s", exc)
            _emit_sequence_recommend_log(
                "CHAT_GRAPH_TRAVERSAL_FAILED",
                error=str(exc),
            )

    if not used_fast_path:
        try:
            graph_or_embedding_part_matches = graph_part_matches or embedding_part_matches
            message_process_matches = _recommend_processes_from_message_fallback(
                message_tokens=message_tokens,
                effective_process_templates=effective_process_templates,
                limit=max(candidate_limit, 20),
            )
            message_process_matches = _filter_process_matches_by_requested_context(
                message_process_matches,
                requested_part_families,
                message_tokens,
            )
            template_process_candidates = _dedupe_process_family_matches(
                [
                    *graph_process_matches,
                    *embedding_process_matches,
                    *message_process_matches,
                ]
            ) or [
                {
                    "processKey": str(process.get("processKey") or "").strip(),
                    "label": process.get("label"),
                    "displayLabel": _resolve_process_display_label(process),
                    "operationLabel": process.get("label"),
                    "partBase": process.get("partBase"),
                    "contextPartBase": process.get("partBase"),
                    "sourceSheet": process.get("sourceSheet"),
                    "score": 1.0,
                    "reason": "AI 채팅 후보",
                }
                for process in effective_process_templates[:120]
                if str(process.get("processKey") or "").strip()
            ]
            template_process_candidates = _filter_process_matches_by_requested_context(
                template_process_candidates,
                requested_part_families,
                message_tokens,
            )
            ai_candidate_parts = list(req.selectedParts or []) + [
                SimpleNamespace(
                    nodeName=item.get("nodeName"),
                    partBase=item.get("partBase"),
                    partName=item.get("partName"),
                    partId=item.get("partId"),
                    sourceSheet=item.get("sourceSheet"),
                    treePath=item.get("treePath") or [],
                    parentName=item.get("parentName"),
                )
                for item in (graph_or_embedding_part_matches or [])
            ]
            if _should_use_ai_chat_rerank(
                message_tokens=message_tokens,
                requested_part_families=requested_part_families,
                selected_parts=list(req.selectedParts or []),
                part_candidates=ai_candidate_parts,
                process_candidates=template_process_candidates,
                limit=limit,
            ):
                recommended_parts, recommended_processes, ai_reply = _build_chat_part_matches_with_ai(
                    message,
                    selected_parts=list(req.selectedParts or []),
                    candidate_parts=ai_candidate_parts,
                    process_candidates=template_process_candidates,
                    limit=limit,
                )
            else:
                recommended_parts, recommended_processes, ai_reply = _build_chat_part_matches_without_ai(
                    message,
                    selected_parts=list(req.selectedParts or []),
                    candidate_parts=ai_candidate_parts,
                    process_candidates=template_process_candidates,
                    limit=limit,
                )
            _emit_sequence_recommend_log(
                "CHAT_AI_PROVIDER_RESULT",
                partCount=len(recommended_parts),
                processCount=len(recommended_processes),
                reply=ai_reply,
            )
        except Exception as exc:
            logger.warning("Provider chat recommendation failed; falling back to heuristic ranking: %s", exc)
            _emit_sequence_recommend_log(
                "CHAT_AI_PROVIDER_FAILED",
                error=str(exc),
            )

    if not used_fast_path and recommended_parts:
        process_source_parts = [
            SimpleNamespace(
                nodeName=item.get("nodeName"),
                partBase=item.get("partBase"),
                partName=item.get("partName"),
                partId=item.get("partId"),
                sourceSheet=item.get("sourceSheet"),
            )
            for item in recommended_parts
        ]
        if not recommended_processes:
            recommended_processes = _recommend_next_processes_from_graph(
                selected_parts=process_source_parts,
                effective_process_templates=effective_process_templates,
                limit=candidate_limit,
                message_tokens=message_tokens,
                apply_message_adjustment=True,
            )
    elif not used_fast_path and req.selectedParts:
        selected_matches = _build_selected_part_matches(
            req.selectedParts,
            max(candidate_limit, len(req.selectedParts)),
            effective_process_templates=effective_process_templates,
        )
        if req.expandSelectedParts is False:
            recommended_parts = selected_matches
        else:
            additional_matches = _build_candidate_part_matches(
                req.candidateParts,
                message_tokens,
                candidate_limit,
                effective_process_templates=effective_process_templates,
            )
            additional_matches = _filter_part_matches_by_requested_families(
                additional_matches,
                requested_part_families,
                message_tokens,
            )
            recommended_parts = _merge_part_recommendations(
                selected_matches,
                additional_matches,
                max(candidate_limit, len(selected_matches), 5),
            )
        process_source_parts = [
            SimpleNamespace(
                nodeName=item.get("nodeName"),
                partBase=item.get("partBase"),
                partName=item.get("partName"),
                partId=item.get("partId"),
                sourceSheet=item.get("sourceSheet"),
            )
            for item in recommended_parts
        ]
    elif not used_fast_path and req.candidateParts:
        recommended_parts = _build_candidate_part_matches(
            req.candidateParts,
            message_tokens,
            candidate_limit,
            effective_process_templates=effective_process_templates,
        )
        recommended_parts = _filter_part_matches_by_requested_families(
            recommended_parts,
            requested_part_families,
            message_tokens,
        )
        recommended_parts = _prioritize_direct_message_part_matches(
            recommended_parts,
            message_tokens,
        )
        process_source_parts = [
            SimpleNamespace(
                nodeName=item.get("nodeName"),
                partBase=item.get("partBase"),
                partName=item.get("partName"),
                partId=item.get("partId"),
                sourceSheet=item.get("sourceSheet"),
            )
            for item in recommended_parts
        ]
    elif not used_fast_path:
        recommended_parts = []
        process_source_parts = []

        recommended_processes = _recommend_next_processes_from_graph(
            selected_parts=process_source_parts,
            effective_process_templates=effective_process_templates,
            limit=candidate_limit,
            message_tokens=message_tokens,
            apply_message_adjustment=True,
        )

    if not used_fast_path and not recommended_parts:
        recommended_processes = _recommend_processes_from_message_fallback(
            message_tokens=message_tokens,
            effective_process_templates=effective_process_templates,
            limit=candidate_limit,
        )
        recommended_processes = _filter_process_matches_by_requested_context(
            recommended_processes,
            requested_part_families,
            message_tokens,
        )
        recommended_parts = _build_part_matches_from_process_recommendations(
            recommended_processes,
            limit=max(candidate_limit, 3),
        )
        process_source_parts = [
            SimpleNamespace(
                nodeName=item.get("nodeName"),
                partBase=item.get("partBase"),
                partName=item.get("partName"),
                partId=item.get("partId"),
                sourceSheet=item.get("sourceSheet"),
            )
            for item in recommended_parts
        ]

    if not used_fast_path and not ai_reply:
        try:
            recommended_parts, recommended_processes, ai_reply = _rerank_chat_part_matches_with_openai(
                message,
                selected_parts=list(req.selectedParts or []),
                part_candidates=recommended_parts,
                process_candidates=recommended_processes,
                limit=limit,
            )
        except Exception as exc:
            logger.warning("Provider chat rerank failed; falling back to heuristic ranking: %s", exc)
            recommended_parts = _dedupe_part_matches(recommended_parts)[:limit]
            recommended_processes = _dedupe_process_family_matches(recommended_processes)[:limit]

    recommended_processes = _filter_process_matches_overlapping_parts(
        _filter_manual_barcode_reading_processes(
            _dedupe_process_family_matches(recommended_processes)
        ),
        recommended_parts,
    )[:limit]
    recommended_processes = _boost_fastener_process_matches(
        recommended_processes,
        effective_process_templates=effective_process_templates,
        requested_part_families=requested_part_families,
        message_tokens=message_tokens,
        limit=limit,
    )[:limit]
    recommended_processes = _boost_requested_action_process_matches(
        recommended_processes,
        effective_process_templates=effective_process_templates,
        message_tokens=message_tokens,
        limit=limit,
    )[:limit]
    recommended_processes = _boost_connector_connection_process_matches(
        recommended_processes,
        effective_process_templates=effective_process_templates,
        message_tokens=message_tokens,
        limit=limit,
    )[:limit]
    recommended_processes = _prioritize_requested_action_part_base_matches(
        recommended_processes,
        message_tokens=message_tokens,
        limit=limit,
    )[:limit]
    recommended_processes = _filter_out_fastener_for_connector_connection(
        recommended_processes,
        message_tokens,
    )[:limit]

    recommended_parts = _dedupe_part_matches(recommended_parts)
    recommended_parts = _apply_chat_part_canonical_resolution(
        recommended_parts,
        message_tokens,
        effective_process_templates,
    )
    recommended_parts = _prioritize_direct_message_part_matches(
        recommended_parts,
        message_tokens,
    )
    recommended_parts = _apply_chat_part_display_labels(recommended_parts, message_tokens)

    if req.candidateParts and not req.selectedParts:
        direct_message_part_matches = _build_direct_message_part_matches(
            list(req.candidateParts or []),
            message_tokens,
            effective_process_templates=effective_process_templates,
            limit=limit,
        )
        if direct_message_part_matches:
            recommended_parts = _apply_chat_part_display_labels(
                direct_message_part_matches,
                message_tokens,
            )

    recommended_options: List[Dict[str, Any]] = []
    for part in recommended_parts:
        part_base = str(part.get("partBase") or "").strip()
        if not part_base:
            continue
        options = _load_options_from_excel(part_base, part.get("sourceSheet"))
        if options:
            recommended_options.append(
                {
                    "targetType": "PART",
                    "targetKey": part_base,
                    "sourceSheet": part.get("sourceSheet"),
                    "options": options,
                }
            )

    for process in recommended_processes:
        part_base = str(process.get("partBase") or "").strip()
        if not part_base:
            continue
        options = _load_process_options_from_excel(
            part_base,
            process.get("label"),
            process.get("sourceSheet"),
        )
        if options:
            recommended_options.append(
                {
                    "targetType": "PROCESS",
                    "targetKey": process.get("label") or process.get("processKey") or part_base,
                    "sourceSheet": process.get("sourceSheet"),
                    "options": options,
                }
            )

    deduped_options: List[Dict[str, Any]] = []
    seen_option_keys = set()
    for item in recommended_options:
        option_key = (item.get("targetType"), item.get("targetKey"))
        if option_key in seen_option_keys:
            continue
        seen_option_keys.add(option_key)
        deduped_options.append(item)

    if req.includePerPartRecommendations and recommended_parts:
        per_part_recommendations = _build_per_part_chat_recommendations(
            part_matches=recommended_parts,
            effective_process_templates=effective_process_templates,
            shared_process_matches=recommended_processes,
            message_text=message,
            message_tokens=message_tokens,
            requested_part_families=requested_part_families,
            candidate_limit=candidate_limit,
            per_part_limit=limit,
        )
        per_part_recommendations = [
            {
                **item,
                "part": _apply_chat_part_display_labels([item.get("part") or {}], message_tokens)[0],
            }
            for item in per_part_recommendations
        ]

    reply = ai_reply or _build_chat_reply(
        message=message,
        part_matches=recommended_parts,
        process_matches=recommended_processes,
        option_matches=deduped_options,
    )

    return SequenceChatResponse(
        reply=reply,
        recommendedParts=recommended_parts,
        recommendedProcesses=recommended_processes,
        recommendedOptions=deduped_options,
        perPartRecommendations=per_part_recommendations,
    )


@router.post("/chat/per-part", response_model=SequenceChatPerPartResponse)
def chat_sequence_per_part_recommendations(req: SequenceChatPerPartRequest):
    message = str(req.message or "").strip()
    if not message:
        raise HTTPException(status_code=400, detail="message가 비어 있습니다.")

    limit = max(1, min(int(req.limit or 5), 5))
    candidate_limit = max(limit * 3, 10)
    message_tokens = _tokenize_chat_text(message)
    requested_part_families = _requested_part_families_from_message(message_tokens)
    effective_process_templates = _resolve_effective_process_templates(req.processTemplates)
    part_matches = _build_chat_part_matches_from_inputs(list(req.selectedParts or []))
    part_matches = _apply_chat_part_canonical_resolution(
        part_matches,
        message_tokens,
        effective_process_templates,
    )
    part_matches = _prioritize_direct_message_part_matches(
        part_matches,
        message_tokens,
    )
    part_matches = _apply_chat_part_display_labels(part_matches, message_tokens)

    if not part_matches:
        return SequenceChatPerPartResponse(perPartRecommendations=[])

    shared_process_matches = _recommend_processes_from_message_fallback(
        message_tokens=message_tokens,
        effective_process_templates=effective_process_templates,
        limit=max(candidate_limit, 20),
    )
    shared_process_matches = _filter_process_matches_by_requested_context(
        shared_process_matches,
        requested_part_families,
        message_tokens,
    )
    shared_process_matches = _boost_fastener_process_matches(
        _dedupe_process_family_matches(shared_process_matches),
        effective_process_templates=effective_process_templates,
        requested_part_families=requested_part_families,
        message_tokens=message_tokens,
        limit=max(candidate_limit, 20),
    )
    shared_process_matches = _boost_requested_action_process_matches(
        shared_process_matches,
        effective_process_templates=effective_process_templates,
        message_tokens=message_tokens,
        limit=max(candidate_limit, 20),
    )
    shared_process_matches = _boost_connector_connection_process_matches(
        shared_process_matches,
        effective_process_templates=effective_process_templates,
        message_tokens=message_tokens,
        limit=max(candidate_limit, 20),
    )
    shared_process_matches = _prioritize_requested_action_part_base_matches(
        shared_process_matches,
        message_tokens=message_tokens,
        limit=max(candidate_limit, 20),
    )
    shared_process_matches = _filter_out_fastener_for_connector_connection(
        shared_process_matches,
        message_tokens,
    )[: max(candidate_limit, 20)]

    per_part_recommendations = _build_per_part_chat_recommendations(
        part_matches=part_matches,
        effective_process_templates=effective_process_templates,
        shared_process_matches=shared_process_matches,
        message_text=message,
        message_tokens=message_tokens,
        requested_part_families=requested_part_families,
        candidate_limit=candidate_limit,
        per_part_limit=limit,
    )
    per_part_recommendations = [
        {
            **item,
            "part": _apply_chat_part_display_labels([item.get("part") or {}], message_tokens)[0],
        }
        for item in per_part_recommendations
    ]
    return SequenceChatPerPartResponse(perPartRecommendations=per_part_recommendations)


@router.post("/recommend-next-processes", response_model=SequenceNextProcessRecommendationResponse)
def recommend_next_processes(req: SequenceNextProcessRecommendationRequest):
    _emit_sequence_recommend_log(
        "NEXT_PROCESS_REQUEST",
        bomId=req.bomId,
        spec=req.spec,
        selectedParts=[
            str(part.partBase or part.partId or part.nodeName or "").strip()
            for part in (req.selectedParts or [])
        ],
        limit=req.limit,
    )
    limit = max(1, min(int(req.limit or 5), 10))
    effective_process_templates = _resolve_effective_process_templates(req.processTemplates)
    selected_parts = list(req.selectedParts or [])
    recommended_sequence: List[Dict[str, Any]] = []
    if len(selected_parts) >= 2:
        recommended_sequence = _recommend_selected_part_combination_sequence(
            selected_parts=selected_parts,
            effective_process_templates=effective_process_templates,
            limit=limit,
        )
        if any(step.get("type") == "PROCESS" for step in recommended_sequence):
            recommended_processes = []
        else:
            recommended_processes = _recommend_common_process_path_from_graph(
                selected_parts=selected_parts,
                effective_process_templates=effective_process_templates,
                limit=limit,
            )
    else:
        recommended_processes = _recommend_next_processes_from_graph(
            selected_parts=selected_parts,
            effective_process_templates=effective_process_templates,
            limit=limit,
            message_tokens=[],
            apply_message_adjustment=False,
        )
    recommended_sequence = _filter_manual_barcode_reading_sequence_steps(recommended_sequence)
    recommended_processes = _filter_manual_barcode_reading_processes(recommended_processes)
    if (
        len(selected_parts) >= 2
        and recommended_sequence
        and not any(step.get("type") == "PROCESS" for step in recommended_sequence)
        and not recommended_processes
    ):
        recommended_sequence = []
        recommended_processes = _filter_manual_barcode_reading_processes(
            _recommend_common_process_path_from_graph(
                selected_parts=selected_parts,
                effective_process_templates=effective_process_templates,
                limit=limit,
            )
        )
    _emit_sequence_recommend_log(
        "NEXT_PROCESS_RESPONSE",
        recommendedCount=len(recommended_processes),
        recommendedLabels=[
            str(item.get("label") or item.get("processKey") or "").strip()
            for item in recommended_processes
        ],
    )
    return SequenceNextProcessRecommendationResponse(
        recommendedProcesses=recommended_processes,
        recommendedSequence=recommended_sequence,
    )


@router.post("/debug-print")
def sequence_debug_print(req: SequenceDebugPrintRequest):
    try:
        _append_sequence_debug_log(req.stage, req.payload)
    except Exception as exc:
        logger.exception("Failed to append sequence debug log")
        raise HTTPException(500, f"시퀀스 디버그 로그 저장 실패: {exc}")
    return {"ok": True}

@router.get("/load")
def load_sequence(bomId: str, spec: str):
    path = DATA_DIR / "data"/ "bom_runs" / bomId / f"{spec}_sequence.json"

    if not path.exists():
        raise HTTPException(404, "시퀀스 파일 없음")

    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


@router.post("/ai-draft")
def create_sequence_ai_draft(req: SequenceAIDraftRequest):
    return generate_sequence_ai_draft(req)

@router.post("/save")
def save_sequence(req: SequenceSaveRequest):
    bom_id = req.bomId
    spec = req.spec

    save_dir = DATA_DIR / "data"/ "bom_runs" / bom_id
    save_dir.mkdir(parents=True, exist_ok=True)

    save_path = save_dir / f"{spec}_sequence.json"

    payload = jsonable_encoder(
        {
            "bomId": bom_id,
            "spec": spec,
            "nodes": req.nodes,
            "edges": req.edges,
            "groups": req.groups or [],
            "workerGroups": req.workerGroups or [],
            "sequenceVersion": int(time.time() * 1000),
        }
    )

    _write_json_atomic(save_path, payload)

    return {"status": "ok"}



