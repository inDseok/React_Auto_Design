import os
import re
import cv2
import math
import time
import numpy as np

from tkinter import Tk, Label, Button, Frame, messagebox, StringVar, Entry
from tkinter import ttk, filedialog

from PIL import Image, ImageTk, ImageGrab
from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

os.environ.setdefault("PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK", "True")

SCALE = 2
CONF_THRESHOLD = 0.9
PREVIEW_MAX_W = 420
PREVIEW_MAX_H = 320


def strict_number(text: str):
    text = text.strip()
    text = text.replace("O", "0").replace("o", "0")
    text = text.replace("I", "1").replace("l", "1")
    text = text.replace(",", ".")
    text = text.replace(" ", "")
    text = re.sub(r"[^0-9+\-\.]", "", text)

    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?", text):
        return text
    return None


class OCRCompareApp:
    def __init__(self, root: Tk):
        self.root = root
        self.root.title("LH/RH OCR Compare")
        self.root.geometry("1050x760")

        self.ocr = PaddleOCR(
            use_doc_orientation_classify=False,
            use_doc_unwarping=False,
            use_textline_orientation=False,
            text_detection_model_name="PP-OCRv5_mobile_det",
            text_recognition_model_name="PP-OCRv5_mobile_rec",
        )

        self.lh_image = None
        self.rh_image = None
        self.lh_preview_ref = None
        self.rh_preview_ref = None
        self.last_lh_numbers = []
        self.last_rh_numbers = []

        self.conf_var = StringVar(value=str(CONF_THRESHOLD))
        self.scale_var = StringVar(value=str(SCALE))
        self.status_var = StringVar(value="LH / RH 영역을 캡처하세요.")

        self._build_ui()

    def _build_ui(self):
        top = Frame(self.root)
        top.pack(fill="x", padx=12, pady=10)

        Button(top, text="LH 스크린샷", command=self.capture_lh).grid(row=0, column=0, padx=5)
        Button(top, text="RH 스크린샷", command=self.capture_rh).grid(row=0, column=1, padx=5)
        Button(top, text="OCR 실행", command=self.run_compare).grid(row=0, column=2, padx=5)
        Button(top, text="엑셀 저장", command=self.save_excel).grid(row=0, column=3, padx=5)

        Label(top, text="Scale").grid(row=0, column=4, padx=(20, 5))
        Entry(top, textvariable=self.scale_var, width=6).grid(row=0, column=5)

        Label(top, text="Conf").grid(row=0, column=6, padx=(20, 5))
        Entry(top, textvariable=self.conf_var, width=6).grid(row=0, column=7)

        preview_frame = Frame(self.root)
        preview_frame.pack(fill="x", padx=12, pady=8)

        self.lh_title = Label(preview_frame, text="LH: 캡처 대기")
        self.lh_title.grid(row=0, column=0, sticky="w")

        self.rh_title = Label(preview_frame, text="RH: 캡처 대기")
        self.rh_title.grid(row=0, column=1, sticky="w")

        self.lh_label = Label(preview_frame, bd=1, relief="solid")
        self.lh_label.grid(row=1, column=0, padx=5, pady=5)

        self.rh_label = Label(preview_frame, bd=1, relief="solid")
        self.rh_label.grid(row=1, column=1, padx=5, pady=5)

        result_frame = Frame(self.root)
        result_frame.pack(fill="both", expand=True, padx=12, pady=10)

        columns = ("index", "rh", "lh", "rh_lh", "lh_rh")
        self.tree = ttk.Treeview(result_frame, columns=columns, show="headings", height=20)

        self.tree.heading("index", text="Index")
        self.tree.heading("rh", text="RH")
        self.tree.heading("lh", text="LH")
        self.tree.heading("rh_lh", text="RH-LH")
        self.tree.heading("lh_rh", text="LH-RH")

        self.tree.column("index", width=70, anchor="center")
        self.tree.column("rh", width=140, anchor="center")
        self.tree.column("lh", width=140, anchor="center")
        self.tree.column("rh_lh", width=140, anchor="center")
        self.tree.column("lh_rh", width=140, anchor="center")

        scrollbar = ttk.Scrollbar(result_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        status = Label(self.root, textvariable=self.status_var, anchor="w", bd=1, relief="sunken")
        status.pack(fill="x", padx=12, pady=(0, 10))

    def capture_region(self):
        self.root.withdraw()
        time.sleep(0.3)

        screenshot_pil = ImageGrab.grab()
        screenshot = cv2.cvtColor(np.array(screenshot_pil), cv2.COLOR_RGB2BGR)
        clone = screenshot.copy()

        state = {
            "start": None,
            "end": None,
            "current": None,
            "dragging": False
        }

        window_name = "Select Area"

        def mouse_callback(event, x, y, flags, param):
            if event == cv2.EVENT_LBUTTONDOWN:
                state["start"] = (x, y)
                state["current"] = (x, y)
                state["dragging"] = True

            elif event == cv2.EVENT_MOUSEMOVE and state["dragging"]:
                state["current"] = (x, y)

            elif event == cv2.EVENT_LBUTTONUP:
                state["end"] = (x, y)
                state["current"] = (x, y)
                state["dragging"] = False

        cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
        cv2.setWindowProperty(window_name, cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)
        cv2.setMouseCallback(window_name, mouse_callback)

        selected_crop = None

        while True:
            display = screenshot.copy()

            if state["start"] is not None and state["current"] is not None:
                x1, y1 = state["start"]
                x2, y2 = state["current"]

                left = min(x1, x2)
                right = max(x1, x2)
                top = min(y1, y2)
                bottom = max(y1, y2)

                cv2.rectangle(display, (left, top), (right, bottom), (0, 255, 0), 2)

            cv2.imshow(window_name, display)
            key = cv2.waitKey(1) & 0xFF

            if key == 27:
                break

            if state["start"] is not None and state["end"] is not None:
                x1, y1 = state["start"]
                x2, y2 = state["end"]

                left = min(x1, x2)
                right = max(x1, x2)
                top = min(y1, y2)
                bottom = max(y1, y2)

                if right - left > 5 and bottom - top > 5:
                    selected_crop = clone[top:bottom, left:right]
                break

        cv2.destroyWindow(window_name)
        self.root.deiconify()
        self.root.lift()
        self.root.focus_force()

        if selected_crop is None:
            return None

        return Image.fromarray(cv2.cvtColor(selected_crop, cv2.COLOR_BGR2RGB))

    def capture_lh(self):
        img = self.capture_region()
        if img is None:
            self.status_var.set("LH 캡처가 취소되었습니다.")
            return

        self.lh_image = img
        self.lh_title.config(text="LH: 캡처 완료")
        self.show_preview(img, self.lh_label, is_lh=True)
        self.status_var.set("LH 영역 캡처 완료")

    def capture_rh(self):
        img = self.capture_region()
        if img is None:
            self.status_var.set("RH 캡처가 취소되었습니다.")
            return

        self.rh_image = img
        self.rh_title.config(text="RH: 캡처 완료")
        self.show_preview(img, self.rh_label, is_lh=False)
        self.status_var.set("RH 영역 캡처 완료")

    def show_preview(self, img, widget, is_lh=True):
        preview = img.copy()
        preview.thumbnail((PREVIEW_MAX_W, PREVIEW_MAX_H))

        tk_img = ImageTk.PhotoImage(preview)
        widget.config(image=tk_img)

        if is_lh:
            self.lh_preview_ref = tk_img
        else:
            self.rh_preview_ref = tk_img

    def rotate_list(self, lst, start_idx):
        return lst[start_idx:] + lst[:start_idx]

    def run_ocr_with_sort(self, pil_img, sort_direction="lh"):
        scale = float(self.scale_var.get())
        conf_threshold = float(self.conf_var.get())

        img = cv2.cvtColor(np.array(pil_img.convert("RGB")), cv2.COLOR_RGB2BGR)

        img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        img_gray = cv2.resize(img_gray, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
        img_ocr = cv2.cvtColor(img_gray, cv2.COLOR_GRAY2BGR)

        result = self.ocr.predict(
            img_ocr,
            text_det_box_thresh=0.8,
            text_det_thresh=0.5,
            text_det_unclip_ratio=1.2,
        )

        detected_items = []

        for res in result:
            boxes = res["rec_boxes"]
            texts = res["rec_texts"]
            scores = res["rec_scores"]

            for box, txt, score in zip(boxes, texts, scores):
                if score < conf_threshold:
                    continue

                value = strict_number(txt)
                if value is None:
                    continue

                x1, y1, x2, y2 = box
                x1 = int(x1 / scale)
                y1 = int(y1 / scale)
                x2 = int(x2 / scale)
                y2 = int(y2 / scale)

                cx = (x1 + x2) / 2
                cy = (y1 + y2) / 2

                detected_items.append({
                    "value": value,
                    "score": float(score),
                    "box": (x1, y1, x2, y2),
                    "cx": cx,
                    "cy": cy,
                })

        if not detected_items:
            return []

        center_x = sum(item["cx"] for item in detected_items) / len(detected_items)
        center_y = sum(item["cy"] for item in detected_items) / len(detected_items)

        for item in detected_items:
            dx = item["cx"] - center_x
            dy = item["cy"] - center_y
            angle = math.atan2(-dy, dx)
            angle_deg = (math.degrees(angle) + 360) % 360

            item["angle_deg"] = angle_deg
            item["rh_sort_key"] = (360 - angle_deg) % 360
            item["lh_sort_key"] = (angle_deg - 180) % 360

        if sort_direction == "rh":
            detected_items = sorted(detected_items, key=lambda x: (x["rh_sort_key"], -x["cx"]))
            start_item = max(detected_items, key=lambda x: x["cx"])
            start_idx = next(i for i, item in enumerate(detected_items) if item is start_item)
            detected_items = self.rotate_list(detected_items, start_idx)

        elif sort_direction == "lh":
            detected_items = sorted(detected_items, key=lambda x: (x["lh_sort_key"], x["cx"]))
            start_item = min(detected_items, key=lambda x: x["cx"])
            start_idx = next(i for i, item in enumerate(detected_items) if item is start_item)
            detected_items = self.rotate_list(detected_items, start_idx)

        else:
            raise ValueError("sort_direction must be 'lh' or 'rh'")

        return [item["value"] for item in detected_items]

    def run_compare(self):
        if self.lh_image is None or self.rh_image is None:
            messagebox.showwarning("경고", "LH와 RH 영역을 모두 캡처하세요.")
            return

        try:
            self.last_lh_numbers = self.run_ocr_with_sort(self.lh_image, sort_direction="lh")
            self.last_rh_numbers = self.run_ocr_with_sort(self.rh_image, sort_direction="rh")
            self._fill_tree()
            self.status_var.set(
                f"OCR 완료 - LH {len(self.last_lh_numbers)}개, RH {len(self.last_rh_numbers)}개 검출"
            )
        except Exception as e:
            messagebox.showerror("오류", str(e))
            self.status_var.set("OCR 실행 중 오류가 발생했습니다.")

    def _fill_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        max_len = max(len(self.last_lh_numbers), len(self.last_rh_numbers))

        for i in range(max_len):
            rh_val = self.last_rh_numbers[i] if i < len(self.last_rh_numbers) else ""
            lh_val = self.last_lh_numbers[i] if i < len(self.last_lh_numbers) else ""

            rh_num = self._to_float_safe(rh_val) if rh_val != "" else None
            lh_num = self._to_float_safe(lh_val) if lh_val != "" else None

            if rh_num is not None and lh_num is not None:
                rh_lh = round(rh_num - lh_num, 6)
                lh_rh = round(lh_num - rh_num, 6)
            else:
                rh_lh = ""
                lh_rh = ""

            self.tree.insert("", "end", values=(i + 1, rh_val, lh_val, rh_lh, lh_rh))

    def _to_float_safe(self, v):
        try:
            return float(v)
        except Exception:
            return None

    def save_excel(self):
        if not self.last_lh_numbers and not self.last_rh_numbers:
            messagebox.showwarning("경고", "먼저 OCR 실행을 해주세요.")
            return

        path = filedialog.asksaveasfilename(
            title="엑셀 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile="ocr_compare.xlsx",
        )
        if not path:
            return

        try:
            self._write_excel(path)
            self.status_var.set(f"엑셀 저장 완료: {path}")
            messagebox.showinfo("완료", f"엑셀 저장 완료\n{path}")
        except Exception as e:
            messagebox.showerror("오류", str(e))

    def _write_excel(self, output_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "OCR Compare"

        headers = ["Index", "RH", "LH", "RH-LH", "LH-RH"]
        ws.append(headers)

        max_len = max(len(self.last_rh_numbers), len(self.last_lh_numbers))

        for i in range(max_len):
            rh_val = self.last_rh_numbers[i] if i < len(self.last_rh_numbers) else ""
            lh_val = self.last_lh_numbers[i] if i < len(self.last_lh_numbers) else ""

            rh_num = self._to_float_safe(rh_val) if rh_val != "" else None
            lh_num = self._to_float_safe(lh_val) if lh_val != "" else None

            if rh_num is not None and lh_num is not None:
                rh_minus_lh = rh_num - lh_num
                lh_minus_rh = lh_num - rh_num
            else:
                rh_minus_lh = ""
                lh_minus_rh = ""

            ws.append([i + 1, rh_val, lh_val, rh_minus_lh, lh_minus_rh])

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=2, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18

        for row in range(2, ws.max_row + 1):
            ws[f"D{row}"].number_format = "0.000"
            ws[f"E{row}"].number_format = "0.000"

        wb.save(output_path)


if __name__ == "__main__":
    root = Tk()
    app = OCRCompareApp(root)
    root.mainloop()